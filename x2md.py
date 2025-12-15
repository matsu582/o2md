#!/usr/bin/env python3
"""
Excel to Markdown Converter
Excelファイルをシートごとに詳細なMarkdown形式に変換するツール

特徴:
- セル内の改行を<br>タグで表現
- 罫線を意識した表の作成
- 図形や埋め込み画像を抽出してMarkdownに挿入

このファイルはxlsxファイルを解析してシートごとにMarkdownを生成します。
デバッグ情報や一時ファイルの出力が含まれます。
"""

import os
import sys
import tempfile
import subprocess
import shutil
from pathlib import Path
from typing import List, Dict, Tuple, Optional, Any, Set
import io
import zipfile
import xml.etree.ElementTree as ET

from utils import get_libreoffice_path, col_letter, normalize_excel_path, get_xml_from_zip, extract_anchor_id, anchor_is_hidden, anchor_has_drawable as utils_anchor_has_drawable
from isolated_group_renderer import IsolatedGroupRenderer
from x2md_tables import _TablesMixin
from x2md_graphics import _GraphicsMixin
from x2md_charts import extract_charts_from_worksheet
from chart_utils import chart_data_to_markdown

try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
except ImportError as e:
    raise ImportError(
        "openpyxlライブラリが必要です: pip install openpyxl または uv sync を実行してください"
    ) from e

try:
    from PIL import Image
except ImportError as e:
    raise ImportError(
        "Pillowライブラリが必要です: pip install pillow または uv sync を実行してください"
    ) from e

try:
    import fitz
except ImportError as e:
    raise ImportError(
        "PyMuPDFライブラリが必要です: pip install PyMuPDF または uv sync を実行してください"
    ) from e

# 設定定数
LIBREOFFICE_PATH = get_libreoffice_path()

# DPI設定
DEFAULT_DPI = 600
IMAGE_QUALITY = 100
IMAGE_BORDER_SIZE = 8

# スキャン設定
MAX_HEAD_SCAN_ROWS = 12
MAX_SCAN_COLUMNS = 60




# グローバルverboseフラグ
_VERBOSE = False

def set_verbose(verbose: bool):
    """verboseモードを設定"""
    global _VERBOSE
    _VERBOSE = verbose

def is_verbose() -> bool:
    """verboseモードかどうかを返す"""
    return _VERBOSE

def debug_print(*args, **kwargs):
    """verboseモード時のみ出力するデバッグ用print"""
    if _VERBOSE:
        print(*args, **kwargs)

class ExcelToMarkdownConverter(_TablesMixin, _GraphicsMixin):
    """ExcelファイルをMarkdown形式に変換するコンバータクラス
    
    機能は以下のMixinクラスから継承されます:
    - _TablesMixin: テーブル検出・構築・出力機能
    - _GraphicsMixin: 画像処理・図形処理・レンダリング機能
    """
    class _LoggingList(list):
        """デバッグ用にappend/insert操作をログ出力するlistのラッパー

        標準出力にログを出力し、可能であればコンバータのデバッグログにも書き込みます。
        """
        def __init__(self, owner, *args):
            super().__init__(*args)
            self._owner = owner

        def append(self, item):
            debug_print(f"[MD_APPEND] {repr(item)}")
            try:
                if isinstance(item, str) and item.strip() == '---' and len(self) and isinstance(self[-1], str) and self[-1].strip() == '---':
                    return
            except (ValueError, TypeError):
                pass

            return super().append(item)

    def __init__(self, excel_file_path: str, output_dir=None, shape_metadata=False, output_format='png'):
        """コンバータインスタンスの初期化

        CLIから使用できるように、最小限で安全なコンストラクタを提供します。
        意図的に保守的な初期化を維持し、メソッド間で使用される共通のシート毎の
        一時的な状態を準備します。
        
        Args:
            excel_file_path: 変換するExcelファイルのパス
            output_dir: 出力ディレクトリ（省略時はデフォルト）
            shape_metadata: 図形メタデータ出力フラグ
            output_format: 出力画像形式 ('png' または 'svg')
        """
        self.excel_file = excel_file_path
        self.base_name = Path(excel_file_path).stem
        if output_dir:
            self.output_dir = output_dir
        else:
            self.output_dir = os.path.join(os.getcwd(), "output")
        self.images_dir = os.path.join(self.output_dir, "images")
        
        self.debug_mode = is_verbose()
        self.shape_metadata = shape_metadata
        self.output_format = output_format.lower() if output_format else 'png'
        
        # 出力形式の検証
        if self.output_format not in ('png', 'svg'):
            print(f"[WARNING] 不明な出力形式 '{output_format}'。'png'を使用します。")
            self.output_format = 'png'

        os.makedirs(self.output_dir, exist_ok=True)
        os.makedirs(self.images_dir, exist_ok=True)

        self.markdown_lines = self._LoggingList(self)
        self.image_counter = 0

        self._init_per_sheet_state()

        self.workbook = load_workbook(excel_file_path, data_only=True)
        print(f"[INFO] Excelワークブック読み込み完了: {excel_file_path}")
        print(f"[INFO] 出力画像形式: {self.output_format.upper()}")

    def _init_per_sheet_state(self):
        """シート毎の状態変数を初期化"""
        self._cell_to_md_index = {}
        self._sheet_shape_images = {}
        self._sheet_shape_next_idx = {}
        self._sheet_shapes_generated = set()
        self._sheet_shape_image_start_rows = {}
        self._sheet_deferred_texts = {}
        self._sheet_deferred_tables = {}
        self._sheet_emitted_texts = {}
        self._sheet_emitted_rows = {}
        self._sheet_emitted_table_titles = {}
        self._emitted_images = set()
        self._embedded_image_cid_by_name = {}
        self._in_canonical_emit = False
        self._global_iso_preserved_ids = set()
        self._image_shape_ids = {}
        self._last_iso_preserved_ids = set()

    def _clear_sheet_state(self, sheet_name: str):
        """特定のシートの状態をクリアする"""
        for dict_attr in ['_cell_to_md_index', '_sheet_shape_images', '_sheet_shape_next_idx',
                          '_sheet_shape_image_start_rows', '_sheet_deferred_texts',
                          '_sheet_deferred_tables', '_sheet_emitted_texts', '_sheet_emitted_rows',
                          '_sheet_emitted_table_titles', '_embedded_image_cid_by_name']:
            getattr(self, dict_attr, {}).pop(sheet_name, None)
        
        self._sheet_shapes_generated.discard(sheet_name)
        self._global_iso_preserved_ids.clear()
        self._last_iso_preserved_ids.clear()

    def _is_canonical_emit(self) -> bool:
        """現在正規出力モードかどうかを確認する"""
        return getattr(self, '_in_canonical_emit', False)

    def _safe_get_cell_value(self, sheet, row: int, col: int) -> Any:
        """セル値を安全に取得し、エラー時はNoneを返す"""
        try:
            return sheet.cell(row, col).value
        except Exception:
            return None

    def convert(self) -> str:
        """トップレベルの変換処理 (軽量ラッパ)

        既存のコードベースには複数の補助メソッドがあるため、ここでは
        最小限のフローを提供して CLI から呼べるようにします。
        - ドキュメント見出しの追加
        - 目次生成 (存在すれば呼ぶ)
        - 各シートを順に変換
        - Markdown ファイルを書き出してパスを返す
        """
        print(f"[INFO] Excel文書変換開始: {self.excel_file}")

        # ドキュメントタイトルを先頭に追加
        self.markdown_lines.append(f"# {self.base_name}")
        self.markdown_lines.append("")

        # ヘルパーが存在する場合は目次を生成
        if hasattr(self, '_generate_toc') and callable(getattr(self, '_generate_toc')):
            try:
                self._generate_toc()
            except Exception as e:
                print(f"[WARNING] 目次生成失敗: {e}")

        # シートを変換
        for sheet_name in self.workbook.sheetnames:
            try:
                sheet = self.workbook[sheet_name]
                
                # 非表示シートをスキップ
                if sheet.sheet_state == 'hidden':
                    print(f"[INFO] シートをスキップ（非表示）: {sheet_name}")
                    continue
                
                print(f"[INFO] シート変換中: {sheet_name}")
                self._convert_sheet(sheet)
            except Exception as e:
                print(f"[WARNING] シート処理中にエラーが発生しました: {sheet_name} -> {e}")
                import traceback
                traceback.print_exc()
                continue

        # Markdown出力を書き込み
        output_file = os.path.join(self.output_dir, f"{self.base_name}.md")
        content = "\n".join(str(x) for x in self.markdown_lines)
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(content)
        print(f"[SUCCESS] 変換完了: {output_file}")
        return output_file

    def _mark_image_emitted(self, img_name: str):
        """Mark an image as emitted only during the canonical emission pass."""
        if self._is_canonical_emit():
            self._emitted_images.add(str(img_name))
        else:
            debug_print(f"[TRACE] Skipping _emitted_images.add({img_name}) in non-canonical pass")

    def _mark_sheet_map(self, sheet_title: str, src_row: int, md_index: int):
        """Record a source-row -> markdown index mapping only during canonical emission."""
        if self._is_canonical_emit():
            self._cell_to_md_index.setdefault(sheet_title, {})[src_row] = int(md_index)
        else:
            debug_print(f"[TRACE] Skipping authoritative sheet_map[{sheet_title}][{src_row}] assignment in non-canonical pass")

    def _mark_emitted_row(self, sheet_title: str, row: int):
        """Mark a row as emitted only during canonical emission."""
        if self._is_canonical_emit():
            self._sheet_emitted_rows.setdefault(sheet_title, set()).add(int(row))
        else:
            debug_print(f"[TRACE] Skipping emitted_rows.add({sheet_title},{row}) in non-canonical pass")

    def _mark_emitted_text(self, sheet_title: str, norm_text: str):
        """Record a normalized emitted text only during canonical emission."""
        if self._is_canonical_emit():
            self._sheet_emitted_texts.setdefault(sheet_title, set()).add(str(norm_text))
        else:
            debug_print(f"[TRACE] Skipping emitted_texts.add({sheet_title},...) in non-canonical pass")
        

    def _escape_angle_brackets(self, text: str) -> str:
        """表示用に角括弧をエスケープして、MarkdownでHTMLタグと解釈されないようにする。

        例: '<Tag>' -> '&lt;Tag&gt;'
        """
        try:
            if text is None:
                return ''
            t = str(text)
            # 安全な表示のためリテラル山括弧をHTMLエンティティに置換
            t = t.replace('<', '&lt;').replace('>', '&gt;')
            return t
        except (ValueError, TypeError):
            return str(text)

    def _normalize_text(self, text: str) -> str:
        """Normalize text for duplicate-detection: collapse whitespace and strip."""
        try:
            if text is None:
                return ''
            import re
            s = str(text)
            s = s.strip()
            s = re.sub(r'\s+', ' ', s)
            return s
        except (ValueError, TypeError):
            return str(text).strip()

    def _add_separator(self):
        """Insert a blank, a Markdown thematic break '---', and a blank.

        Returns True if inserted, False if skipped due to dedupe or non-canonical mode.
        """
        if not self._is_canonical_emit():
            return False

        # 重複セパレータの出力を避けるため最後の数行をチェック
        tail = [x for x in self.markdown_lines[-6:] if isinstance(x, str)]
        for t in reversed(tail):
            if t.strip() == '':
                continue
            if t.strip() == '---':
                debug_print("[DEBUG][_add_separator] skipping duplicate separator '---'")
                return False
            break

        self.markdown_lines.append("")
        self.markdown_lines.append('---')
        self.markdown_lines.append("")
        return True

    def _emit_free_text(self, sheet, src_row: Optional[int], text: str):
        """Append a free-form text line if its normalized form hasn't been emitted for this sheet.

        - sheet: worksheet object
        - src_row: source row number (or None)
        - text: raw text to emit

        Returns True if emitted, False if skipped as duplicate.
        """
        try:
            if text is None:
                return False
            norm = self._normalize_text(text)
            # ここで正式なemitted_textsエントリを作成しない; getを使用して
            # 正規エミッタの外で正式なストアを変更することを避けます。
            emitted_texts = self._sheet_emitted_texts.get(sheet.title, set())
            if norm in emitted_texts:
                return False

            if self._is_canonical_emit():
                # 正規の出力: markdownバッファに追加
                self.markdown_lines.append(self._escape_angle_brackets(text) + "  ")
                
                # ソース行をmarkdownインデックスにマップ
                if src_row is not None:
                    md_index = len(self.markdown_lines) - 1
                    self._mark_sheet_map(sheet.title, src_row, md_index)
                    debug_print(f"[DEBUG][_text_emit] sheet={sheet.title} src_row={src_row} md_index={md_index} text_norm='{norm}'")
                
                # 出力済みとしてマーク
                if src_row is not None:
                    self._mark_emitted_row(sheet.title, src_row)
                self._mark_emitted_text(sheet.title, norm)
                return True
            else:
                # 後の正規パスのため出力を延期
                lst = self._sheet_deferred_texts.setdefault(sheet.title, [])
                
                # 重複する延期テキストをチェック
                already_deferred = any(
                    dt is not None and self._normalize_text(dt) == norm
                    for _, dt in lst
                )
                
                if not already_deferred:
                    lst.append((src_row, text))
                return True
        except Exception as e:
            print(f"[ERROR] _emit_free_text failed: {e}")
            return False
    
    def _insert_markdown_image(self, insert_at: Optional[int], md_line: str, img_name: str, sheet=None):
        """Insert or append an image markdown line and immediately mark it as emitted.

        Returns the new insert index (one past the inserted block) when inserted,
        or the current length of markdown_lines when appended.
        """
        try:
            import traceback
            stk = traceback.extract_stack()
            caller = stk[-3] if len(stk) >= 3 else None
            caller_info = f"{caller.filename}:{caller.lineno}:{caller.name}" if caller else 'unknown'
            debug_print(f"[DEBUG][_insert_markdown_image_called] insert_at={insert_at} img_name={img_name} caller={caller_info}")
            # 正規の出力パス中でなく、即座の画像
            # 挿入が明示的に許可されていない場合、このリクエストを
            # 延期登録に変換し、正規エミッタが配置を制御するようにします。
            if not getattr(self, '_in_canonical_emit', False) and not getattr(self, '_allow_immediate_image_inserts', False):
                try:
                    # markdownのaltテキストからシートタイトルを推測: '![<title>](images/...)'
                    import re
                    m = re.search(r'!\[(.*?)\]', md_line or "")
                    sheet_title = None
                    if m:
                        sheet_title = m.group(1)
                        # 末尾の'の図'が存在する場合は削除（一般的なaltテキストパターン）
                        if sheet_title.endswith('の図'):
                            sheet_title = sheet_title[:-2]
                except Exception:
                    sheet_title = None
                key = sheet_title if sheet_title is not None else 'unknown'
                # sheet_shape_imagesは延期された非正式なコレクションで
                # ここで安全に変更できます。
                lst = self._sheet_shape_images.setdefault(key, [])
                # 不明な場合は安全なデフォルトとして代表的な行=1を使用します。
                # 同じシートに対して同じ画像を複数回登録することを避けます。
                already = any((isinstance(it, (list, tuple)) and len(it) >= 2 and it[1] == img_name) or (str(it) == img_name) for it in lst)
                if not already:
                    lst.append((1, img_name))
                    debug_print(f"[DEBUG][_insert_markdown_image_deferred] img_name={img_name} sheet={key}")

                # 変更は実行されません; 挿入を期待する呼び出し側は
                # 追加されたかのように現在のmarkdown長を受け取ります。
                return len(self.markdown_lines)

            if insert_at is None:
                self.markdown_lines.append(md_line)
                self.markdown_lines.append("")
                
                if sheet is not None:
                    try:
                        filter_ids = self._image_shape_ids.get(img_name)
                        shapes_metadata = self._extract_all_shapes_metadata(sheet, filter_ids=filter_ids)
                        if shapes_metadata:
                            text_metadata = self._format_shape_metadata_as_text(shapes_metadata)
                            if text_metadata:
                                self.markdown_lines.append("")
                                for line in text_metadata.split('\n'):
                                    self.markdown_lines.append(line)
                                self.markdown_lines.append("")
                            
                            json_metadata = self._format_shape_metadata_as_json(shapes_metadata)
                            if json_metadata and json_metadata != "{}":
                                self.markdown_lines.append("<details>")
                                self.markdown_lines.append("<summary>JSON形式の図形情報</summary>")
                                self.markdown_lines.append("")
                                self.markdown_lines.append("```json")
                                for line in json_metadata.split('\n'):
                                    self.markdown_lines.append(line)
                                self.markdown_lines.append("```")
                                self.markdown_lines.append("")
                                self.markdown_lines.append("</details>")
                                self.markdown_lines.append("")
                    except Exception as e:
                        print(f"[WARNING] Failed to add shape metadata: {e}")
                
                self._mark_image_emitted(img_name)
                return len(self.markdown_lines)

            # insert_atをクランプ
            try:
                if insert_at < 0:
                    insert_at = 0
            except Exception:
                insert_at = 0
            if insert_at > len(self.markdown_lines):
                insert_at = len(self.markdown_lines)

            # 複数挿入の相対順序を保持するため空行とmd行を挿入
            self.markdown_lines.insert(insert_at, "")
            self.markdown_lines.insert(insert_at, md_line)
            
            lines_added = 2
            
            if sheet is not None:
                try:
                    filter_ids = self._image_shape_ids.get(img_name)
                    shapes_metadata = self._extract_all_shapes_metadata(sheet, filter_ids=filter_ids)
                    if shapes_metadata:
                        text_metadata = self._format_shape_metadata_as_text(shapes_metadata)
                        if text_metadata:
                            self.markdown_lines.insert(insert_at + lines_added, "")
                            lines_added += 1
                            for line in text_metadata.split('\n'):
                                self.markdown_lines.insert(insert_at + lines_added, line)
                                lines_added += 1
                            self.markdown_lines.insert(insert_at + lines_added, "")
                            lines_added += 1
                        
                        json_metadata = self._format_shape_metadata_as_json(shapes_metadata)
                        if json_metadata and json_metadata != "{}":
                            self.markdown_lines.insert(insert_at + lines_added, "<details>")
                            lines_added += 1
                            self.markdown_lines.insert(insert_at + lines_added, "<summary>JSON形式の図形情報</summary>")
                            lines_added += 1
                            self.markdown_lines.insert(insert_at + lines_added, "")
                            lines_added += 1
                            self.markdown_lines.insert(insert_at + lines_added, "```json")
                            lines_added += 1
                            for line in json_metadata.split('\n'):
                                self.markdown_lines.insert(insert_at + lines_added, line)
                                lines_added += 1
                            self.markdown_lines.insert(insert_at + lines_added, "```")
                            lines_added += 1
                            self.markdown_lines.insert(insert_at + lines_added, "")
                            lines_added += 1
                            self.markdown_lines.insert(insert_at + lines_added, "</details>")
                            lines_added += 1
                            self.markdown_lines.insert(insert_at + lines_added, "")
                            lines_added += 1
                except Exception as e:
                    print(f"[WARNING] Failed to add shape metadata: {e}")
            
            self._mark_image_emitted(img_name)
            return insert_at + lines_added
        except Exception:
            # フォールバック: 追加
            self.markdown_lines.append(md_line)
            self.markdown_lines.append("")
            self._mark_image_emitted(img_name)
            return len(self.markdown_lines)
    
    def _set_excel_fit_to_one_page(self, xlsx_path: str) -> bool:
        """ExcelファイルのpageSetupを縦横1ページに設定
        
        Args:
            xlsx_path: 対象のExcelファイルパス
            
        Returns:
            設定成功時True、失敗時False
        """
        try:
            import zipfile
            import tempfile
            import shutil
            import xml.etree.ElementTree as ET
            
            # 一時ディレクトリに解凍
            tmpdir = tempfile.mkdtemp(prefix='xls2md_fitpage_')
            try:
                with zipfile.ZipFile(xlsx_path, 'r') as zin:
                    zin.extractall(tmpdir)
                
                # 全シートのpageSetupを設定
                xl_worksheets = os.path.join(tmpdir, 'xl', 'worksheets')
                if os.path.exists(xl_worksheets):
                    for fname in os.listdir(xl_worksheets):
                        if fname.endswith('.xml') and fname.startswith('sheet'):
                            sheet_path = os.path.join(xl_worksheets, fname)
                            try:
                                tree = ET.parse(sheet_path)
                                root = tree.getroot()
                                ns = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
                                
                                # 既存のpageSetupを削除
                                for ps in root.findall('.//{%s}pageSetup' % ns):
                                    root.remove(ps)
                                
                                # 新しいpageSetupを追加（scaleで縮小）
                                # LibreOfficeはfitToWidth/fitToHeightを無視するため、scaleを使用
                                ps = ET.Element('{%s}pageSetup' % ns)
                                # 8ページ→1ページなので大幅縮小が必要
                                # まずは25%で試行（後で調整可能）
                                ps.set('scale', '25')
                                ps.set('orientation', 'landscape')  # 横向きで大きな図形に対応
                                ps.set('paperSize', '9')  # A4サイズ
                                ps.set('useFirstPageNumber', '1')
                                root.append(ps)
                                
                                # pageMargins を調整（余白を最小化）
                                for pm in root.findall('.//{%s}pageMargins' % ns):
                                    root.remove(pm)
                                pm = ET.Element('{%s}pageMargins' % ns)
                                pm.set('left', '0.25')
                                pm.set('right', '0.25')
                                pm.set('top', '0.25')
                                pm.set('bottom', '0.25')
                                pm.set('header', '0.0')
                                pm.set('footer', '0.0')
                                root.append(pm)
                                
                                # ファイルに書き戻し
                                tree.write(sheet_path, encoding='utf-8', xml_declaration=True)
                                debug_print(f"[DEBUG] {fname} のpageSetupを縦横1ページに設定")
                            except Exception as e:
                                print(f"[WARNING] {fname} のpageSetup設定に失敗: {e}")
                
                # 変更を元のファイルに上書き保存
                with zipfile.ZipFile(xlsx_path, 'w', zipfile.ZIP_DEFLATED) as zout:
                    for root_dir, dirs, files in os.walk(tmpdir):
                        for file in files:
                            file_path = os.path.join(root_dir, file)
                            arcname = os.path.relpath(file_path, tmpdir)
                            zout.write(file_path, arcname)
                
                return True
            finally:
                try:
                    shutil.rmtree(tmpdir)
                except Exception:
                    pass  # 一時ファイルの削除失敗は無視
        except Exception as e:
            print(f"[ERROR] pageSetup設定に失敗: {e}")
            return False
    
    def _convert_sheet(self, sheet):
        """シートを変換"""
        sheet_name = sheet.title
        
        # 前のシート毎の状態をクリアしてデフォルトを初期化
        self._clear_sheet_state(sheet_name)
        
        # このシートのデフォルトを初期化
        self._cell_to_md_index.setdefault(sheet_name, {})
        self._sheet_shape_images.setdefault(sheet_name, [])
        self._sheet_shape_next_idx.setdefault(sheet_name, 0)
        self._sheet_deferred_texts.setdefault(sheet_name, [])
        self._sheet_deferred_tables.setdefault(sheet_name, [])
        self._sheet_emitted_texts.setdefault(sheet_name, set())
        self._sheet_emitted_rows.setdefault(sheet_name, set())
        self._embedded_image_cid_by_name.setdefault(sheet_name, {})
        # 描画アンカーcNvPr IDの軽量マッピングを構築（順序付き）することで
        # クラスタループが候補クラスタが以前の分離レンダリングで
        # 既に保持されたアンカーを含むかどうかを迅速に判定できます。
        anchors_cid_list = []
        try:
            try:
                ztmp = zipfile.ZipFile(self.excel_file)
                sheet_index_tmp = self.workbook.sheetnames.index(sheet.title)
                rels_path_tmp = f"xl/worksheets/_rels/sheet{sheet_index_tmp+1}.xml.rels"
                if rels_path_tmp in ztmp.namelist():
                    rels_xml_tmp = ET.fromstring(ztmp.read(rels_path_tmp))
                    drawing_target_tmp = None
                    for rel in rels_xml_tmp.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                        if rel.attrib.get('Type','').endswith('/drawing'):
                            drawing_target_tmp = rel.attrib.get('Target')
                            break
                    if drawing_target_tmp:
                        drawing_path_tmp = normalize_excel_path(drawing_target_tmp)
                        if drawing_path_tmp not in ztmp.namelist():
                            drawing_path_tmp = drawing_path_tmp.replace('worksheets', 'drawings')
                        if drawing_path_tmp in ztmp.namelist():
                            drawing_xml_tmp = ET.fromstring(ztmp.read(drawing_path_tmp))
                            for node_tmp in drawing_xml_tmp:
                                lname_tmp = node_tmp.tag.split('}')[-1].lower()
                                if lname_tmp in ('twocellanchor', 'onecellanchor'):
                                    cid_tmp = extract_anchor_id(node_tmp, allow_idx=True)
                                    anchors_cid_list.append(str(cid_tmp) if cid_tmp is not None else None)
            except (ET.ParseError, KeyError, AttributeError):
                anchors_cid_list = []
        except (ValueError, TypeError):
            anchors_cid_list = []
        
        # シート見出し（番号とアンカーIDを削除）
        self.markdown_lines.append(f"## {sheet_name} (Sheet Data)")
        self.markdown_lines.append("")
        # シート先頭の説明文を、データ範囲の手前まで（最大12行）スキャンして
        # 表示順どおりに出力する。これにより「MailBoxより先の処理は」のような
        # 任意の説明文が欠落したり順序が入れ替わる問題を防止する。
        try:
            # 検出されたdata_range開始の直前にある連続した非空行のみを出力します。
            # これによりシート上の上から下への順序を保持しつつ、
            # データの遥か上に表示される可能性のある無関係なヘッダーブロックの
            # 出力を回避します。シートにまだdata_rangeがない場合は、
            # 保守的なヒューリスティックとして最初の12行をスキャンすることにフォールバックします。
            # 現在のmarkdown挿入インデックスに設定され、延期された
            # 画像処理を要求する際に使用されます。以前はinsert_posが
            # 代入前に参照されてUnboundLocalErrorを引き起こしていました。
            insert_pos = len(self.markdown_lines)
            max_head_scan = min(12, sheet.max_row)
            data_range = self._get_data_range(sheet)  # UnboundLocalErrorを避けるためここで初期化
            head_rows = []
            # max_head_scanまで各行をスキャンし、行毎に結合された非空
            # セルテキストを収集します。ここではmarkdownを出力しません --- 出力
            # （「このシートには表示可能なデータがありません」メッセージを含む）
            # は正規の出力パス中に行う必要があります。
            # プレスキャン中の出力により、同じメッセージが全ての非空セルに対して
            # 繰り返し追加されていました; 代わりに、今は行テキストのみを収集し
            # 挿入を延期します。
            for r in range(1, max_head_scan + 1):
                row_texts = []
                for c in range(1, min(20, sheet.max_column) + 1):
                    try:
                        cell = sheet.cell(r, c)
                        v = cell.value
                    except Exception:
                        cell = None
                        v = None
                    if v is not None:
                        s = str(v).strip()
                        if s:
                            # セルの書式を適用
                            if cell is not None:
                                s = self._apply_cell_formatting(cell, s)
                            row_texts.append(s)
                # この行のセル値を結合; 空行にはNoneを保持
                if row_texts:
                    combined = ' '.join(row_texts)
                else:
                    combined = None
                head_rows.append(combined)

            emitted_any = False
            # 収集されたhead_rowsを延期テキストとして登録し、
            # 正規の出力パス中にのみ出力されるようにします
            # （同じメッセージのセル毎の繰り返し出力を防ぎます）。
            for idx_row, combined in enumerate(head_rows, start=1):
                if not combined:
                    continue
                # 隣接する同一の延期テキストの重複を回避
                lst = self._sheet_deferred_texts.setdefault(sheet.title, [])
                if len(lst) > 0 and lst[-1][1].strip() == combined.strip():
                    continue
                lst.append((idx_row, combined))

            if data_range:
                start_row = data_range[0]
                # start_rowの直前の連続した非空行を出力（スキャンされたhead_rows内）
                # 1..min(max_head_scan, start_row-1)の中から候補行を検索
                cand_end = min(max_head_scan, start_row - 1)
                if cand_end >= 1:
                    # cand_endから逆方向に歩いて連続した非空ブロックを検索
                    block = []
                    for rr in range(cand_end, 0, -1):
                        content = head_rows[rr-1]
                        if content is None:
                            # 最初の空白に遭遇したら停止
                            break
                        block.insert(0, (rr, content))
                    # ヘッダーブロックを後でソート出力するため延期バッファに収集
                    for (rnum, combined) in block:
                        if len(self.markdown_lines) > 0 and self.markdown_lines[-1].strip() == combined:
                            continue
                        self._sheet_deferred_texts.setdefault(sheet.title, []).append((rnum, combined))
                        emitted_any = True
            else:
                # まだdata_rangeが検出されていません: 元の保守的な動作にフォールバック
                for r in range(1, max_head_scan + 1):
                    combined = head_rows[r-1]
                    if not combined:
                        continue
                    if len(self.markdown_lines) > 0 and self.markdown_lines[-1].strip() == combined:
                        continue
                    # 出力を延期: 後でソート出力するためヘッダー行を収集
                    self._sheet_deferred_texts.setdefault(sheet.title, []).append((r, combined))
                    emitted_any = True

        except Exception as e:
            print(f"[WARNING] シートヘッダー処理でエラー: {e}")
            # エラー時はdata_rangeを再取得
            data_range = self._get_data_range(sheet)

        # シート内のデータ範囲を確認
        if not data_range:
            # シートにデータが無い場合でも、描画が存在するなら図を出力する
            try:
                insert_pos = len(self.markdown_lines)
                # 描画がある場合、_process_sheet_imagesは挿入を延期し
                # 正規エミッタ（_reorder_sheet_output_by_row_order）が
                # 画像を決定論的に配置できるようにします。延期挿入を要求します。
                self._process_sheet_images(sheet, insert_index=insert_pos, insert_images=False)
                if not self._sheet_has_drawings(sheet):
                    # 複数の
                    # 異なるブランチによる同一メッセージの追加を避けるため正規対応の自由テキストエミッタを使用します。
                    self._emit_free_text(sheet, None, "*このシートには表示可能なデータがありません*")
                    # 正規の出力パス中のみ末尾の空行を追加
                    if getattr(self, '_in_canonical_emit', False):
                        self.markdown_lines.append("")
                else:
                    # 描画が処理されました; セパレータを追加して続行
                    self._add_separator()
                return
            except Exception:
                self._emit_free_text(sheet, None, "*このシートには表示可能なデータがありません*")
                if getattr(self, '_in_canonical_emit', False):
                    self.markdown_lines.append("")
                return
        else:
            # data_range が存在しても、セルの実体（テキスト等）が無い場合がある
            # （罫線や書式のみで範囲が検出されるケース）。その場合は空の表を出力しない。
            try:
                r1, r2, c1, c2 = data_range
                has_content = False
                for rr in range(r1, r2 + 1):
                    for cc in range(c1, c2 + 1):
                        try:
                            v = sheet.cell(row=rr, column=cc).value
                        except Exception:
                            v = None
                        if v is not None and str(v).strip():
                            has_content = True
                            break
                    if has_content:
                        break
                if not has_content:
                    # セル内容が無いため、図のみを挿入して終了する
                    insert_pos = len(self.markdown_lines)
                    # 正規パス中に画像が出力されるよう挿入を延期
                    self._process_sheet_images(sheet, insert_index=insert_pos, insert_images=False)
                    if not self._sheet_has_drawings(sheet):
                        self._emit_free_text(sheet, None, "*このシートには表示可能なデータがありません*")
                        if getattr(self, '_in_canonical_emit', False):
                            self.markdown_lines.append("")
                    else:
                        self._add_separator()
                    return
            except Exception:
                # 何か失敗した場合は従来の処理にフォールバック
                pass
        # まずデータをテーブルとして変換（図は表の下に出力したいので後で処理）
        # 注意: ここではreturnしません — 画像の処理を続け、
        # 延期されたテキストと画像が決定論的に出力されるよう
        # 以下の正規の行順序出力パスを実行します。以前は早期returnが
        # 正規出力をバイパスし、テーブル/段落出力の欠落を引き起こしていました。
        self._convert_sheet_data(sheet, data_range)

        # テーブル出力後、図形を生成し（即座の挿入なし）、その後
        # シートテキストと画像を厳密に行番号の昇順で出力し、
        # MarkdownがExcelの上から下への順序と一致するようにします。これは各グループの
        # self._sheet_shape_imagesに格納された代表的なstart_rowを使用します。
        try:
            # 図形が生成され記録されることを確認; 延期挿入を要求
            insert_pos = len(self.markdown_lines)
            self._process_sheet_images(sheet, insert_index=insert_pos, insert_images=False)

            # 厳密な行順序出力を実行: 全てのテキスト行を収集
            # （既に出力されたものを除く）と全ての画像start_rowsを収集し、その後
            # 1..max_rowの行を歩き、テキストまたは画像が存在する場合に出力します。
            self._reorder_sheet_output_by_row_order(sheet)
        except Exception:
            # 保守的な動作にフォールバック: 保留中の画像を挿入しセパレータを追加
            # フォールバックの場合でも、正規
            # エミッタが配置と重複抑制を制御するよう延期挿入を優先します。
            self._process_sheet_images(sheet, insert_index=len(self.markdown_lines), insert_images=False)
        finally:
            # チャートデータを抽出してMarkdownテーブルとして出力
            self._process_sheet_charts(sheet)
            # シート処理後の最終セパレータ
            self._add_separator()

    def _process_sheet_charts(self, sheet):
        """シート内のチャートを個別に画像化し、画像の下にデータを配置する
        
        各チャートを個別にレンダリングし、画像とデータテーブルを出力する。
        チャートがある場合、シート全体画像は出力から除外する。
        
        Args:
            sheet: openpyxlのワークシートオブジェクト
        """
        try:
            ws_charts = getattr(sheet, "_charts", [])
            chart_data_list = extract_charts_from_worksheet(sheet, self.workbook)
            
            if not ws_charts:
                return
            
            self._remove_sheet_image_from_output(sheet.title)
            
            print(f"[INFO] シート '{sheet.title}' から {len(ws_charts)} 個のチャートを画像化")
            
            for i, chart in enumerate(ws_charts):
                image_filename = self._render_chart_as_image(sheet, chart, i)
                
                if image_filename:
                    self.markdown_lines.append(f"![](images/{image_filename})")
                    self.markdown_lines.append("")
                
                if i < len(chart_data_list):
                    md_content = chart_data_to_markdown(chart_data_list[i])
                    for line in md_content.split('\n'):
                        self.markdown_lines.append(line)
                    
        except Exception as e:
            print(f"[WARNING] チャート処理中にエラー: {e}")
            import traceback
            traceback.print_exc()
    
    def _remove_sheet_image_from_output(self, sheet_title):
        """シート全体画像をMarkdown出力から除外する
        
        チャートがあるシートでは、シート全体画像（*_sheet.png/svg）は
        冗長なため除外する。
        
        Args:
            sheet_title: シート名
        """
        pattern = f"_sheet."
        lines_to_remove = []
        
        for i, line in enumerate(self.markdown_lines):
            if f"![" in line and pattern in line and sheet_title in line:
                lines_to_remove.append(i)
        
        for i in reversed(lines_to_remove):
            del self.markdown_lines[i]
            if i < len(self.markdown_lines) and self.markdown_lines[i] == "":
                del self.markdown_lines[i]
    
    def _render_chart_as_image(self, sheet, chart, chart_index: int):
        """チャートを個別に画像としてレンダリングする
        
        Args:
            sheet: ワークシートオブジェクト
            chart: openpyxlのチャートオブジェクト
            chart_index: チャートのインデックス
            
        Returns:
            str: 画像ファイル名、失敗時はNone
        """
        try:
            from copy import deepcopy
            import re
            
            temp_wb = openpyxl.Workbook()
            temp_ws = temp_wb.active
            temp_ws.title = sheet.title
            
            cell_refs = self._extract_chart_cell_references(chart)
            self._copy_chart_data_cells(sheet, temp_ws, cell_refs)
            
            chart_copy = deepcopy(chart)
            chart_copy.anchor = "A1"
            temp_ws.add_chart(chart_copy)
            
            temp_xlsx = tempfile.mktemp(suffix='.xlsx')
            temp_wb.save(temp_xlsx)
            temp_wb.close()
            
            temp_pdf = self._convert_chart_excel_to_pdf(temp_xlsx)
            if not temp_pdf:
                os.unlink(temp_xlsx)
                return None
            
            self.image_counter += 1
            ext = self.output_format
            image_filename = f"{self.base_name}_{sheet.title}_chart_{chart_index + 1:03d}.{ext}"
            image_path = os.path.join(self.images_dir, image_filename)
            
            if self.output_format == 'svg':
                success = self._convert_chart_pdf_to_svg(temp_pdf, image_path)
            else:
                success = self._convert_chart_pdf_to_png(temp_pdf, image_path)
            
            os.unlink(temp_xlsx)
            os.unlink(temp_pdf)
            
            if success:
                print(f"[SUCCESS] チャート画像生成: {image_filename}")
                return image_filename
            
            return None
            
        except Exception as e:
            print(f"[WARNING] チャート画像レンダリングエラー: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _extract_chart_cell_references(self, chart):
        """チャートが参照しているセル範囲を抽出する
        
        Args:
            chart: openpyxlのチャートオブジェクト
            
        Returns:
            set: セル参照の集合（例: {('A', 2), ('A', 3), ('B', 2), ('B', 3)}）
        """
        import re
        cell_refs = set()
        
        def parse_range_ref(ref_str):
            """セル範囲参照を解析してセル座標のリストを返す"""
            if not ref_str:
                return []
            
            match = re.match(r"'?[^'!]+'?\!\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)", ref_str)
            if match:
                col_start, row_start, col_end, row_end = match.groups()
                row_start, row_end = int(row_start), int(row_end)
                
                from openpyxl.utils import column_index_from_string
                col_start_idx = column_index_from_string(col_start)
                col_end_idx = column_index_from_string(col_end)
                
                for row in range(row_start, row_end + 1):
                    for col_idx in range(col_start_idx, col_end_idx + 1):
                        cell_refs.add((col_idx, row))
            return []
        
        for series in chart.series:
            if hasattr(series, 'val') and series.val:
                if hasattr(series.val, 'numRef') and series.val.numRef:
                    parse_range_ref(series.val.numRef.f)
            
            if hasattr(series, 'cat') and series.cat:
                if hasattr(series.cat, 'numRef') and series.cat.numRef:
                    parse_range_ref(series.cat.numRef.f)
                if hasattr(series.cat, 'strRef') and series.cat.strRef:
                    parse_range_ref(series.cat.strRef.f)
            
            if hasattr(series, 'xVal') and series.xVal:
                if hasattr(series.xVal, 'numRef') and series.xVal.numRef:
                    parse_range_ref(series.xVal.numRef.f)
            
            if hasattr(series, 'yVal') and series.yVal:
                if hasattr(series.yVal, 'numRef') and series.yVal.numRef:
                    parse_range_ref(series.yVal.numRef.f)
        
        return cell_refs
    
    def _copy_chart_data_cells(self, src_sheet, dst_sheet, cell_refs):
        """チャートが参照しているセルのデータをコピーする
        
        Args:
            src_sheet: コピー元のワークシート
            dst_sheet: コピー先のワークシート
            cell_refs: セル参照の集合（(col_idx, row)のタプル）
        """
        from openpyxl.utils import get_column_letter
        
        for col_idx, row in cell_refs:
            col_letter = get_column_letter(col_idx)
            src_cell = src_sheet[f"{col_letter}{row}"]
            dst_cell = dst_sheet[f"{col_letter}{row}"]
            dst_cell.value = src_cell.value
    
    def _convert_chart_excel_to_pdf(self, xlsx_path: str):
        """チャート用ExcelファイルをPDFに変換する
        
        Args:
            xlsx_path: Excelファイルのパス
            
        Returns:
            str: PDFファイルのパス、失敗時はNone
        """
        try:
            from utils import get_libreoffice_path
            
            lo_path = get_libreoffice_path()
            temp_dir = tempfile.mkdtemp()
            
            cmd = [
                lo_path,
                '--headless',
                '--convert-to', 'pdf',
                '--outdir', temp_dir,
                xlsx_path
            ]
            
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
            
            if result.returncode != 0:
                debug_print(f"[DEBUG] LibreOffice変換エラー: {result.stderr}")
                shutil.rmtree(temp_dir)
                return None
            
            pdf_name = os.path.splitext(os.path.basename(xlsx_path))[0] + '.pdf'
            pdf_path = os.path.join(temp_dir, pdf_name)
            
            if os.path.exists(pdf_path):
                final_pdf = tempfile.mktemp(suffix='.pdf')
                shutil.move(pdf_path, final_pdf)
                shutil.rmtree(temp_dir)
                print(f"[INFO] Excel→PDF変換完了: {final_pdf}")
                return final_pdf
            
            shutil.rmtree(temp_dir)
            return None
            
        except Exception as e:
            print(f"[WARNING] Excel→PDF変換エラー: {e}")
            return None
    
    def _convert_chart_pdf_to_svg(self, pdf_path: str, output_path: str) -> bool:
        """チャートPDFをSVGに変換する（PyMuPDF使用）
        
        Args:
            pdf_path: PDFファイルのパス
            output_path: 出力SVGファイルのパス
            
        Returns:
            bool: 成功時True
        """
        try:
            import fitz
            import numpy as np
            from PIL import Image as PILImage
            import re
            
            doc = fitz.open(pdf_path)
            if len(doc) == 0:
                print("[ERROR] PDFにページが含まれていません")
                doc.close()
                return False
            
            page = doc[0]
            
            dpi = 300
            zoom = dpi / 72
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat, alpha=False)
            
            img = PILImage.frombytes("RGB", [pix.width, pix.height], pix.samples)
            img_array = np.array(img)
            
            if len(img_array.shape) == 3:
                gray = np.mean(img_array, axis=2)
            else:
                gray = img_array
            
            threshold = 250
            non_white_pixels = gray < threshold
            
            rows = np.any(non_white_pixels, axis=1)
            cols = np.any(non_white_pixels, axis=0)
            
            svg_content = page.get_svg_image()
            width_units = page.rect.width
            height_units = page.rect.height
            
            if rows.any() and cols.any():
                row_indices = np.where(rows)[0]
                col_indices = np.where(cols)[0]
                
                top_px = row_indices[0]
                bottom_px = row_indices[-1] + 1
                left_px = col_indices[0]
                right_px = col_indices[-1] + 1
                
                scale_x = width_units / pix.width
                scale_y = height_units / pix.height
                
                left_u = left_px * scale_x
                top_u = top_px * scale_y
                width_u = (right_px - left_px) * scale_x
                height_u = (bottom_px - top_px) * scale_y
                
                svg_content = self._update_chart_svg_content(
                    svg_content, left_u, top_u, width_u, height_u
                )
            
            doc.close()
            
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(svg_content)
            
            print(f"[INFO] チャートSVG変換完了: {output_path}")
            return True
            
        except Exception as e:
            print(f"[WARNING] チャートSVG変換エラー: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _update_chart_svg_content(self, svg_content, left, top, width, height, scale=2.0):
        """チャートSVGのviewBoxとサイズを更新する
        
        Args:
            svg_content: SVG文字列
            left, top, width, height: 新しいviewBox座標
            scale: 表示サイズの倍率
            
        Returns:
            str: 更新されたSVG文字列
        """
        import re
        
        new_viewbox = f'viewBox="{left:.2f} {top:.2f} {width:.2f} {height:.2f}"'
        svg_content = re.sub(
            r'viewBox="[^"]*"',
            new_viewbox,
            svg_content,
            count=1
        )
        
        display_width = width * scale
        svg_content = re.sub(
            r'width="[^"]*"',
            f'width="{display_width:.2f}"',
            svg_content,
            count=1
        )
        
        display_height = height * scale
        svg_content = re.sub(
            r'height="[^"]*"',
            f'height="{display_height:.2f}"',
            svg_content,
            count=1
        )
        
        return svg_content
    
    def _convert_chart_pdf_to_png(self, pdf_path: str, output_path: str) -> bool:
        """チャートPDFをPNGに変換する（PyMuPDF使用）
        
        Args:
            pdf_path: PDFファイルのパス
            output_path: 出力PNGファイルのパス
            
        Returns:
            bool: 成功時True
        """
        try:
            import fitz
            from PIL import Image as PILImage
            import io
            
            doc = fitz.open(pdf_path)
            if len(doc) == 0:
                print("[ERROR] PDFにページが含まれていません")
                doc.close()
                return False
            
            page = doc[0]
            
            mat = fitz.Matrix(300/72, 300/72)
            pix = page.get_pixmap(matrix=mat, alpha=False)
            
            img_data = pix.tobytes("png")
            pix = None
            doc.close()
            
            img = PILImage.open(io.BytesIO(img_data))
            
            if img.mode == 'RGBA':
                background = PILImage.new('RGB', img.size, (255, 255, 255))
                background.paste(img, mask=img.split()[3] if len(img.split()) > 3 else None)
                img = background
            elif img.mode != 'RGB':
                img = img.convert('RGB')
            
            img = self._trim_chart_margins(img)
            
            width, height = img.size
            new_width = int(width * 2)
            new_height = int(height * 2)
            img = img.resize((new_width, new_height), PILImage.Resampling.LANCZOS)
            
            img.save(output_path, 'PNG', quality=95)
            
            print(f"[INFO] チャートPNG変換完了: {output_path}")
            return True
            
        except Exception as e:
            print(f"[WARNING] チャートPNG変換エラー: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _trim_chart_margins(self, img):
        """チャート画像の余白をトリムする
        
        Args:
            img: PIL Imageオブジェクト
            
        Returns:
            PIL Image: トリムされた画像
        """
        try:
            import numpy as np
            
            img_array = np.array(img)
            
            if len(img_array.shape) == 3:
                gray = np.mean(img_array, axis=2)
            else:
                gray = img_array
            
            threshold = 250
            non_white = gray < threshold
            
            rows = np.any(non_white, axis=1)
            cols = np.any(non_white, axis=0)
            
            if not np.any(rows) or not np.any(cols):
                return img
            
            row_indices = np.where(rows)[0]
            col_indices = np.where(cols)[0]
            
            top, bottom = row_indices[0], row_indices[-1]
            left, right = col_indices[0], col_indices[-1]
            
            padding = 10
            top = max(0, top - padding)
            bottom = min(img_array.shape[0], bottom + padding)
            left = max(0, left - padding)
            right = min(img_array.shape[1], right + padding)
            
            return img.crop((left, top, right, bottom))
            
        except Exception as e:
            debug_print(f"[DEBUG] 余白トリムエラー: {e}")
            return img
    
    def _reorder_sheet_output_by_row_order(self, sheet):
        """Emit sheet content (text and deferred images) strictly by source row order.

        - Uses self._sheet_shape_images[sheet.title] which is a list of (start_row, filename)
        - Uses self._emit_free_text to avoid duplicates
        - Updates self._cell_to_md_index so images can anchor to emitted md indices
        """
        try:
            # 計測: reorderルーチンへのエントリをマーク
            debug_print(f"[DEBUG][_reorder_entry] sheet={sheet.title}")
            # 実行毎のディスク上マーカーを削除; デバッグトレースのみを保持。
            debug_print(f"[DEBUG][_reorder_entry_marker] sheet={sheet.title}")
            max_row = sheet.max_row
            # ここでシートごとの出力済み行セットを作成しない。
            # 正規エミッタのみがヘルパーを通じて_sheet_emitted_rowsを変更すべき。
            emitted = self._sheet_emitted_rows.get(sheet.title, set())
            # マッピングを構築: 行 -> 画像ファイル名のリスト（_sheet_shape_imagesから）
            img_map = {}
            pairs = self._sheet_shape_images.get(sheet.title, []) or []
            # pairsはファイル名のリストまたは(row, filename)のリストの可能性
            normalized_pairs = []
            for item in pairs:
                if isinstance(item, (list, tuple)) and len(item) >= 2:
                    try:
                        r = int(item[0]) if item[0] is not None else 1
                    except (ValueError, TypeError):
                        r = 1
                    normalized_pairs.append((r, item[1]))
                else:
                    # start_row=1のファイル名として扱う
                    normalized_pairs.append((1, str(item)))
            # normalized_pairsから代表的なstart_rowを直接使用し、
            # 出力が元のExcel行順序に従うようにします。これにより
            # テキスト内容に基づくシート固有のヒューリスティックを回避します。
            for r, fn in normalized_pairs:
                img_map.setdefault(r, []).append(fn)

            # また、現在のsheet_map（行->mdインデックス）を出力して比較できるようにします
            sheet_map = self._cell_to_md_index.get(sheet.title, {})
            debug_print(f"[DEBUG][_img_insertion_debug] sheet={sheet.title} sheet_map={sheet_map}")

            # 注意: eventsログを永続化できるようmarkdownダンプを後に移動
            # 現在のmarkdown状態のデバッグ出力の前に。

            # 正規出力モードに入る: 延期テキストは今
            # _emit_free_textによって実際にmarkdownバッファに追加されるべきです。
            self._in_canonical_emit = True

            # デバッグ: 出力済み行と延期テキストを
            # 正規出力に入った直後にダンプして、以前にマークされたものを確認できるようにします。
            emitted_rows = self._sheet_emitted_rows.get(sheet.title, set()) if hasattr(self, '_sheet_emitted_rows') else set()
            deferred_texts = self._sheet_deferred_texts.get(sheet.title, []) if hasattr(self, '_sheet_deferred_texts') else []
            try:
                debug_print(f"[DEBUG][_canonical_enter] sheet={sheet.title} emitted_rows={sorted(list(emitted_rows))[:50]} deferred_texts_count={len(deferred_texts)}")
            except (ValueError, TypeError):
                debug_print(f"[DEBUG][_canonical_enter] sheet={getattr(sheet, 'title', None)} emitted_rows=<error> deferred_texts_count=<error>")

            # 全ての項目がstart_row==1に集約された場合（保存されたリストがファイル名のみを含む場合に一般的）、
            # 描画セル範囲から代表的なstart_rowsを再計算し、
            # それらの計算された行に順番に画像を再分配することを試みます。これにより
            # _render_sheet_fallbackがstart rowsを永続化しなかった場合により正確な配置が生成されます。
            try:
                all_rows = [r for r, _ in normalized_pairs]
                filenames_only = all(r == 1 for r in all_rows) and len(normalized_pairs) > 0
            except (ValueError, TypeError):
                filenames_only = False
            if filenames_only:
                cell_ranges = self._extract_drawing_cell_ranges(sheet) or []
                if cell_ranges:
                    # 各アンカーインデックスをそのstart_rowにマップ
                    start_rows = [cr[2] for cr in cell_ranges]
                    # start_rowでアンカーインデックスをソート
                    idxs = list(range(len(start_rows)))
                    idxs.sort(key=lambda i: start_rows[i])
                    # 割り当てる画像グループの数
                    nimgs = len(normalized_pairs)
                    # インデックスをnimgsの連続したバケットに分割（カウントで）
                    buckets = [[] for _ in range(nimgs)]
                    for i, idx in enumerate(idxs):
                        buckets[i % nimgs].append(idx)
                    # 各バケットの代表的な行を計算し、順番にファイル名を割り当て
                    new_img_map = {}
                    for bi, bucket in enumerate(buckets):
                        insert_r = 1
                        try:
                            if bucket:
                                vals = [start_rows[i] for i in bucket if isinstance(i, int) and i < len(start_rows)]
                                if vals:
                                    insert_r = int(min(vals))
                        except (ValueError, TypeError):
                            insert_r = 1
                        # インデックスをnimgsで割った余りがこのバケットインデックスと等しいファイル名を割り当て
                        try:
                            for j, (_, fn) in enumerate(normalized_pairs):
                                if j % nimgs == bi:
                                    new_img_map.setdefault(insert_r, []).append(fn)
                        except (ValueError, TypeError) as e:
                            debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")
                    img_map = new_img_map
                    # より簡単な
                    # 実行後の検査のため、stdoutとloggerの両方にログ出力（利用可能な場合）。
                    msg = f"[DEBUG][_img_fallback_row] sheet={sheet.title} assigned_images_row={insert_r} images={list(img_map.get(insert_r,[]))}"
                    debug_print(msg)
                    # 正規の
                    # 以下の出力ループが調整された行アンカーを使用するようimg_mapからnormalized_pairsを再構築します。
                    new_normalized = []
                    for rr in sorted(img_map.keys()):
                        for fn in img_map.get(rr, []):
                            new_normalized.append((int(rr), fn))
                    if new_normalized:
                        normalized_pairs = new_normalized

            # 上記で実行された調整（例: フォールバック再アンカー）を
            # 反映するようnormalized_pairsからimg_mapを再構築します。
            try:
                img_map = {}
                for r, fn in normalized_pairs:
                    img_map.setdefault(r, []).append(fn)
            except (ValueError, TypeError) as e:
                debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")

            # 同じ行にiso_group（トリミング/グループ）画像が存在する場合、
            # それを優先し、その行の個別の埋め込み画像を抑制します。
            # これによりグループ化された
            # レンダリングが埋め込み画像を1つの合成PNGにキャプチャした場合、同じビジュアルコンテンツを2回出力することを回避します。
            try:
                for rr, fns in list(img_map.items()):
                    try:
                        has_group = any((('iso_group' in (fn or '')) or ('.fixed' in (fn or '')))
                                        for fn in fns)
                    except Exception:
                        has_group = False
                    if has_group:
                        kept = [fn for fn in fns if (('iso_group' in (fn or '')) or ('.fixed' in (fn or '')))]
                        suppressed = [fn for fn in fns if fn not in kept]
                        if suppressed:
                            debug_print(f"[DEBUG][_img_suppress] sheet={sheet.title} row={rr} suppressed={suppressed} kept={kept}")
                        img_map[rr] = kept
                    
            except (ValueError, TypeError) as e:
                debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")

            # フィルタされた可能性のあるimg_mapからnormalized_pairsを再構築し、
            # 以下の正規出力ループが更新されたセットを使用するようにします。
            try:
                new_normalized = []
                for rr in sorted(img_map.keys()):
                    for fn in img_map.get(rr, []):
                        new_normalized.append((int(rr), fn))
                if new_normalized:
                    normalized_pairs = new_normalized
            except (ValueError, TypeError) as e:
                debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")

            # 各非空ソース行のテキストを収集（既に出力された行はスキップ）
            # これは画像アンカーを決定する前に行う必要があり、新しく検出された
            # ヘッダー/テキスト行（まだself._cell_to_md_indexに存在しない可能性がある）が
            # 近くの画像のアンカーとして使用できるようにします。また、ヘッダースキャン中に
            # 以前に延期されたヘッダー/テキスト行をマージし、それらが
            # 正規のソート済み出力パスでのみ出力されるようにします（早期書き込みを防ぐ）。
            texts_by_row = {}
            try:
                # 以前に収集された延期ヘッダー/テキスト行を取得
                deferred = []
                if hasattr(self, '_sheet_deferred_texts'):
                    try:
                        deferred = self._sheet_deferred_texts.pop(sheet.title, []) or []
                    except Exception:
                        deferred = []
                if deferred:
                    try:
                        # 延期テキストをtexts_by_rowに統合し、出力済みセットを尊重
                        # ここでは行を出力済みとしてマークしないでください; 実際のマーキングは
                        # テキストが正規のmarkdownバッファに正常に書き込まれたときにのみ
                        # 出力ループ中に行われるべきです。ここで
                        # 行をマークすると、正式な出力済みセットが
                        # 早期に設定され、正当なテーブル行の刈り込みにつながっていました。
                        for dr, dtxt in deferred:
                            try:
                                rr = int(dr) if dr is not None else 1
                            except (ValueError, TypeError):
                                rr = 1
                            if rr in emitted:
                                continue
                            if dtxt:
                                texts_by_row[rr] = dtxt
                    except (ValueError, TypeError) as e:
                        debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")
            except (ValueError, TypeError) as e:
                debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")
            for r in range(1, max_row + 1):
                if r in emitted:
                    continue
                # deferred_textsから来たテキストがある行はスキップ（書式が適用済み）
                if r in texts_by_row:
                    continue
                row_texts = []
                for c in range(1, min(60, sheet.max_column) + 1):
                    try:
                        v = sheet.cell(r, c).value
                    except Exception:
                        v = None
                    if v is not None:
                        s = str(v).strip()
                        if s:
                            row_texts.append(s)
                if row_texts:
                    texts_by_row[r] = " ".join(row_texts)
                    # ここでは行を出力済みとしてマークしないでください。実際の正式な
                    # マーキングは正規の出力パス中に行う必要があります
                    # （_emit_free_text内、または画像/テキストが書き込まれたとき）
                    # テーブル行の早期刈り込みを避けるため。

            # よりシンプルな決定論的出力: 統一されたイベントリストを構築
            # テキスト項目（src_row -> コンテンツ）と画像項目（start_row -> ファイル名）の、
            # その後行でソートし順番に出力します。同一行の場合、テキストを
            # 画像の前に出力し、画像がそのアンカーテキストの直後に表示されるようにします。
            try:
                # 実際に出力されるイベントのリストを構築します。
                # 二重メンテナンスを防ぐため、出力リストの構築中はロギングリストを
                # 変更しません。代わりに、まず`events_emit`を構築し、その後
                # 既存のシートマッピングと最終的な出力リストから`events_log`を
                # 合成します。
                events_emit = []

                # まだ出力されていない新しく収集されたテキスト行を追加
                for r, txt in texts_by_row.items():
                    try:
                        events_emit.append((int(r), 0, 'text', txt))
                    except (ValueError, TypeError):
                        events_emit.append((1, 0, 'text', txt))

                # normalized_pairsから画像イベントを追加（順序1）
                for start_row, fn in normalized_pairs:
                    try:
                        r = int(start_row) if start_row is not None else 1
                    except (ValueError, TypeError):
                        r = 1
                    events_emit.append((r, 1, 'image', str(fn)))

                # 延期されたテーブル出力をイベントとして追加（順序0.5でテキストの後、画像の前に配置）
                try:
                    deferred_tables = self._sheet_deferred_tables.get(sheet.title, []) if hasattr(self, '_sheet_deferred_tables') else []
                    for entry in deferred_tables:
                        try:
                            # deferred_tablesエントリは(anchor, table_data, src_rows)
                            # または(anchor, table_data, src_rows, meta)の可能性。両方を正規化。
                            if isinstance(entry, (list, tuple)) and len(entry) >= 3:
                                anchor_row = entry[0]
                                t_data = entry[1]
                                src_rows = entry[2]
                                meta = entry[3] if len(entry) >= 4 else None
                            else:
                                # 予期しない形状: スキップ
                                continue
                            # 同じ行のテキストの後にテーブルを配置するため順序0.5を使用
                            events_emit.append((int(anchor_row) if anchor_row else 1, 0.5, 'table', (t_data, src_rows, meta)))
                        except (ValueError, TypeError):
                            try:
                                events_emit.append((int(anchor_row) if anchor_row else 1, 0.5, 'table', (t_data, src_rows, meta)))
                            except (ValueError, TypeError):
                                # このエントリを諦める
                                continue
                except (ValueError, TypeError) as e:
                    debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")

                # 延期されたテーブルのソース行と重複するテキストイベントを削除
                try:
                    table_src_rows = set()
                    try:
                        deferred_tables = self._sheet_deferred_tables.get(sheet.title, []) if hasattr(self, '_sheet_deferred_tables') else []
                    except Exception:
                        deferred_tables = []

                    for entry in deferred_tables:
                        try:
                            # 可能なエントリ形状を正規化: (anchor, table_data, src_rows) または
                            # (anchor, table_data, src_rows, meta)
                            anchor_row = None
                            src_rows = None
                            meta = None
                            tdata = None
                            if isinstance(entry, (list, tuple)) and len(entry) >= 3:
                                anchor_row = entry[0]
                                tdata = entry[1]
                                src_rows = entry[2]
                                meta = entry[3] if len(entry) >= 4 else None
                            else:
                                # 予期しない形状: ベストエフォートで展開を試行
                                try:
                                    anchor_row = entry[0]
                                except Exception:
                                    anchor_row = None
                                try:
                                    tdata = entry[1]
                                except Exception:
                                    tdata = None
                                src_rows = None

                            # 明示的なsrc_rowsが提供された場合、サニタイズして追加
                            added_any = False
                            if src_rows:
                                try:
                                    for rr in src_rows:
                                        try:
                                            if rr is None:
                                                continue
                                            table_src_rows.add(int(rr))
                                            added_any = True
                                        except (ValueError, TypeError):
                                            # 整数でないエントリをスキップ
                                            continue
                                except (ValueError, TypeError) as e:
                                    debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")

                            # ヒューリスティック: src_rowsが1つずれている場合があり、
                            # 実際のテーブルにはリストされたsrc_rowsの直後に
                            # 追加の行が含まれることがあります。次の行（max+1）に
                            # テキストが含まれている（先にtexts_by_rowで収集済み）か、
                            # 非空のセルがある場合、保守的にそれを含めます。
                            try:
                                if added_any:
                                    try:
                                        mx = max(int(x) for x in src_rows if x is not None)
                                    except (ValueError, TypeError):
                                        mx = None
                                    if mx is not None:
                                        cand = mx + 1
                                        if cand not in table_src_rows:
                                            has_text = False
                                            try:
                                                # texts_by_rowはこのスコープで先に構築済み
                                                if isinstance(texts_by_row, dict) and texts_by_row.get(cand):
                                                    has_text = True
                                            except Exception:
                                                has_text = False
                                            if not has_text:
                                                # フォールバック: 最初の60列で非空セルを検査
                                                try:
                                                    for cc in range(1, min(60, sheet.max_column) + 1):
                                                        v = sheet.cell(cand, cc).value
                                                        if v is not None and str(v).strip():
                                                            has_text = True
                                                            break
                                                except (ValueError, TypeError):
                                                    has_text = False
                                            if has_text:
                                                try:
                                                    table_src_rows.add(int(cand))
                                                except (ValueError, TypeError) as e:
                                                    debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")
                            except (ValueError, TypeError) as e:
                                debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")

                            # フォールバック: 明示的なsrc_rowsがないが、アンカーと
                            # テーブルデータがある場合、保守的にanchor..anchor+len(table)-1を追加
                            if (not added_any) and anchor_row and tdata and isinstance(tdata, list) and len(tdata) > 0:
                                start = int(anchor_row)
                                cnt = len(tdata)
                                for rr in range(start, start + cnt):
                                    table_src_rows.add(int(rr))

                            # この延期されたテーブルが検出されたタイトル（meta.title）を持つ場合、
                            # テーブルのアンカー行も重複テキストとして扱い、同じ行の
                            # フリーテキスト（例：生のタイトル）を抑制します。
                            # これは正規の出力器がテーブルメタデータパスを通じて
                            # タイトルを出力するためです。
                            if meta and isinstance(meta, dict) and meta.get('title') and anchor_row:
                                table_src_rows.add(int(anchor_row))
                        except (ValueError, TypeError):
                            continue

                    if table_src_rows:
                        filtered = []
                        for r, order, kind, payload in events_emit:
                            try:
                                if kind == 'text' and int(r) in table_src_rows:
                                    # タイトルはテーブルメタデータ経由で出力されます。重複出力を
                                    # 避けるため、同じソース行のフリーテキストをスキップします。
                                    continue
                            except (ValueError, TypeError):
                                # エラー時は保守的にテキストをスキップ
                                continue
                            filtered.append((r, order, kind, payload))
                        events_emit = filtered
                except (ValueError, TypeError) as e:
                    debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")

                # (row, order)で決定論的にソートし、元の相対順序を保持
                events_emit.sort(key=lambda e: (e[0], e[1]))
                # 決定論的トレース用に最終的なソート済み出力リストをログ出力。
                # スキャン/行/種類の取得を容易にするため、イベントごとに1行出力。
                try:
                    for e in events_emit:
                        try:
                            row = int(e[0])
                        except (ValueError, TypeError):
                            row = e[0]
                        order = e[1]
                        kind = e[2]
                        payload = e[3]
                        # 可読性のために簡潔なペイロードサマリーを構築。
                        try:
                            if kind == 'table':
                                tdata, src_rows, meta = payload
                                p_summary = f"rows={len(tdata)} src_rows={src_rows} meta={meta}"
                            elif kind == 'image':
                                p_summary = os.path.basename(str(payload))
                            else:
                                p_summary = str(payload)
                        except (ValueError, TypeError):
                            p_summary = str(payload)
                        log_line = f"row={row} order={order} kind={kind} payload={p_summary}"
                        debug_print(f"[DEBUG][_events_emit_sorted] sheet={sheet.title} {log_line}")
                except (ValueError, TypeError) as e:
                    debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")

                # 権威あるシートマッピング（以前に出力された行）と最終的な出力リストから
                # events_logを合成します。
                events_log = []
                try:
                    sheet_map_all = self._cell_to_md_index.get(sheet.title, {})
                    for r, md_idx in sheet_map_all.items():
                        try:
                            md_idx_i = int(md_idx)
                        except (ValueError, TypeError):
                            continue
                        try:
                            if 0 <= md_idx_i < len(self.markdown_lines):
                                text_line = (self.markdown_lines[md_idx_i] or "").rstrip()
                                if text_line.endswith("  "):
                                    text_line = text_line[:-2]
                                events_log.append((int(r), 0, 'text', text_line))
                        except (ValueError, TypeError):
                            continue
                except (ValueError, TypeError):
                    pass  # データ構造操作失敗は無視

                # 呼び出し元が完全な正規シーケンスを確認できるように、
                # 各event_emit項目のログ表現を追加します。
                for r, order, kind, payload in events_emit:
                    try:
                        if kind == 'text':
                            events_log.append((int(r), order, 'text', payload))
                        elif kind == 'table':
                            # payloadは(table_data, src_rows, meta)
                            try:
                                tdata = payload[0]
                                rows_count = len(tdata) if isinstance(tdata, list) else 0
                            except (ValueError, TypeError):
                                rows_count = 0
                            events_log.append((int(r), order, 'table', f"rows={rows_count}"))
                        else:
                            events_log.append((int(r), order, 'image', os.path.basename(str(payload))))
                    except (ValueError, TypeError):
                        events_log.append((int(r) if r else 1, order, kind, payload))

                # ログの決定論的順序も確保
                events_log.sort(key=lambda e: (e[0], e[1]))

                # 現在のメモリ内markdownを事前スキャンして、既に挿入された
                # 画像参照を検出し、出力済みとしてマークします。これにより
                # 正規の出力パスが重複を挿入しないようにします。これは
                # 延期された決定論的出力が実行される前に、以前のコードパスが
                # 画像を挿入した（insert_images=True）場合を処理します。
                try:
                    for ln in list(self.markdown_lines):
                        try:
                            m = re.search(r"!\[.*?\]\(images/([^\)]+)\)", ln or "")
                            if m:
                                imgnm = m.group(1)
                                self._mark_image_emitted(imgnm)
                        except Exception:
                            continue
                except Exception as e:
                    print(f"[WARNING] ファイル操作エラー: {e}")

                # 強力なデバッグ: 空/充填を検出できるように常にイベント数を出力
                debug_print(f"[DEBUG][_events_sorted] sheet={sheet.title} events_count_emit={len(events_emit)}")

                # markdownを変更する前に、events_emitから直接導出された
                # 正規の決定論的ログスナップショットを出力します。これにより
                # 別のevents_logリストを維持する必要がなくなり、ログが
                # 実際の出力シーケンスと一致することが保証されます。
                try:
                    debug_print(f"[DEBUG][_sorted_events_block] sheet={sheet.title} events_count={len(events_emit)}")

                    # 診断/ログ専用パス: デバッグ用にevents_emitシーケンスの
                    # 決定論的で人間が読めるスナップショットを出力します。このパスは
                    # self.markdown_linesを変更したり、副作用（ファイル、マッピング）を
                    # 作成する出力ヘルパーを呼び出してはいけません。下の正規の
                    # 出力ループが権威ある書き込みを担当します。
                    for row, _, kind, payload in events_emit:
                        try:
                            if kind == 'text':
                                print(f"  [LOG] text @{row}: {payload}")
                            elif kind == 'table':
                                # payloadは(table_data, src_rows, meta)の可能性
                                tdata = None
                                src_rows = None
                                meta = None
                                if isinstance(payload, (list, tuple)):
                                    tdata = payload[0] if len(payload) >= 1 else None
                                    src_rows = payload[1] if len(payload) >= 2 else None
                                    meta = payload[2] if len(payload) >= 3 else None
                                title = None
                                try:
                                    if isinstance(meta, dict):
                                        title = meta.get('title')
                                except Exception:
                                    title = None
                                if title:
                                    print(f"  [LOG] table @{row} title: {title} rows={len(tdata) if isinstance(tdata, list) else 'N/A'} src_rows={src_rows}")
                                else:
                                    print(f"  [LOG] table @{row} rows={len(tdata) if isinstance(tdata, list) else 'N/A'} src_rows={src_rows}")
                            else:  # 画像
                                print(f"  [LOG] image @{row}: {payload}")
                        except (ValueError, TypeError):
                            # 診断パスでは堅牢に; 例外を発生させない
                            print(f"  [LOG] event @{row} kind={kind} (payload unstable)")
                except (ValueError, TypeError) as e:
                    debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")

                # 決定論的順序でイベントを出力し、位置を記録します。
                # これはself.markdown_linesを変更します（正規の出力パス）。
                for row, _, kind, payload in events_emit:
                    if kind == 'text':
                        try:
                            emitted_ok = self._emit_free_text(sheet, row, payload)
                        except (ValueError, TypeError):
                            emitted_ok = False
                        if not emitted_ok:
                            # ベストエフォートフォールバック: エスケープ/正規化されたテキストを追加
                            try:
                                txt = self._escape_angle_brackets(payload) + "  "
                                debug_print(f"[DEBUG][_emit_fallback] row={row} text={txt} >>")
                                self.markdown_lines.append(txt)
                                # 正規パス中のみ権威あるマッピングを割り当て
                                # 正規パス中のみ権威あるマッピングを記録
                                md_idx = len(self.markdown_lines) - 1
                                self._mark_sheet_map(sheet.title, row, md_idx)
                                self._mark_emitted_row(sheet.title, row)
                                self._mark_emitted_text(sheet.title, self._normalize_text(payload))
                                self.markdown_lines.append("")
                            except (ValueError, TypeError) as e:
                                debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")
                    elif kind == 'table':
                        try:
                            # payloadは(table_data, src_rows)または
                            # (table_data, src_rows, meta). Normalize both shapes.
                            table_data = None
                            src_rows = None
                            meta = None
                            try:
                                if isinstance(payload, (list, tuple)):
                                    if len(payload) == 2:
                                        table_data, src_rows = payload
                                    elif len(payload) >= 3:
                                        table_data, src_rows, meta = payload[0], payload[1], payload[2]
                                    else:
                                        # 予期しない小さな形状
                                        table_data = payload[0] if len(payload) >= 1 else None
                                else:
                                    # 予期しない: payload全体をtable_dataとして扱う
                                    table_data = payload
                            except Exception:
                                table_data = payload

                            debug_print(f"[DEBUG][_emit_table] row={row} table_rows={len(table_data) if isinstance(table_data, list) else 0} src_rows={src_rows} meta={meta} >>")

                            # タイトル/メタが存在する場合、最初に出力（正規パス）
                            try:
                                title = None
                                if isinstance(meta, dict):
                                    title = meta.get('title')
                                if title:
                                    normalized_title = ' '.join(str(title).strip().split())
                                    
                                    if sheet.title not in self._sheet_emitted_table_titles:
                                        self._sheet_emitted_table_titles[sheet.title] = set()
                                    
                                    if normalized_title in self._sheet_emitted_table_titles[sheet.title]:
                                        debug_print(f"[DEBUG] タイトル '{title}' は既に出力済みのため抑制")
                                        should_emit_title = False
                                    else:
                                        should_emit_title = True
                                        try:
                                            if table_data and len(table_data) > 0:
                                                header_row = table_data[0]
                                                if isinstance(header_row, (list, tuple)):
                                                    header_text = ' '.join(str(cell).strip() for cell in header_row if cell)
                                                    normalized_header = ' '.join(header_text.split())
                                                    
                                                    if normalized_title and normalized_header:
                                                        if normalized_title in normalized_header or normalized_header.startswith(normalized_title):
                                                            should_emit_title = False
                                                            debug_print(f"[DEBUG] タイトル '{title}' はヘッダー行と重複しているため出力を抑制")
                                        except Exception as e:
                                            debug_print(f"[DEBUG] タイトル冗長性チェックエラー（無視）: {e}")
                                            should_emit_title = True
                                    
                                    self._sheet_emitted_table_titles[sheet.title].add(normalized_title)
                                    
                                    if should_emit_title:
                                        try:
                                            h = f"{self._escape_angle_brackets(title)}  "
                                            self.markdown_lines.append(h)
                                            md_idx = len(self.markdown_lines) - 1
                                            self._mark_sheet_map(sheet.title, row, md_idx)
                                            self._mark_emitted_text(sheet.title, self._normalize_text(title))
                                            self._mark_emitted_row(sheet.title, row)
                                        except Exception:
                                            self._emit_free_text(sheet, row, title)
                            except Exception as e:
                                pass  # XML解析エラーは無視

                            # 正規出力: テーブルを出力しマッピングを記録
                            try:
                                if table_data is not None:
                                    # ガード付きヘルパーを使用する出力を呼び出す
                                    self._output_markdown_table(table_data, source_rows=src_rows, sheet_title=sheet.title)
                                    self.markdown_lines.append("")
                            except Exception:
                                if table_data is not None:
                                    self._output_markdown_table(table_data)

                            # プルーニングロジック用にソース行を出力済みとしてマーク
                            if src_rows:
                                for rr in src_rows:
                                    self._mark_emitted_row(sheet.title, rr)
                        except (ValueError, TypeError) as e:
                            debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")
                    else:  # 画像
                        img_fn = payload
                        debug_print(f"[DEBUG][_emit_image] row={row} image={img_fn} >> ")
                        ref = f"images/{img_fn}"
                        # 重複画像を決定論的にゲート
                        if ref in self._emitted_images or img_fn in self._emitted_images:
                            continue
                        md = f"![{sheet.title}](images/{img_fn})"
                        # ヘルパーメソッドを使用してメタデータ付き画像を挿入
                        try:
                            filter_ids = self._image_shape_ids.get(img_fn)
                            shapes_metadata = self._extract_all_shapes_metadata(sheet, filter_ids=filter_ids)
                            
                            if shapes_metadata:
                                debug_print(f"[DEBUG] 図形メタデータ抽出成功: {img_fn} -> {len(shapes_metadata)} shapes")
                                text_metadata = self._format_shape_metadata_as_text(shapes_metadata)
                                
                                self.markdown_lines.append(md)
                                self.markdown_lines.append("")
                                
                                if text_metadata:
                                    self.markdown_lines.append("")
                                    for line in text_metadata.split('\n'):
                                        self.markdown_lines.append(line)
                                    self.markdown_lines.append("")
                                
                                if self.shape_metadata:
                                    json_metadata = self._format_shape_metadata_as_json(shapes_metadata)
                                    if json_metadata and json_metadata != "{}":
                                        self.markdown_lines.append("<details>")
                                        self.markdown_lines.append("<summary>JSON形式の図形情報</summary>")
                                        self.markdown_lines.append("")
                                        self.markdown_lines.append("```json")
                                        for line in json_metadata.split('\n'):
                                            self.markdown_lines.append(line)
                                        self.markdown_lines.append("```")
                                        self.markdown_lines.append("")
                                        self.markdown_lines.append("</details>")
                                        self.markdown_lines.append("")
                            else:
                                self.markdown_lines.append(md)
                                self.markdown_lines.append("")
                        except Exception as e:
                            print(f"[WARNING] 図形メタデータ追加失敗: {e}")
                            self.markdown_lines.append(md)
                            self.markdown_lines.append("")
                        # ヘルパーを通じてのみ正式なマッピングを記録
                        try:
                            md_idx = len(self.markdown_lines) - 2
                            self._mark_sheet_map(sheet.title, row, md_idx)
                        except (ValueError, TypeError):
                            debug_print(f"WARNING self._mark_sheet_map({sheet.title}, {row}, {md_idx})")
                        try:
                            # 重複を防ぐためemitted_imagesを常にマーク。これは安全です。
                            # emitted_imagesはファイル名のみを追跡し、プルーニングに影響しないため。
                            self._mark_image_emitted(img_fn)
                        except (ValueError, TypeError):
                            debug_print(f"WARNING self._mark_image_emitted({img_fn})")
            except (ValueError, TypeError):
                # 簡略化フローで何か問題が発生した場合、再スローして
                # 外部の例外ハンドラが後で保守的な挿入を実行できるように
                # 元の複雑な挿入パスにフォールバックします。
                raise

            # 上記の正規イベント出力が既に画像を決定論的に配置し、
            # self._emitted_imagesに記録したため、後続のアンカーベース/
            # フォールバック挿入パスを意図的にスキップします。
            # これにより複数の挿入コードパスによる二重挿入を回避します。
            # 正規出力モードを終了し、将来の_emit_free_text呼び出しが
            # 次の正規パスまで再び延期されるようにします。
            self._in_canonical_emit = False

            # 最後に、start_rowが範囲外またはstart_row==1で挿入されなかった
            # 画像が末尾に追加されることを確認（セーフティネット）。
            remaining_imgs = []
            for r, imgs in img_map.items():
                if r > max_row or r < 1:
                    remaining_imgs.extend(imgs)
            for r, imgs in img_map.items():
                if r == 1:
                    for i in imgs:
                        ref = f"images/{i}"
                        if ref not in '\n'.join(self.markdown_lines):
                            remaining_imgs.append(i)
            for img in remaining_imgs:
                ref = f"images/{img}"
                if ref in self._emitted_images or img in self._emitted_images:
                    continue
                md = f"![{sheet.title}](images/{img})"
                self.markdown_lines.append(md)
                self.markdown_lines.append("")
                try:
                    self._mark_sheet_map(sheet.title, 1, len(self.markdown_lines) - 2)
                except Exception:
                    debug_print(f"WARNING: Exception self._mark_sheet_map({sheet.title}, {1}, {len(self.markdown_lines) - 2})")
                try:
                    self._mark_image_emitted(img)
                except (ValueError, TypeError):
                    debug_print(f"WARNING: Exception self._mark_image_emitted({img})")
            # このシートの延期テーブルをクリア（既に出力済み）
            if hasattr(self, '_sheet_deferred_tables') and sheet.title in self._sheet_deferred_tables:
                del self._sheet_deferred_tables[sheet.title]
            # 最終ソートイベントフォールバック削除: ここでは追加ログなし
        except Exception as _exc:
            # デバッグ: 簡略化フローが失敗した理由を確認するため例外情報を出力
            debug_print(f"[DEBUG][_reorder_exception] sheet={sheet.title} exc={_exc!r}")
            import traceback
            traceback.print_exc()
            # エラー時は、全ての延期画像の即座挿入にフォールバック
            for item in self._sheet_shape_images.get(sheet.title, []) or []:
                fn = item[1] if isinstance(item, (list, tuple)) and len(item) >= 2 else str(item)
                md = f"![{sheet.title}](images/{fn})"
                self.markdown_lines.append(md)
                self.markdown_lines.append("")
    
    def _get_data_range(self, sheet) -> Optional[Tuple[int, int, int, int]]:
        """シート内のデータ範囲を取得 (min_row, max_row, min_col, max_col)"""
        try:
            # 実際にデータが存在する範囲を取得
            if self._is_empty_sheet(sheet):
                return None
                
            # 実際に値のあるセルの範囲を計算
            return self._calculate_data_bounds(sheet)
            
        except Exception as e:
            print(f"[WARNING] データ範囲取得エラー: {e}")
            return None

    def _prune_emitted_rows(self, sheet_title: str, table_data: List[List[str]], source_rows: Optional[List[int]]):
        """既に事前出力された行があれば、table_data と source_rows からその行を除去する。

        戻り値: (pruned_table_data, pruned_source_rows)
        """
        try:
            emitted = self._sheet_emitted_rows.get(sheet_title, set()) if hasattr(self, '_sheet_emitted_rows') else set()
        except (ValueError, TypeError):
            emitted = set()

        # 記録された出力済み行のうち、実際にmarkdownマッピングを持つもののみを
        # 刈り込みます。一部のコードパスは正規の書き込みが発生する前に、
        # 保守的に（または誤って）行を出力済みセットに追加した可能性があります。
        # _sheet_emitted_rowsに存在し、かつ_cell_to_md_indexに具体的な
        # マッピングを持つ（つまりself.markdown_linesに書き込まれた）行のみを
        # 刈り込みます。これにより、登録されただけでまだ出力されていない行が
        # 削除されることを回避します。
        try:
            sheet_map = self._cell_to_md_index.get(sheet_title, {}) if hasattr(self, '_cell_to_md_index') else {}
        except Exception:
            sheet_map = {}

        try:
            # 両方の構造に存在する行のみを正式なものとして扱う
            authoritative_emitted = set(r for r in emitted if r in sheet_map)
            sample_emitted = sorted(list(authoritative_emitted))[:20]
            debug_print(f"[TRACE][_prune_emitted_rows_entry] sheet={sheet_title} emitted_count_total={len(emitted)} emitted_count_auth={len(authoritative_emitted)} emitted_sample={sample_emitted} source_rows_count={len(source_rows) if source_rows else 0}")
        except (ValueError, TypeError):
            debug_print(f"[TRACE][_prune_emitted_rows_entry] sheet={sheet_title} unable to snapshot emitted set")

        if not authoritative_emitted or not source_rows:
            return table_data, source_rows

        pruned_table = []
        pruned_src = []
        for row, src in zip(table_data, source_rows):
            try:
                # ソース行が実際にmarkdownに出力された場合のみ刈り込み
                # （authoritative_emittedに存在）。より広いemittedセットにのみ
                # リストされ、markdownマッピングを持たない行はここで保持されます。
                if src not in authoritative_emitted:
                    pruned_table.append(row)
                    pruned_src.append(src)
                else:
                    # デバッグ: このソース行は以前の正式な出力により削除された
                    debug_print(f"[TRACE][_prune_emitted_rows_removed] sheet={sheet_title} removed_src_row={src}")
            except (ValueError, TypeError):
                pruned_table.append(row)
                pruned_src.append(src)

        debug_print(f"[TRACE][_prune_emitted_rows_exit] sheet={sheet_title} in={len(source_rows)} out={len(pruned_src)}")

        return pruned_table, pruned_src
    
    def _is_empty_sheet(self, sheet) -> bool:
        """シートが空かどうかをチェック"""
        return (sheet.max_row == 1 and 
                sheet.max_column == 1 and 
                not sheet.cell(1, 1).value)
    
    def _calculate_data_bounds(self, sheet) -> Optional[Tuple[int, int, int, int]]:
        """データの境界を計算"""
        min_row, max_row = None, None
        min_col, max_col = None, None
        
        for row in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row, col)
                if cell.value is not None or self._has_cell_formatting(cell):
                    min_row, max_row = self._update_row_bounds(min_row, max_row, row)
                    min_col, max_col = self._update_col_bounds(min_col, max_col, col)
        
        return (min_row, max_row, min_col, max_col) if min_row is not None else None
    
    def _update_row_bounds(self, min_row: Optional[int], _max_row: Optional[int], row: int) -> Tuple[int, int]:
        """行の境界を更新"""
        new_min_row = row if min_row is None else min_row
        new_max_row = row
        return new_min_row, new_max_row
    
    def _update_col_bounds(self, min_col: Optional[int], max_col: Optional[int], col: int) -> Tuple[int, int]:
        """列の境界を更新"""
        new_min_col = col if min_col is None or col < min_col else min_col
        new_max_col = col if max_col is None or col > max_col else max_col
        return new_min_col, new_max_col
    
    def _has_cell_formatting(self, cell) -> bool:
        """セルに特別な書式設定があるかチェック"""
        try:
            # セルの背景色、罫線、フォントスタイルなどをチェック
            if (cell.fill and cell.fill.fgColor and hasattr(cell.fill.fgColor, 'rgb') and 
                cell.fill.fgColor.rgb and cell.fill.fgColor.rgb != 'FFFFFFFF'):
                return True
            
            if cell.border and any([
                cell.border.left.style, cell.border.right.style, 
                cell.border.top.style, cell.border.bottom.style
            ]):
                return True
                
            if cell.font and (cell.font.bold or cell.font.italic):
                return True
                
            return False
        except Exception:
            return False
    
    def _convert_sheet_data(self, sheet, data_range: Tuple[int, int, int, int]):
        """シートデータをテーブルとして変換（複数テーブル対応）"""
        min_row, max_row, min_col, max_col = data_range
        
        print(f"[INFO] データ範囲: 行{min_row}〜{max_row}, 列{min_col}〜{max_col}")
        
        # 罫線で囲まれた矩形領域のみを表として抽出
        print("[INFO] 罫線で囲まれた領域によるテーブル抽出を開始...")
        table_regions = self._detect_bordered_tables(sheet, min_row, max_row, min_col, max_col)
        debug_print(f"[DEBUG][_convert_sheet_data] bordered_table_regions_count={len(table_regions)} sample={table_regions[:5]}")

        # 離散データ領域を追跡するセット（strict_column_bounds=Trueで処理するため）
        # 罫線テーブルも含める（罫線テーブルも列範囲の拡張を制限する）
        discrete_region_set = set()
        
        # 罫線テーブルが占有するセルのマスクを作成（離散領域検出で除外するため）
        occupied_cells = set()
        for tr in table_regions:
            tr_r1, tr_r2, tr_c1, tr_c2 = tr
            for r in range(tr_r1, tr_r2 + 1):
                for c in range(tr_c1, tr_c2 + 1):
                    occupied_cells.add((r, c))
            # 罫線テーブルも列範囲の拡張を制限するためセットに追加
            discrete_region_set.add(tr)

        # 罫線テーブルがある場合、その外側の領域に対してのみ離散領域検出を実行
        # これにより、罫線なしテーブルが罫線テーブルの横に配置されている場合も検出可能
        # five_sheet_.xlsxのように罫線テーブルがないシートでは離散領域検出は実行されない
        if table_regions:
            debug_print(f"[DEBUG] bordered tables found; trying discrete region detection for non-occupied areas")
            try:
                # 離散データ領域検出を実行（罫線テーブルの占有セルを除外）
                discrete_regions = self._find_discrete_data_regions(
                    sheet, min_row, max_row, min_col, max_col, occupied_cells
                )
                if discrete_regions:
                    # 既存の罫線テーブルと重複しない離散領域のみを追加
                    new_discrete_regions = []
                    for dr in discrete_regions:
                        dr_r1, dr_r2, dr_c1, dr_c2 = dr
                        is_overlapping = False
                        for tr in table_regions:
                            tr_r1, tr_r2, tr_c1, tr_c2 = tr
                            # 重複チェック（行と列の両方が重複している場合）
                            if not (dr_r2 < tr_r1 or dr_r1 > tr_r2 or dr_c2 < tr_c1 or dr_c1 > tr_c2):
                                is_overlapping = True
                                break
                        if not is_overlapping:
                            new_discrete_regions.append(dr)
                            discrete_region_set.add(dr)
                    if new_discrete_regions:
                        debug_print(f"[DEBUG] adding {len(new_discrete_regions)} non-overlapping discrete regions")
                        table_regions = table_regions + new_discrete_regions
            except Exception as _e:
                debug_print(f"[DEBUG] discrete region detection failed: {_e}")

        # 罫線テーブルが見つからない場合、または罫線テーブルが上部の行（1-4）を含まない場合、
        # ヒューリスティック（結合セル、注釈、列分離）を使用するより広範なテーブル領域検出を試行。
        # これにより、シート上部のヘッダー行が適切に検出されることを保証
        top_region_in_bordered = any(r[0] == 1 and r[1] <= 4 for r in table_regions)
        if not table_regions or not top_region_in_bordered:
            try:
                if not table_regions:
                    debug_print("[DEBUG] no bordered tables found; trying heuristic _detect_table_regions first")
                    # まずヒューリスティック検出を試行（mainブランチとの互換性維持）
                    heur_tables, heur_annotations = self._detect_table_regions(sheet, min_row, max_row, min_col, max_col)
                    if heur_tables:
                        debug_print(f"[DEBUG] heuristic detection found {len(heur_tables)} table regions")
                        table_regions = heur_tables
                    else:
                        # 離散領域検出は一時的に無効化（テキストがテーブルとして検出される問題を回避）
                        debug_print("[DEBUG] heuristic detection failed; discrete region detection is disabled")
                else:
                    debug_print(f"[DEBUG] bordered tables found but no top region (rows 1-4); trying heuristic _detect_table_regions to find header rows")
                    heur_tables, heur_annotations = self._detect_table_regions(sheet, min_row, max_row, min_col, max_col)
                    try:
                        debug_print(f"[TRACE][_detect_table_regions_result] sheet={sheet.title} heur_tables_count={len(heur_tables) if heur_tables else 0} heur_annotations_count={len(heur_annotations) if heur_annotations else 0}")
                        if heur_tables:
                            debug_print(f"[TRACE][_detect_table_regions_result_sample] {heur_tables[:10]}")
                        if heur_annotations:
                            debug_print(f"[TRACE][_detect_table_regions_annotations_sample] {heur_annotations[:10]}")
                    except (ValueError, TypeError) as e:
                        debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")
                    if heur_tables:
                        top_heur_tables = [r for r in heur_tables if r[0] == 1 and r[1] <= 4]
                        if top_heur_tables:
                            debug_print(f"[DEBUG] adding {len(top_heur_tables)} top regions from heuristic detection to bordered tables")
                            table_regions = top_heur_tables + table_regions
            except Exception as _e:
                debug_print(f"[DEBUG] table detection failed: {_e}")

        # 変更: 描画（図形/画像）が占有するセル領域を検出し、重複する表領域は
        # テーブルとして出力せずプレーンテキスト扱いで出力する（ユーザ要望）。
        try:
            drawing_cell_ranges = self._extract_drawing_cell_ranges(sheet)
        except (ValueError, TypeError):
            drawing_cell_ranges = []

        def region_overlaps_drawings(region, drawing_ranges, overlap_threshold=0.25):
            """
            テーブル領域(region)が描画領域と重複するか確認する（重複割合ベース）。
            - region: (start_row, end_row, start_col, end_col)
            - drawing_ranges: list of (start_col, end_col, start_row, end_row)
            - overlap_threshold: テーブルセル総数に対する重複セル数の割合の閾値。デフォルト0.25

            戻り値: Trueなら重複とみなす（テーブルを除外）。
            """
            if not drawing_ranges:
                return False

            r1, r2, c1, c2 = region if len(region) == 4 else (region[0], region[1], region[2], region[3])

            table_cells = max(0, (r2 - r1 + 1)) * max(0, (c2 - c1 + 1))
            if table_cells <= 0:
                return False

            overlap_cells = 0
            for dr in drawing_ranges:
                try:
                    d_c1, d_c2, d_r1, d_r2 = dr
                except Exception:
                    continue
                inter_r1 = max(r1, d_r1)
                inter_r2 = min(r2, d_r2)
                inter_c1 = max(c1, d_c1)
                inter_c2 = min(c2, d_c2)
                if inter_r1 <= inter_r2 and inter_c1 <= inter_c2:
                    overlap_cells += (inter_r2 - inter_r1 + 1) * (inter_c2 - inter_c1 + 1)

            frac = overlap_cells / table_cells if table_cells > 0 else 0.0
            
            num_rows = r2 - r1 + 1
            num_cols = c2 - c1 + 1
            is_very_large_table = num_rows > 50 and num_cols > 30
            
            if is_very_large_table and len(drawing_ranges) >= 10:
                adjusted_threshold = 0.02
                debug_print(f"[DEBUG] Large table with many drawings detected - using stricter threshold {adjusted_threshold}")
            else:
                adjusted_threshold = overlap_threshold
            
            debug_print(f"[DEBUG] table_region={region} overlap_cells={overlap_cells} table_cells={table_cells} frac={frac:.3f} threshold={adjusted_threshold:.3f}")

            return frac >= adjusted_threshold

        # table_regionsを保持するものと重複で除外するものに分割
        kept_table_regions = []
        excluded_table_regions = []
        for tr in table_regions:
            # _detect_bordered_tables returns (r1, r2, c1, c2)
            if region_overlaps_drawings(tr, drawing_cell_ranges):
                print(f"[INFO] テーブル領域が描画と重複しているため除外: {tr}")
                excluded_table_regions.append(tr)
            else:
                kept_table_regions.append(tr)

        filtered_table_regions = []
        for i, region_a in enumerate(kept_table_regions):
            r1_a, r2_a, c1_a, c2_a = region_a
            width_a = c2_a - c1_a + 1
            height_a = r2_a - r1_a + 1
            
            is_nested = False
            for j, region_b in enumerate(kept_table_regions):
                if i == j:
                    continue
                r1_b, r2_b, c1_b, c2_b = region_b
                width_b = c2_b - c1_b + 1
                height_b = r2_b - r1_b + 1
                
                if (r1_b <= r1_a and r2_a <= r2_b and 
                    c1_b <= c1_a and c2_a <= c2_b):
                    if width_a == 1 and width_b >= 2:
                        overlap_ratio = height_a / height_b if height_b > 0 else 0
                        if overlap_ratio > 0.8:
                            debug_print(f"[DEBUG] ネストされた1列テーブルを除外: {region_a} (含まれる先: {region_b}, 幅: {width_a} vs {width_b}, 重複率: {overlap_ratio:.2f})")
                            is_nested = True
                            break
            
            if not is_nested:
                filtered_table_regions.append(region_a)

        table_regions = filtered_table_regions
        debug_print(f"[DEBUG][_convert_sheet_data] kept_table_regions_count={len(table_regions)} kept_sample={table_regions[:5]}")

        # 重複テーブルを除外する処理
        def regions_overlap(r1, r2, threshold=0.5):
            """2つのテーブル領域が重複しているかチェック"""
            row1_start, row1_end, col1_start, col1_end = r1
            row2_start, row2_end, col2_start, col2_end = r2
            
            overlap_row_start = max(row1_start, row2_start)
            overlap_row_end = min(row1_end, row2_end)
            overlap_col_start = max(col1_start, col2_start)
            overlap_col_end = min(col1_end, col2_end)
            
            if overlap_row_start > overlap_row_end or overlap_col_start > overlap_col_end:
                return False
            
            overlap_cells = (overlap_row_end - overlap_row_start + 1) * (overlap_col_end - overlap_col_start + 1)
            
            r1_cells = (row1_end - row1_start + 1) * (col1_end - col1_start + 1)
            r2_cells = (row2_end - row2_start + 1) * (col2_end - col2_start + 1)
            smaller_cells = min(r1_cells, r2_cells)
            
            overlap_ratio = overlap_cells / smaller_cells if smaller_cells > 0 else 0
            
            return overlap_ratio >= threshold
        
        # 重複テーブルを除外（大きいテーブルを優先）
        deduplicated_regions = []
        for i, region in enumerate(table_regions):
            is_duplicate = False
            for j, other_region in enumerate(table_regions):
                if i != j and regions_overlap(region, other_region):
                    r1_cells = (region[1] - region[0] + 1) * (region[3] - region[2] + 1)
                    r2_cells = (other_region[1] - other_region[0] + 1) * (other_region[3] - other_region[2] + 1)
                    
                    if r1_cells < r2_cells:
                        debug_print(f"[DEBUG] 重複テーブルを除外: {region} (重複先: {other_region})")
                        is_duplicate = True
                        break
                    elif r1_cells == r2_cells and i > j:
                        debug_print(f"[DEBUG] 重複テーブルを除外: {region} (重複先: {other_region})")
                        is_duplicate = True
                        break
            
            if not is_duplicate:
                deduplicated_regions.append(region)
        
        table_regions = deduplicated_regions
        debug_print(f"[DEBUG][_convert_sheet_data] deduplicated_table_regions_count={len(table_regions)} deduplicated_sample={table_regions[:5]}")

        processed_rows = set()
        # 検出されたテーブル領域を実際のテーブルとして出力（行の予約だけでなく）。
        # ここで検出された各テーブル領域をマークダウンに変換し、
        # 行を処理済みとしてマークして、後続のプレーンテキスト収集がスキップするようにする。
        table_index = 0
        for region in table_regions:
            debug_print(f"[DEBUG] emitting detected table region: {region}")
            try:
                # 検出された領域をマークダウンテーブルに変換。単調増加する
                # table_indexを使用して、ファイル名/IDが実行間で
                # 決定論的になるようにする。
                # 離散データ領域の場合は列範囲の拡張を制限
                # all_table_regionsを渡して、他テーブル領域内の行をタイトル候補から除外
                is_discrete = region in discrete_region_set
                self._convert_table_region(sheet, region, table_number=table_index, strict_column_bounds=is_discrete, all_table_regions=table_regions)
                table_index += 1
            except Exception as _e:
                debug_print(f"[DEBUG] _convert_table_region failed for region={region}: {_e}")
            # 変換の成否に関わらず行を処理済みとしてマークし、
            # プレーンテキストとして再収集されないようにする。
            for r in range(region[0], region[1]+1):
                processed_rows.add(r)

        # 除外されたテーブル領域（描画と重複）がある場合、行テキストを収集するが、
        # 他のプレーンテキストとマージするまで実際の出力を遅延させ、
        # 最終出力がシートの厳密な行順序を保持するようにする。
        excluded_blocks = []  # (start_row, end_row, [(row, text), ...])のリスト
        excluded_end_rows = set()
        if 'excluded_table_regions' in locals() and excluded_table_regions:
            for excl in excluded_table_regions:
                try:
                    print(f"[INFO] 描画重複のためプレーンテキストとして収集: {excl}")
                    srow, erow, sc, ec = excl
                    lines = []
                    for rr in range(srow, erow + 1):
                        # 行テキストを収集
                        row_texts = []
                        for col_num in range(sc, ec + 1):
                            if rr <= sheet.max_row and col_num <= sheet.max_column:
                                cell = sheet.cell(row=rr, column=col_num)
                                cell_value = cell.value
                                if cell_value is not None:
                                    text = str(cell_value).strip()
                                    if text:
                                        # セルの書式を適用
                                        text = self._apply_cell_formatting(cell, text)
                                        row_texts.append(text)
                        if row_texts:
                            lines.append((rr, " ".join(row_texts)))
                    if lines:
                        excluded_blocks.append((srow, erow, lines))
                        excluded_end_rows.add(erow)
                        # 後で再収集されないよう行を処理済みとしてマーク
                        # 除外された領域内で実際にテキストをキャプチャした場合のみマーク。
                        # テキストがキャプチャされなかった場合（例：関連テキストが
                        # 除外された列の外側にある場合）、行をマークせず、
                        # 後でプレーンテキストとして発見できるようにする。
                        for (rr, _) in lines:
                            processed_rows.add(rr)
                except Exception:
                    pass  # データ構造操作失敗は無視

        # プレーンテキスト領域を先に走査して収集する
        # 変更点: プレーン判定でTrueにならない場合でも、非空の行を"説明文"として出力するフォールバックを追加
        plain_texts = []  # (row_num, text)のリスト
        for row_num in range(min_row, max_row + 1):
            if row_num in processed_rows:
                continue
            region = (row_num, row_num, min_col, max_col)
            # 行のテキストを結合
            row_texts = []
            for col_num in range(min_col, max_col + 1):
                if row_num <= sheet.max_row and col_num <= sheet.max_column:
                    cell = sheet.cell(row=row_num, column=col_num)
                    v = cell.value
                    if v is not None:
                        s = str(v).strip()
                        if s:
                            # セルの書式を適用
                            s = self._apply_cell_formatting(cell, s)
                            row_texts.append(s)
            if not row_texts:
                continue
            line = " ".join(row_texts)
            # プレーンテキスト判定を試みるが、判定がFalseでも説明的な単一カラムの行などは
            # ユーザ期待の文書テキストである可能性が高いためフォールバックで出力対象にする
            is_plain = self._is_plain_text_region(sheet, region)
            if is_plain or len(row_texts) == 1:
                plain_texts.append((row_num, line))
            else:
                # フォールバック: 非表形式として出力する候補に含める
                plain_texts.append((row_num, line))
            # 注意: ここで行を処理済みとしてマークしない。以前は
            # 行を即座に追加していたため、後続の
            # 暗黙テーブル検出が連続した複数列の
            # 実行を見ることができなかった（candidate_rowsが空になった）。
            # 暗黙テーブル検出と実際の出力後までマークを延期。

        # プレーンテキストはシート上の行順でそのまま出力する。
        # plain_texts（processed_rowsにない行）を収集し、
        # 先に収集したexcluded_blocksとマージし、行番号でソートして
        # 最終出力がシートの上から下への順序に従うようにする。
        # plain_texts は (row_num, line) のリストなので行番号でソートしておく。
        # 除外ブロック行をplain_textsコンテナにマージ
        merged_texts = []
        for (srow, erow, lines) in excluded_blocks:
            for (r, txt) in lines:
                merged_texts.append((r, txt, True if r == erow else False))

        plain_texts.sort(key=lambda x: x[0])
        for (r, line) in plain_texts:
            merged_texts.append((r, line, False))

        # 検出されたテーブルとして既に処理された行がmerged_textsに
        # 存在しないことを確認。これにより、最終ソートと出力段階で
        # 同じコンテンツが変換されたテーブルとフリーフォームテキストの
        # 両方として出力されることを回避。
        if processed_rows:
            try:
                before_count = len(merged_texts)
                merged_texts = [t for t in merged_texts if t[0] not in processed_rows]
                debug_print(f"[DEBUG] filtered merged_texts: removed {before_count - len(merged_texts)} rows that were already processed as tables")
            except (ValueError, TypeError) as e:
                debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")

        # merged_textsを出力する前に、複数の非空列を持つ連続した行で
        # 形成される暗黙のテーブルを検出しようとする。これにより、
        # テーブル検出ヒューリスティックがテーブルを見逃したが、データが
        # 明らかに表形式である場合（例：2列以上の非空列を持つ多くの行）を回復。
        try:
            candidate_rows = [r for r in range(min_row, max_row + 1) if r not in processed_rows]
            # 行 -> 非空列インデックスのリストをマッピング
            row_cols = {}
            for r in candidate_rows:
                cols = [c for c in range(min_col, max_col + 1) if r <= sheet.max_row and c <= sheet.max_column and sheet.cell(r, c).value is not None and str(sheet.cell(r, c).value).strip()]
                row_cols[r] = cols

            # 各行が少なくとも2つの非空列を持つ連続した実行を検索
            runs = []
            cur_run = []
            for r in candidate_rows:
                if len(row_cols.get(r, [])) >= 2:
                    cur_run.append(r)
                else:
                    if cur_run:
                        runs.append((cur_run[0], cur_run[-1]))
                        cur_run = []
            if cur_run:
                runs.append((cur_run[0], cur_run[-1]))

            # 第1パス: 暗黙テーブル領域を収集（変換はまだ行わない）
            implicit_table_regions = []
            for (srow, erow) in runs:
                if (erow - srow + 1) >= 3:
                    cols_used = [c for r in range(srow, erow + 1) for c in row_cols.get(r, [])]
                    if cols_used:
                        smin = min(cols_used)
                        smax = max(cols_used)
                        
                        if self._is_colon_separated_list(sheet, srow, erow, smin, smax):
                            debug_print(f"[DEBUG] implicit table is colon-separated list; skipping rows={srow}-{erow}")
                            continue
                        
                        # 番号付きリストのチェック
                        skip_run = False
                        try:
                            content_cols_set = set(cols_used)
                            if len(content_cols_set) == 2:
                                sorted_cols = sorted(list(content_cols_set))
                                lcol, rcol = sorted_cols[0], sorted_cols[1]
                                l_texts = []
                                r_texts = []
                                for rr in range(srow, erow + 1):
                                    try:
                                        lv = sheet.cell(rr, lcol).value
                                    except Exception:
                                        lv = None
                                    try:
                                        rv = sheet.cell(rr, rcol).value
                                    except Exception:
                                        rv = None
                                    if lv is not None and str(lv).strip():
                                        l_texts.append(str(lv).strip())
                                    if rv is not None and str(rv).strip():
                                        r_texts.append(str(rv).strip())

                                if l_texts and r_texts and len(l_texts) >= 2:
                                    import re
                                    circled = '①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬⑭⑮⑯⑰⑱⑲⑳'
                                    num_matches = 0
                                    for t in l_texts:
                                        tt = t.strip()
                                        if any(ch in circled for ch in tt):
                                            num_matches += 1
                                        elif re.match(r'^[0-9]+[\.)]?$|^[A-Za-z]$|^[IVXivx]+$', tt):
                                            num_matches += 1
                                        elif len(tt) <= 2:
                                            num_matches += 1

                                    ratio = num_matches / len(l_texts) if l_texts else 0.0
                                    r_avg = sum(len(x) for x in r_texts) / len(r_texts) if r_texts else 0
                                    if ratio >= 0.8 and r_avg >= 8:
                                        debug_print(f"[DEBUG] implicit run looks like enumerated list; skipping rows={srow}-{erow}")
                                        skip_run = True
                        except (ValueError, TypeError) as e:
                            debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")
                        
                        if not skip_run:
                            implicit_table_regions.append((srow, erow, smin, smax))
                            debug_print(f"[DEBUG] implicit table detected rows={srow}-{erow} cols={smin}-{smax}")
            
            # 暗黙テーブル領域をtable_regionsに追加（タイトル検出用）
            all_implicit_regions = table_regions + implicit_table_regions if table_regions else implicit_table_regions
            
            # 第2パス: 収集した領域を変換（all_table_regionsとして全領域を渡す）
            for (srow, erow, smin, smax) in implicit_table_regions:
                try:
                    self._convert_table_region(sheet, (srow, erow, smin, smax), table_number=0, all_table_regions=all_implicit_regions)
                    for rr in range(srow, erow + 1):
                        processed_rows.add(rr)
                except Exception:
                    pass  # データ構造操作失敗は無視
        except Exception:
            pass  # データ構造操作失敗は無視

        # merged_textsを行番号（昇順）でソートしてシート順序を保持
        merged_texts.sort(key=lambda x: x[0])
        # 上記の暗黙テーブル検出で行が変換された場合、merged_textsを確認
        # これらの行がもう含まれていないことを確認。この二重チェックにより、
        # merged_textsが最初に構築された後に暗黙テーブルが見つかった場合の重複を防止。
        if processed_rows:
            try:
                before_count2 = len(merged_texts)
                merged_texts = [t for t in merged_texts if t[0] not in processed_rows]
                debug_print(f"[DEBUG] post-implicit-filter: removed {before_count2 - len(merged_texts)} rows processed by implicit-table conversion")
            except (ValueError, TypeError) as e:
                debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")
        # マージされたフリーフォームテキストエントリを行番号の昇順で出力。
        last_emitted_row = None
        if merged_texts:
            debug_print(f"[DEBUG] merged_texts出力開始: {len(merged_texts)}件")
            for (r, txt, is_excl_end) in merged_texts:
                debug_print(f"[DEBUG] merged_texts出力: 行{r}, text='{txt[:50]}...' (is_excl_end={is_excl_end})")
                self._emit_free_text(sheet, r, txt)
                # この行が除外ブロックの終端の場合、空行を追加
                # 正規出力パス中のみ出力済み行をマーク
                if is_excl_end and getattr(self, '_in_canonical_emit', False):
                    self.markdown_lines.append("")
                    # end_rowを空行インデックスにマッピングし出力済み行をマーク
                    try:
                        self._mark_sheet_map(sheet.title, r, len(self.markdown_lines) - 1)
                    except (ValueError, TypeError) as e:
                        debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")
                    try:
                        # 対応する除外ブロック内のすべての行を出力済みとしてマーク
                        for (srow, erow, lines) in excluded_blocks:
                            if erow == r:
                                for rr in range(srow, erow + 1):
                                    self._mark_emitted_row(sheet.title, rr)
                                break
                    except Exception as e:
                        pass  # XML解析エラーは無視
                last_emitted_row = r
            # マージされたフリーテキスト領域の後に区切りの空行を追加（実際に出力する場合のみ）
            if getattr(self, '_in_canonical_emit', False):
                self.markdown_lines.append("")
    
    def _detect_and_process_plain_text_regions(self, sheet, min_row: int, max_row: int, min_col: int, max_col: int, processed_rows: set = None) -> set:
        """プレーンテキスト領域を検出して処理し、処理済み行のセットを返す"""
        if processed_rows is None:
            processed_rows = set()
        for row_num in range(min_row, max_row + 1):
            if row_num in processed_rows:
                continue
            # この行がプレーンテキスト行かチェック
            region = (row_num, row_num, min_col, max_col)
            if self._is_plain_text_region(sheet, region):
                debug_print(f"[DEBUG] プレーンテキスト行を検出: 行{row_num}")
                # 連続するプレーンテキスト行を検索
                text_end_row = row_num
                for next_row in range(row_num + 1, max_row + 1):
                    next_region = (next_row, next_row, min_col, max_col)
                    if self._is_plain_text_region(sheet, next_region):
                        text_end_row = next_row
                    else:
                        break
                # プレーンテキスト領域を出力
                self._output_plain_text_region(sheet, row_num, text_end_row, min_col, max_col)
                # 処理済み行を記録
                for r in range(row_num, text_end_row + 1):
                    processed_rows.add(r)
        return processed_rows
    
    def _process_excluded_region_as_text(self, sheet, region: Tuple[int, int, int, int]):
        """フィルタで除外されたテーブル領域をプレーンテキストとして処理"""
        start_row, end_row, min_col, max_col = region
        
        for row_num in range(start_row, end_row + 1):
            # 行のテキストを収集
            row_texts = []
            for col_num in range(min_col, max_col + 1):
                if row_num <= sheet.max_row and col_num <= sheet.max_column:
                    cell = sheet.cell(row=row_num, column=col_num)
                    cell_value = cell.value
                    if cell_value is not None:
                        text = str(cell_value).strip()
                        if text:
                            # セルの書式を適用
                            text = self._apply_cell_formatting(cell, text)
                            row_texts.append(text)
            
            # 行にテキストがある場合は出力
            if row_texts:
                line_text = " ".join(row_texts)
                self._emit_free_text(sheet, row_num, line_text)
        
        if end_row >= start_row:  # 何らかのテキストが処理された場合
            # 正規出力時のみセパレータを追加し、出力済み行をマーク
            if getattr(self, '_in_canonical_emit', False):
                self.markdown_lines.append("")  # 空行を追加
                # end_rowを空行インデックスにマッピングし出力済み行をマーク (helper already registered normalized texts)
                try:
                    self._mark_sheet_map(sheet.title, end_row, len(self.markdown_lines) - 1)
                except Exception as e:
                    pass  # XML解析エラーは無視
                try:
                    for r in range(start_row, end_row + 1):
                        self._mark_emitted_row(sheet.title, r)
                except Exception as e:
                    pass  # XML解析エラーは無視
            else:
                debug_print(f"[TRACE] Skipping authoritative mapping for excluded_region rows {start_row}-{end_row} (non-canonical)")
    
    def _output_plain_text_region(self, sheet, start_row: int, end_row: int, min_col: int, max_col: int):
        """プレーンテキスト領域をMarkdownに出力"""
        text_content = []

        for row_num in range(start_row, end_row + 1):
            row_text = []
            for col_num in range(min_col, max_col + 1):
                cell = sheet.cell(row=row_num, column=col_num)
                if cell.value is not None:
                    text = str(cell.value).strip()
                    if text:
                        # セルの書式を適用
                        text = self._apply_cell_formatting(cell, text)
                        row_text.append(text)

            if row_text:
                text_content.append(" ".join(row_text))

        # 空でないテキストのみ出力
        if text_content:
            # ここでマッピングが存在しない場合は作成しない。マッピングは権威的であり、
            # _mark_sheet_mapを介した正規出力時にのみ設定されるべき。
            sheet_map = self._cell_to_md_index.get(sheet.title, {})
            for i, text in enumerate(text_content):
                if text.strip():
                    src_row = start_row + i
                    emitted = self._emit_free_text(sheet, src_row, text)
                    # emittedがFalseの場合は重複としてスキップ
            # 空行を追加し、最後のソース行を空白セパレータインデックスにマップ
            # 正規出力時のみ実際に追加し、出力済み行をマーク
            if getattr(self, '_in_canonical_emit', False):
                self.markdown_lines.append("")  # 空行を追加
                try:
                    self._mark_sheet_map(sheet.title, end_row, len(self.markdown_lines) - 1)
                except Exception as e:
                    pass  # XML解析エラーは無視
                # すべての出力済み行をマーク
                try:
                    for r in range(start_row, end_row + 1):
                        self._mark_emitted_row(sheet.title, r)
                except Exception as e:
                    pass  # XML解析エラーは無視
            debug_print(f"[DEBUG] プレーンテキスト出力: {len(text_content)}行")


def convert_xls_to_xlsx(xls_file_path: str) -> Optional[str]:
    """XLSファイルをXLSXに変換"""
    print(f"[INFO] XLSファイルをXLSXに変換中: {xls_file_path}")
    
    # 一時ディレクトリを作成
    temp_dir = tempfile.mkdtemp(prefix='xls2md_conversion_')
    
    try:
        # 出力ファイル名を決定
        xls_path = Path(xls_file_path)
        xlsx_filename = xls_path.stem + '.xlsx'
        xlsx_output_path = os.path.join(temp_dir, xlsx_filename)
        
        # LibreOfficeを使用してXLSをXLSXに変換
        cmd = [
            LIBREOFFICE_PATH,
            '--headless',
            '--convert-to', 'xlsx',
            '--outdir', temp_dir,
            xls_file_path
        ]
        
        debug_print(f"[DEBUG] LibreOffice変換コマンド: {' '.join(cmd)}")
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
        
        if result.returncode != 0:
            print(f"[ERROR] LibreOffice変換失敗: {result.stderr}")
            shutil.rmtree(temp_dir)
            return None
        
        # 変換されたファイルが存在するか確認
        if not os.path.exists(xlsx_output_path):
            print(f"[ERROR] 変換後のXLSXファイルが見つかりません: {xlsx_output_path}")
            shutil.rmtree(temp_dir)
            return None
        
        print(f"[SUCCESS] XLS→XLSX変換完了: {xlsx_output_path}")
        return xlsx_output_path
        
    except subprocess.TimeoutExpired:
        print("[ERROR] LibreOffice変換がタイムアウトしました")
        shutil.rmtree(temp_dir)
        return None
    except Exception as e:
        print(f"[ERROR] XLS変換エラー: {e}")
        shutil.rmtree(temp_dir)
        return None


def main():
    """メイン関数"""
    import argparse
    
    parser = argparse.ArgumentParser(description='ExcelファイルをMarkdownに変換')
    parser.add_argument('excel_file', help='変換するExcelファイル（.xlsx/.xls）')
    parser.add_argument('-o', '--output-dir', type=str, 
                       help='出力ディレクトリを指定（デフォルト: ./output）')
    parser.add_argument('--shape-metadata', action='store_true',
                       help='図形メタデータを画像の後に出力（テキスト形式とJSON形式）')
    parser.add_argument('--format', choices=['png', 'svg'], default='svg',
                       help='出力画像形式を指定（デフォルト: png）')
    parser.add_argument('-v', '--verbose', action='store_true',
                       help='デバッグ情報を出力し、debug_workbooks/pdfs/diagnosticsフォルダを保存')
    
    args = parser.parse_args()
    
    set_verbose(args.verbose)
    
    if not os.path.exists(args.excel_file):
        debug_print(f"エラー: ファイル '{args.excel_file}' が見つかりません。")
        sys.exit(1)
    
    if not args.excel_file.endswith(('.xlsx', '.xls')):
        debug_print("エラー: .xlsxまたは.xls形式のファイルを指定してください。")
        sys.exit(1)
    
    # XLSファイルの場合は事前にXLSXへ変換
    processing_file = args.excel_file
    converted_file = None
    converted_temp_dir = None
    
    if args.excel_file.endswith('.xls'):
        debug_print("XLSファイルが指定されました。XLSXに変換します...")
        converted_file = convert_xls_to_xlsx(args.excel_file)
        if converted_file is None:
            debug_print("XLS→XLSX変換に失敗しました。")
            sys.exit(1)
        processing_file = converted_file
        converted_temp_dir = Path(converted_file).parent
        debug_print(f"XLS→XLSX変換完了: {converted_file}")
    
    try:
        converter = ExcelToMarkdownConverter(
            processing_file, 
            output_dir=args.output_dir, 
            shape_metadata=args.shape_metadata,
            output_format=args.format
        )
        output_file = converter.convert()
        debug_print("\n変換完了!")
        debug_print(f"出力ファイル: {output_file}")
        debug_print(f"画像フォルダ: {converter.images_dir}")
        
    except Exception as e:
        debug_print(f"変換エラー: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        # 一時的に作成したXLSXファイルとその親ディレクトリを削除
        if converted_temp_dir:
            try:
                if converted_temp_dir.exists() and converted_temp_dir.name.startswith('xls2md_conversion_'):
                    shutil.rmtree(converted_temp_dir)
                    debug_print(f"一時ディレクトリを削除: {converted_temp_dir}")
            except Exception as cleanup_error:
                debug_print(f"一時ファイル削除に失敗: {cleanup_error}")


if __name__ == "__main__":
    main()
