#!/usr/bin/env python3
"""
画像・図形処理Mixinモジュール

ExcelToMarkdownConverterクラスの画像処理・図形処理機能を提供します。
このモジュールはMixinクラスとして設計されており、メインクラスから継承されます。

機能:
- シート画像の処理と抽出
- PDF/PNG/SVG変換
- 図形の抽出とクラスタリング
- 分離グループのレンダリング
"""

import os
import sys
import tempfile
import subprocess
import shutil
from pathlib import Path
from typing import List, Dict, Tuple, Optional, Any, Set
import io
import zipfile
import xml.etree.ElementTree as ET

from utils import get_libreoffice_path, col_letter, normalize_excel_path, get_xml_from_zip, extract_anchor_id, anchor_is_hidden, anchor_has_drawable as utils_anchor_has_drawable
from isolated_group_renderer import IsolatedGroupRenderer

try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
except ImportError as e:
    raise ImportError(
        "openpyxlライブラリが必要です: pip install openpyxl または uv sync を実行してください"
    ) from e

try:
    from PIL import Image
except ImportError as e:
    raise ImportError(
        "Pillowライブラリが必要です: pip install pillow または uv sync を実行してください"
    ) from e

try:
    import fitz
except ImportError as e:
    raise ImportError(
        "PyMuPDFライブラリが必要です: pip install PyMuPDF または uv sync を実行してください"
    ) from e

# 設定定数
LIBREOFFICE_PATH = get_libreoffice_path()

# DPI設定
DEFAULT_DPI = 600
IMAGE_QUALITY = 100
IMAGE_BORDER_SIZE = 8


def debug_print(*args, **kwargs):
    """デバッグ出力（x2mdモジュールに委譲）"""
    try:
        from x2md import debug_print as _dp
        _dp(*args, **kwargs)
    except ImportError:
        pass


class _GraphicsMixin:
    """画像・図形処理機能を提供するMixinクラス
    
    このクラスはExcelToMarkdownConverterに継承され、
    画像処理、図形抽出、レンダリング機能を提供します。
    
    注意: このクラスは単独では使用できません。
    ExcelToMarkdownConverterクラスと組み合わせて使用してください。
    """

    def _process_sheet_images(self, sheet, insert_index: Optional[int] = None, insert_images: bool = True):
        """シート内の画像を処理"""
        try:
            debug_print(f"[DEBUG][_process_sheet_images_entry] sheet={sheet.title} insert_index={insert_index} insert_images={insert_images}")
            debug_print(f"[DEBUG][_process_sheet_images_entry] sheet={sheet.title} insert_index={insert_index} insert_images={insert_images}")
            # 重複した重い処理を防止: 図形が既に生成されている場合
            # このシートの実行中に既に生成されている場合、処理をスキップして
            # tmp_xlsxの繰り返し作成と外部コンバーターの呼び出しを回避します。
            if sheet.title in self._sheet_shapes_generated:
                debug_print(f"[DEBUG][_process_sheet_images] sheet={sheet.title} shapes already generated; skipping repeated processing")
                return False
            images_found = False
            # 描画図形（ベクトル図形、コネクタなど）を確認
            # 埋め込み画像が見つかったかどうかに関係なく確認します。これにより
            # ベクトル図形のみ（埋め込み画像なし）のシートも正しく処理されます。
            # Phase 2-D修正: images_found=Trueの時だけでなく、常に描画図形を確認
            if True:  # 描画図形の分離グループ処理を常に実行
                    debug_print(f"[DEBUG] {len(sheet._images)} 個の埋め込み画像が検出されました。描画要素を調査中...")
                    # 埋め込み画像が1つ（またはゼロ）の場合、
                    # コストのかかるisolated-groupクラスタリングとトリミングされた
                    # ワークブックレンダリングを実行する代わりに、その画像を直接使用します。
                    # これによりtmp_xlsx/.fixed.xlsxの作成と
                    # 不要な場合の外部コンバーターを回避します（input_files/three_sheet_.xlsx
                    # のような単純なシートで一般的）。
                    try:
                        emb_count = len(getattr(sheet, '_images', []) or [])
                        # 埋め込み画像がちょうど1つ存在する場合、その画像を優先
                        # 直接使用し、重いisolated-group/フォールバックレンダリングをスキップします。
                        # これは単一の埋め込みグラフィックが存在する場合に
                        # クラスタリングを避けるユーザーのリクエストを尊重します。
                        if emb_count == 1:
                            debug_print(f"[DEBUG][_process_sheet_images_shortcircuit] sheet={sheet.title} single embedded image detected; using embedded image without clustering")
                            for image in sheet._images:
                                img_name = self._process_excel_image(image, f"{sheet.title} (Image)")
                                if img_name:
                                    start_row = 1
                                    try:
                                        pos = self._get_image_position(image)
                                        if pos and isinstance(pos, dict) and 'row' in pos:
                                            start_row = pos['row']
                                    except Exception:
                                        start_row = 1
                                    self._sheet_shape_images.setdefault(sheet.title, [])
                                    self._sheet_shape_images[sheet.title].append((start_row, img_name))
                            try:
                                self._sheet_shapes_generated.add(sheet.title)
                            except (ValueError, TypeError):
                                pass  # データ構造操作失敗は無視
                            return True
                        # 埋め込み画像がゼロの場合、描画チェックにフォールスルー
                        # し、必要に応じてisolated-groupまたはフルシートフォールバックを実行します。
                    except (ValueError, TypeError):
                        pass  # データ構造操作失敗は無視
                    try:
                        z = zipfile.ZipFile(self.excel_file)
                        sheet_index = self.workbook.sheetnames.index(sheet.title)
                        rels_path = f"xl/worksheets/_rels/sheet{sheet_index+1}.xml.rels"
                        rels_xml = get_xml_from_zip(z, rels_path)
                        if rels_xml is not None:
                            drawing_target = None
                            for rel in rels_xml.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                                t = rel.attrib.get('Type','')
                                if t.endswith('/drawing'):
                                    drawing_target = rel.attrib.get('Target')
                                    break
                            if drawing_target:
                                drawing_path = normalize_excel_path(drawing_target)
                                if drawing_path not in z.namelist():
                                    drawing_path = drawing_path.replace('worksheets', 'drawings')
                                drawing_xml = get_xml_from_zip(z, drawing_path)
                                if drawing_xml is not None:
                                    # 簡素化されバランスの取れた解析: アンカーIDを収集
                                    # and count pic/sp anchors. Also attempt to map any
                                    # embedded image filenames to their cNvPr ids. Keep
                                    # errors non-fatal and avoid deep nesting of try/except.
                                    anchors_cid_list = []
                                    total_anchors = 0
                                    pic_anchors = 0
                                    sp_anchors = 0
                                    try:
                                        # collect anchor ids and basic counts
                                        for node in list(drawing_xml):
                                            lname = node.tag.split('}')[-1].lower()
                                            if lname not in ('twocellanchor', 'onecellanchor'):
                                                continue
                                            total_anchors += 1
                                            for sub in node.iter():
                                                t = sub.tag.split('}')[-1].lower()
                                                if t == 'pic':
                                                    pic_anchors += 1
                                                if t == 'sp':
                                                    sp_anchors += 1
                                                if t == 'cnvpr':
                                                    cid_val = sub.attrib.get('id') or sub.attrib.get('idx')
                                                    anchors_cid_list.append(str(cid_val) if cid_val is not None else None)

                                        # attempt to read drawing relationships and map embedded images
                                        self._embedded_image_cid_by_name.setdefault(sheet.title, {})
                                        drawing_rels_path = os.path.dirname(drawing_path) + '/_rels/' + os.path.basename(drawing_path) + '.rels'
                                        rels_xml = get_xml_from_zip(z, drawing_rels_path)
                                        if rels_xml is not None:
                                            rid_to_target = {}
                                            for rel in rels_xml.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                                                rid = rel.attrib.get('Id') or rel.attrib.get('Id')
                                                tgt = rel.attrib.get('Target')
                                                if rid and tgt:
                                                    tgtp = normalize_excel_path(tgt)
                                                    rid_to_target[rid] = tgtp

                                            for node_c in list(drawing_xml):
                                                lname_c = node_c.tag.split('}')[-1].lower()
                                                if lname_c not in ('twocellanchor', 'onecellanchor'):
                                                    continue
                                                cid_val = extract_anchor_id(node_c, allow_idx=True)
                                                for sub in node_c.iter():
                                                    if sub.tag.split('}')[-1].lower() == 'blip':
                                                        rid = sub.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed') or sub.attrib.get('embed')
                                                        if rid and rid in rid_to_target:
                                                            target = rid_to_target[rid]
                                                            fname = os.path.basename(target)
                                                            try:
                                                                self._embedded_image_cid_by_name[sheet.title][fname] = str(cid_val) if cid_val is not None else None
                                                            except (ValueError, TypeError) as e:
                                                                debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")
                                                            break
                                    except (ValueError, TypeError):
                                        # non-fatal: ensure we have defaults
                                        anchors_cid_list = anchors_cid_list if 'anchors_cid_list' in locals() else []
                                        total_anchors = total_anchors if 'total_anchors' in locals() else 0
                                        pic_anchors = pic_anchors if 'pic_anchors' in locals() else 0
                                        sp_anchors = sp_anchors if 'sp_anchors' in locals() else 0
                                    
                                    debug_print(f"[DEBUG] Sheet '{sheet.title}': total_anchors={total_anchors}, pic_anchors={pic_anchors}, sp_anchors={sp_anchors}, sheet._images={len(sheet._images)}")
                                    debug_print(f"[DEBUG] Sheet '{sheet.title}': total_anchors={total_anchors}, pic_anchors={pic_anchors}, sp_anchors={sp_anchors}, sheet._images={len(sheet._images)}")
                                    
                                    # 埋め込み画像よりアンカーが多く、少なくとも1つの図形がある場合、
                                    # attempt isolated-group rendering to capture vector shapes
                                    debug_print(f"[DEBUG] Checking condition: total_anchors({total_anchors}) > len(sheet._images)({len(sheet._images)}) = {total_anchors > len(sheet._images)} AND sp_anchors({sp_anchors}) > 0 = {sp_anchors > 0}")
                                    if total_anchors > len(sheet._images) and sp_anchors > 0:
                                        debug_print(f"[DEBUG] Condition TRUE - entering isolated group rendering block for sheet '{sheet.title}'")
                                        debug_print(f"[DEBUG] Detected additional drawing shapes (anchors={total_anchors}, pics={pic_anchors}, sps={sp_anchors}) - attempting isolated-group rendering")
                                        try:
                                            # 図形のバウンディングボックスを抽出
                                            shapes = None
                                            try:
                                                debug_print(f"[DEBUG] Calling _extract_drawing_shapes for sheet '{sheet.title}'")
                                                shapes = self._extract_drawing_shapes(sheet)
                                                debug_print(f"[DEBUG] _extract_drawing_shapes returned {len(shapes) if shapes else 0} shapes")
                                            except Exception as shape_ex:
                                                print(f"[WARNING] _extract_drawing_shapes failed: {shape_ex}")
                                                import traceback
                                                traceback.print_exc()
                                            
                                            debug_print(f"[DEBUG] _extract_drawing_shapes returned: {len(shapes) if shapes else 'None'} shapes")
                                            debug_print(f"[DEBUG] Checking shapes: shapes={'Not None' if shapes else 'None'}, len={len(shapes) if shapes else 0}")
                                            if shapes and len(shapes) > 0:
                                                debug_print(f"[DEBUG] Shapes condition TRUE - entering clustering block")
                                                # 適切なクラスタリングロジックを使用して図形をクラスタリング
                                                # 行ベースのギャップ分割のためセル範囲を抽出
                                                try:
                                                    cell_ranges_all = self._extract_drawing_cell_ranges(sheet)
                                                except (ValueError, TypeError):
                                                    cell_ranges_all = []
                                                
                                                # 適切なクラスタリングのため_cluster_shapes_commonを使用
                                                # max_groups=1 means cluster into 1 group if possible (no splitting)
                                                # ただし、このメソッドは大きなギャップがある場合は分割します
                                                debug_print(f"[DEBUG] Calling _cluster_shapes_common with {len(shapes)} shapes")
                                                clusters, debug_info = self._cluster_shapes_common(
                                                    sheet, shapes, cell_ranges=cell_ranges_all, max_groups=1
                                                )
                                                debug_print(f"[DEBUG] _cluster_shapes_common returned {len(clusters)} clusters")
                                                debug_print(f"[DEBUG] clustered into {len(clusters)} groups: sizes={[len(c) for c in clusters]}")
                                                debug_print(f"[DEBUG] clustering debug_info: {debug_info}")
                                                
                                                # 各クラスタを分離グループとしてレンダリング
                                                # 安定した_render_sheet_isolated_groupメソッドを使用（v2ではない）
                                                # v2は実験的で不完全（コネクタの外観処理が欠落）
                                                isolated_produced = False
                                                isolated_images = []  # (filename, row)タプルのリスト
                                                debug_print(f"[DEBUG] Starting to render {len(clusters)} clusters for sheet '{sheet.title}'")
                                                for idx, cluster in enumerate(clusters):
                                                    if len(cluster) > 0:
                                                        debug_print(f"[DEBUG] Rendering cluster {idx+1}/{len(clusters)} with {len(cluster)} shapes")
                                                        result = self._render_sheet_isolated_group(sheet, cluster)
                                                        debug_print(f"[DEBUG] Cluster {idx+1} rendering result: {result}")
                                                        if result:
                                                            if isinstance(result, tuple) and len(result) == 2:
                                                                img_name, cluster_row = result
                                                            else:
                                                                img_name = result
                                                                cluster_row = 1
                                                            
                                                            isolated_produced = True
                                                            isolated_images.append((cluster_row, img_name))
                                                            print(f"[INFO] シート '{sheet.title}' のクラスタ {idx+1} をisolated groupとして出力: {img_name} (row={cluster_row})")
                                                
                                                if isolated_produced:
                                                    print(f"[INFO] シート '{sheet.title}' の図形をisolated groupとして出力しました")
                                                    debug_print(f"[DEBUG] isolated_images count: {len(isolated_images)}")
                                                    # isolated group画像をMarkdownに追加するため、images_foundをTrueに設定
                                                    images_found = True
                                                    # 各画像を登録（row情報を使用）
                                                    for cluster_row, img_name in isolated_images:
                                                        debug_print(f"[DEBUG] Processing isolated group image: {img_name} at row={cluster_row}")
                                                        try:
                                                            self._mark_image_emitted(img_name)
                                                            debug_print(f"[DEBUG] _mark_image_emitted succeeded for: {img_name}")
                                                        except Exception as e:
                                                            print(f"[WARNING] _mark_image_emitted failed: {e}")
                                                        
                                                        try:
                                                            # _sheet_shape_images に追加（クラスタの最小行を使用）
                                                            if not hasattr(self, '_sheet_shape_images'):
                                                                self._sheet_shape_images = {}
                                                            self._sheet_shape_images.setdefault(sheet.title, [])
                                                            # クラスタの最小行に配置
                                                            self._sheet_shape_images[sheet.title].append((cluster_row, img_name))
                                                            debug_print(f"[DEBUG] isolated group画像を_sheet_shape_imagesに追加: {img_name} at row={cluster_row}")
                                                            
                                                            if hasattr(self, '_last_iso_preserved_ids') and self._last_iso_preserved_ids:
                                                                self._image_shape_ids[img_name] = set(self._last_iso_preserved_ids)
                                                                debug_print(f"[DEBUG] 図形IDマッピングを保存: {img_name} -> {len(self._last_iso_preserved_ids)} shapes")
                                                        except Exception as e:
                                                            print(f"[WARNING] Failed to add to _sheet_shape_images: {e}")
                                                            import traceback
                                                            traceback.print_exc()
                                            else:
                                                isolated_produced = False
                                        except Exception as e:
                                            print(f"[WARNING] isolated-group rendering failed: {e}")
                                            import traceback
                                            traceback.print_exc()
                                            isolated_produced = False
                                        
                                        # end of drawing parsing block
                    except (ValueError, TypeError) as e:
                        debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")
            # パーサーで検出された画像が見つからなかった場合、保守的な
            # フォールバックを試行: LibreOffice経由でシートをPDFにレンダリングし、
            # ImageMagickを使用して対応するPDFページをPNGにラスタライズします。
            # これによりopenpyxlが画像として公開しないベクトル図形や描画をキャプチャします。
            if hasattr(sheet, '_images') and sheet._images:
                print(f"[INFO] シート '{sheet.title}' 内の画像を処理中...")
                images_found = True
                    # 埋め込みメディアからのマッピングを事前に設定（描画relsから）
                    # cNvPr IDへのマッピングを行い、下で埋め込み画像を処理する際に
                    # クラスタ化/グループレンダリングが既に同じ描画アンカーを
                    # 保持している場合に抑制するかどうかを決定できるようにします。
                try:
                    z = zipfile.ZipFile(self.excel_file)
                    sheet_index = self.workbook.sheetnames.index(sheet.title)
                    rels_path = f"xl/worksheets/_rels/sheet{sheet_index+1}.xml.rels"
                    rels_xml = get_xml_from_zip(z, rels_path)
                    if rels_xml is not None:
                        drawing_target = None
                        for rel in rels_xml.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                            t = rel.attrib.get('Type','')
                            if t.endswith('/drawing'):
                                drawing_target = rel.attrib.get('Target')
                                break
                        if drawing_target:
                            drawing_path = normalize_excel_path(drawing_target)
                            if drawing_path not in z.namelist():
                                drawing_path = drawing_path.replace('worksheets', 'drawings')
                            drawing_xml = get_xml_from_zip(z, drawing_path)
                            if drawing_xml is not None:
                                # ensure map exists
                                self._embedded_image_cid_by_name.setdefault(sheet.title, {})
                                # attempt to read drawing rels if present and map rId -> target
                                drawing_rels_path = os.path.dirname(drawing_path) + '/_rels/' + os.path.basename(drawing_path) + '.rels'
                                try:
                                    rels_xml2 = get_xml_from_zip(z, drawing_rels_path)
                                    if rels_xml2 is not None:
                                        rid_to_target = {}
                                        for rel2 in rels_xml2.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                                            rid = rel2.attrib.get('Id') or rel2.attrib.get('Id')
                                            tgt = rel2.attrib.get('Target')
                                            if rid and tgt:
                                                tgtp = normalize_excel_path(tgt)
                                                rid_to_target[rid] = tgtp
                                        # iterate anchors and map both media basename and media SHA8 -> cNvPr
                                        import hashlib as _hashlib
                                        for node_c in list(drawing_xml):
                                            lname_c = node_c.tag.split('}')[-1].lower()
                                            if lname_c not in ('twocellanchor', 'onecellanchor'):
                                                continue
                                            cid_val = extract_anchor_id(node_c, allow_idx=True)
                                            for sub in node_c.iter():
                                                if sub.tag.split('}')[-1].lower() == 'blip':
                                                    rid = sub.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed') or sub.attrib.get('embed')
                                                    if rid and rid in rid_to_target:
                                                        target = rid_to_target[rid]
                                                        # normalize path
                                                        tgtp = normalize_excel_path(target)
                                                        # extract basename
                                                        fname = os.path.basename(tgtp)
                                                        try:
                                                            media_bytes = z.read(tgtp) if tgtp in z.namelist() else None
                                                        except Exception:
                                                            media_bytes = None
                                                        sha8 = None
                                                        if media_bytes:
                                                            try:
                                                                sha8 = _hashlib.sha1(media_bytes).hexdigest()[:8]
                                                            except Exception:
                                                                sha8 = None
                                                        if cid_val is not None:
                                                            try:
                                                                # map by original basename
                                                                self._embedded_image_cid_by_name[sheet.title][fname] = str(cid_val)
                                                            except (ValueError, TypeError):
                                                                pass  # データ構造操作失敗は無視
                                                            try:
                                                                # map by short sha if available
                                                                if sha8:
                                                                    self._embedded_image_cid_by_name[sheet.title][sha8] = str(cid_val)
                                                            except (ValueError, TypeError):
                                                                pass  # データ構造操作失敗は無視
                                                        else:
                                                            try:
                                                                self._embedded_image_cid_by_name[sheet.title][fname] = None
                                                            except (ValueError, TypeError):
                                                                pass  # データ構造操作失敗は無視
                                                            try:
                                                                if sha8:
                                                                    self._embedded_image_cid_by_name[sheet.title][sha8] = None
                                                            except Exception as e:
                                                                print(f"[WARNING] ファイル操作エラー: {e}")
                                except Exception as e:
                                    print(f"[WARNING] ファイル操作エラー: {e}")
                except Exception as e:
                    print(f"[WARNING] ファイル操作エラー: {e}")
                md_lines = []
                for image in sheet._images:
                    # _process_excel_image now returns the saved image filename (basename)
                    img_name = self._process_excel_image(image, f"{sheet.title} (Image)")
                    if img_name:
                            # この画像の代表的なstart_rowを決定（利用可能な場合）
                            start_row = 1
                            try:
                                pos = None
                                try:
                                    pos = self._get_image_position(image)
                                except Exception:
                                    pos = None
                                if pos and isinstance(pos, dict):
                                    # prefer explicit row if provided
                                    if 'row' in pos and isinstance(pos['row'], int):
                                        start_row = pos['row']
                            except Exception:
                                start_row = 1

                            # 正規出力パス中の場合、即座に挿入し
                            # the image appears inline with emitted text. Otherwise,
                            # defer by registering into self._sheet_shape_images so the
                            # canonical emission will place it deterministically.
                            if getattr(self, '_in_canonical_emit', False):
                                md_line = f"![{sheet.title}の図](images/{img_name})"
                                ref = f"images/{img_name}"
                                # この埋め込み画像が描画アンカーに対応する場合
                                # that has already been preserved by a grouped render,
                                # skip emitting it to avoid duplicate presentation.
                                try:
                                    cid_map = self._embedded_image_cid_by_name.get(sheet.title, {}) if hasattr(self, '_embedded_image_cid_by_name') else {}
                                    mapped_cid = cid_map.get(img_name)
                                    # ファイル名に_<sha8>.extのような短いハッシュサフィックスが含まれる場合、それを抽出してキーとして試行
                                    if mapped_cid is None:
                                        try:
                                            # try extracting trailing 8-hex from filename
                                            import re
                                            m = re.search(r'([0-9a-f]{8})', img_name)
                                            if m:
                                                maybe = m.group(1)
                                                mapped_cid = cid_map.get(maybe)
                                        except Exception as e:
                                            print(f"[WARNING] ファイル操作エラー: {e}")
                                    # まだ不明な場合、ディスク上の既存ファイルから短いshaを計算して試行
                                    if mapped_cid is None:
                                        try:
                                            fp = os.path.join(self.images_dir, img_name)
                                            if os.path.exists(fp):
                                                import hashlib as _hashlib
                                                with open(fp, 'rb') as _f:
                                                    d = _f.read()
                                                sha8 = _hashlib.sha1(d).hexdigest()[:8]
                                                mapped_cid = cid_map.get(sha8)
                                        except (OSError, IOError, FileNotFoundError):
                                            print(f"[WARNING] ファイル操作エラー: {e if 'e' in locals() else '不明'}")
                                    global_iso_preserved_ids = getattr(self, '_global_iso_preserved_ids', set()) or set()
                                    if mapped_cid and str(mapped_cid) in global_iso_preserved_ids:
                                        debug_print(f"[DEBUG][_emit_image_skip] sheet={sheet.title} embedded image {img_name} suppressed (cid={mapped_cid} already preserved)")
                                        continue
                                except (OSError, IOError, FileNotFoundError):
                                    print(f"[WARNING] ファイル操作エラー: {e if 'e' in locals() else '不明'}")
                                if ref in self._emitted_images or img_name in self._emitted_images:
                                    continue
                                try:
                                    new_idx = self._insert_markdown_image(insert_index, md_line, img_name, sheet=sheet)
                                    try:
                                        if insert_index is not None:
                                            insert_index = new_idx
                                    except Exception as e:
                                        print(f"[WARNING] ファイル操作エラー: {e}")
                                except Exception:
                                    try:
                                        self.markdown_lines.append(md_line)
                                        self.markdown_lines.append("")
                                        try:
                                            self._mark_image_emitted(img_name)
                                        except Exception as e:
                                            print(f"[WARNING] ファイル操作エラー: {e}")
                                    except Exception as e:
                                        print(f"[WARNING] ファイル操作エラー: {e}")
                                # 挿入を延期: 正規の行ソート済み出力のため登録
                                try:
                                    # check mapped cNvPr for this embedded image and
                                    # skip deferral if already preserved by a group render
                                    cid_map = self._embedded_image_cid_by_name.get(sheet.title, {}) if hasattr(self, '_embedded_image_cid_by_name') else {}
                                    mapped_cid = cid_map.get(img_name)
                                    global_iso_preserved_ids = getattr(self, '_global_iso_preserved_ids', set()) or set()
                                    if mapped_cid and str(mapped_cid) in global_iso_preserved_ids:
                                        debug_print(f"[DEBUG][_defer_image_skip] sheet={sheet.title} embedded image {img_name} suppressed on defer (cid={mapped_cid} already preserved)")
                                    else:
                                        self._sheet_shape_images.setdefault(sheet.title, [])
                                        self._sheet_shape_images[sheet.title].append((start_row, img_name))
                                except (ValueError, TypeError) as e:
                                    debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")
                            else:
                                # 非正規コンテキスト: 画像を登録/延期し
                                # canonical emitter will place it deterministically.
                                try:
                                    cid_map = self._embedded_image_cid_by_name.get(sheet.title, {}) if hasattr(self, '_embedded_image_cid_by_name') else {}
                                    mapped_cid = cid_map.get(img_name)
                                    if mapped_cid is None:
                                        try:
                                            import re
                                            m = re.search(r'([0-9a-f]{8})', img_name)
                                            if m:
                                                maybe = m.group(1)
                                                mapped_cid = cid_map.get(maybe)
                                        except Exception as e:
                                            print(f"[WARNING] ファイル操作エラー: {e}")
                                    if mapped_cid is None:
                                        try:
                                            fp = os.path.join(self.images_dir, img_name)
                                            if os.path.exists(fp):
                                                import hashlib as _hashlib
                                                with open(fp, 'rb') as _f:
                                                    d = _f.read()
                                                sha8 = _hashlib.sha1(d).hexdigest()[:8]
                                                mapped_cid = cid_map.get(sha8)
                                        except (OSError, IOError, FileNotFoundError):
                                            print(f"[WARNING] ファイル操作エラー: {e if 'e' in locals() else '不明'}")
                                    global_iso_preserved_ids = getattr(self, '_global_iso_preserved_ids', set()) or set()
                                    if mapped_cid and str(mapped_cid) in global_iso_preserved_ids:
                                        debug_print(f"[DEBUG][_noncanonical_image_skip] sheet={sheet.title} embedded image {img_name} suppressed (cid={mapped_cid} already preserved)")
                                        continue
                                    
                                    md_line = f"![{sheet.title}の図](images/{img_name})"
                                    new_idx = self._insert_markdown_image(insert_index, md_line, img_name, sheet=sheet)
                                    try:
                                        if insert_index is not None:
                                            insert_index = new_idx
                                    except Exception as e:
                                        print(f"[WARNING] ファイル操作エラー: {e}")
                                except Exception:
                                    # フォールバック: sheet_shape_imagesに直接登録
                                    try:
                                        self._sheet_shape_images.setdefault(sheet.title, [])
                                        self._sheet_shape_images[sheet.title].append((start_row, img_name))
                                    except Exception as e:
                                        print(f"[WARNING] ファイル操作エラー: {e}")

            if not images_found:
                debug_print(f"[DEBUG] イメージが見つかりませんでした。")
                # セルテキストのみを含むシートのレンダリングを回避。シートに
                # 描画要素がある場合のみフォールバックします。
                if not self._sheet_has_drawings(sheet):
                    return False
                # 重いPDF->PNGパイプラインを起動する前に、描画XMLから
                # 描画バウンディングボックスの抽出を試みます。XML解析が
                # 空のリストを返した場合、レンダリングする可視図形がない可能性が高く、
                # フルページ画像の生成をスキップする必要があります。
                shapes = None
                try:
                    shapes = self._extract_drawing_shapes(sheet)
                except Exception:
                    shapes = None

                # 抽出が成功し空のリストを返した場合、描画可能な要素がない時に
                # フルシートラスターを挿入することを避けるためフォールバックをスキップします。
                # 抽出がエラー（shapesがNone）または非空を返した場合、
                # 以前と同様にレンダリングを続行します。
                if shapes == []:
                    print(f"[INFO] シート '{sheet.title}' に描画要素が見つかりませんでした（XML解析結果）。フォールバックレンダリングをスキップします。")
                    return False

                print(f"[INFO] シート '{sheet.title}' に検出されたラスタ画像がありません。フォールバックレンダリングを試行します...")
                try:
                    # シートレベルの図形画像を生成（images_dirに保存されます）
                    rendered = self._render_sheet_fallback(sheet, insert_index=insert_index, insert_images=insert_images)
                    if rendered:
                        # mark shapes as generated for this sheet
                        self._sheet_shapes_generated.add(sheet.title)
                        # initialize next index
                        if sheet.title not in self._sheet_shape_next_idx:
                            self._sheet_shape_next_idx[sheet.title] = 0
                        # 図形が作成された場合、markdownのinsert_index（テーブル末尾）に挿入します。
                        try:
                            imgs = self._sheet_shape_images.get(sheet.title, [])
                            if imgs:
                                # 行順序のマージに基づいて図形を挿入することを優先し、
                                # テキストと画像が最終的なMarkdownでExcelシートと
                                # 同じ上から下への順序で表示されるようにします。
                                imgs_by_row = {}
                                assigned = self._sheet_shape_images.get(sheet.title, []) or []

                                # assigned may be list of pairs or list of filenames (backcompat)
                                normalized = []
                                for item in assigned:
                                    if isinstance(item, (list, tuple)) and len(item) >= 2:
                                        try:
                                            row_key = int(item[0]) if item[0] is not None else 1
                                        except (ValueError, TypeError):
                                            row_key = 1
                                        normalized.append((row_key, item[1]))
                                    else:
                                        # fallback: treat as filename with default row=1
                                        try:
                                            normalized.append((1, str(item)))
                                        except (ValueError, TypeError) as e:
                                            debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")

                                # 小さな調整を許可: 代表的なstart_rowが既存のテキストアンカーの
                                # 近く（SNAP_DIST行以内）にある場合、グループ画像を近くの
                                # ヘッダーに隣接させるためにその画像をそのアンカーにスナップします。
                                # 同点の場合は早い行を優先します。
                                SNAP_DIST = getattr(self, '_anchor_snap_distance', 3)
                                # スナッピング用のsheet->mdマッピングが利用可能であることを確認
                                sheet_map = self._cell_to_md_index.get(sheet.title, {}) or {}
                                # スナッピング用にソートされたテキスト行を事前計算
                                try:
                                    text_rows_sorted = sorted(list(sheet_map.keys()))
                                except Exception:
                                    text_rows_sorted = []
                                for r, img in normalized:
                                    adjusted_row = r
                                    try:
                                        if text_rows_sorted:
                                            # 最も近いテキスト行を見つけ、SNAP_DIST以内ならスナップ
                                            nearest = min(text_rows_sorted, key=lambda tr: (abs(tr - r), tr))
                                            if abs(nearest - r) <= SNAP_DIST:
                                                adjusted_row = nearest
                                    except Exception:
                                        pass  # データ構造操作失敗は無視
                                    imgs_by_row.setdefault(adjusted_row, []).append(img)

                                # このシートの既存のtext->mdマッピングを取得
                                # sheet_mapは上で既に定義済み。再利用（または新規取得）
                                sheet_map = self._cell_to_md_index.get(sheet.title, {}) or sheet_map

                                # 注: レガシーコードは永続化されたstart_map（self._sheet_shape_image_start_rows）を
                                # 使用して実行間で画像挿入行を再割り当てしていました。そのロジックは
                                # 複数の異なるグループ画像を単一の挿入バケットに折りたたむ可能性がありました。
                                # normalizedに格納された新しく計算された代表的なstart_row値を優先し、
                                # ここでは永続化されたstart_mapを参照しません。
                                if hasattr(self, '_sheet_shape_image_start_rows') and self._sheet_shape_image_start_rows.get(sheet.title):
                                    # このシートの永続化されたヒントをクリアして、
                                    # 生成したばかりの計算されたstart_rowペアを上書きしないようにします。
                                    try:
                                        # log for diagnostics but do not use it
                                        debug_print(f"[DEBUG] Ignoring persisted start_map for sheet={sheet.title}")
                                    except (ValueError, TypeError) as e:
                                        debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")

                                # 反復する適切な行セットを決定（テキスト行と画像行の和集合）
                                rows = sorted(set(list(sheet_map.keys()) + list(imgs_by_row.keys())))

                                # 診断デバッグ: ソース行 -> markdownインデックスの現在のマッピングを出力
                                try:
                                    debug_print(f"[DEBUG][_img_insertion_debug] sheet={sheet.title} sheet_map={sheet_map}")
                                    debug_print(f"[DEBUG][_img_insertion_debug] imgs_by_row={imgs_by_row}")
                                    debug_print(f"[DEBUG][_img_insertion_debug] normalized_pairs={normalized}")
                                except (ValueError, TypeError) as e:
                                    debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")

                                # 行にマップされたテキストが存在しない場合の基本挿入インデックス
                                insert_base = insert_index if insert_index is not None else len(self.markdown_lines)

                                # 最近傍マッピング用にソートされたテキスト行を事前計算
                                text_rows_sorted = sorted(list(sheet_map.keys()))

                                # 各画像行について、代表的なstart_rowを最もよく反映する挿入ポイントを選択:
                                # start_row以上の最初のテキスト行を優先（画像がそのテキストブロックの
                                # 直前に表示されるように）。そのような行が存在しない場合、
                                # start_row未満の最後のテキスト行を優先（その後に挿入）。
                                # シートにテキストマッピングがまったくない場合、start_rowの昇順で
                                # insert_baseに順次挿入するフォールバック。これにより、グループ画像は
                                # 同じテキストアンカーに真にマップされない限り、同じ挿入バケットに
                                # 折りたたまれることなく、代表的なstart_rowに保持されます。
                                # 検証用の最終挿入マッピングを収集: md_index -> [filenames]
                                md_index_map = {}
                                for row_num in sorted(imgs_by_row.keys()):
                                    imgs_for_row = imgs_by_row.get(row_num, [])
                                    # 候補テキストアンカーを決定
                                    md_pos = None
                                    if row_num in sheet_map:
                                        md_pos = sheet_map.get(row_num)
                                        insert_at = md_pos + 1 if md_pos is not None else insert_base
                                    else:
                                        # この画像のstart_rowに最も近いテキストアンカーを選択。
                                        # 最も近いアンカーを使用することで、論理的なアンカーが
                                        # 近くのヘッダー（例: row3）である場合に、遠い後のブロック
                                        # （例: row26）にバインドすることを回避します。同点の場合は早い行を優先。
                                        if text_rows_sorted:
                                            try:
                                                nearest = min(text_rows_sorted, key=lambda tr: (abs(tr - row_num), tr))
                                                md_pos = sheet_map.get(nearest)
                                                insert_at = (md_pos + 1) if md_pos is not None else insert_base
                                            except Exception:
                                                insert_at = insert_base
                                        else:
                                            # テキストマッピングなし。insert_baseに順次追加
                                            insert_at = insert_base

                                    # insert_atをクランプ to valid markdown range
                                    if insert_at < 0:
                                        insert_at = 0
                                    if insert_at > len(self.markdown_lines):
                                        insert_at = len(self.markdown_lines)

                                    # このグループ用に特に選択したテキストアンカー（md_pos）の前に
                                    # 画像を挿入しないようにします。以前のグローバルクランプ
                                    # （最新のアンカーインデックスを使用）は、無関係な後のアンカーの
                                    # 後に画像を移動させ、画像が論理的なテキストコンテキストから
                                    # 遠く離れて表示される可能性がありました。この画像に使用した
                                    # アンカー（存在する場合）に対してのみ最小値を強制します。
                                    try:
                                        if md_pos is not None:
                                            # md_posは選択されたアンカーのmarkdownインデックス
                                            # insert_atは少なくともその1行後である必要があります。
                                            if insert_at <= md_pos:
                                                insert_at = md_pos + 1
                                    except Exception:
                                        # 保守的なフォールバック: insert_atを変更しない
                                        pass

                                    # この行の各画像を挿入し、元の相対順序を保持
                                    for img in imgs_for_row:
                                        if not insert_images:
                                            # 呼び出し元が遅延挿入を要求した場合、マッピングを記録するだけ
                                            md_index_map.setdefault(row_num, []).append(img)
                                            continue
                                        ref = f"images/{img}"
                                        already = any(ref in (ln or '') for ln in self.markdown_lines)
                                        if already:
                                            continue
                                        md = f"![{sheet.title}](images/{img})"
                                        # ヘルパーを使用して挿入し、出力済みとしてマーク
                                        try:
                                            new_at = self._insert_markdown_image(insert_at, md, img, sheet=sheet)
                                            md_index_map.setdefault(insert_at, []).append(img)
                                            insert_at = new_at
                                        except Exception:
                                            # fallback
                                            try:
                                                self.markdown_lines.append(md)
                                                self.markdown_lines.append("")
                                                self._mark_image_emitted(img)
                                            except Exception as e:
                                                print(f"[WARNING] ファイル操作エラー: {e}")

                                    # グローバルinsert_base位置に挿入した場合、それを進める
                                    if (row_num not in sheet_map) and insert_at > insert_base:
                                        insert_base = insert_at

                                    # それに依存する後続の挿入のためにsheet_mapオフセットを更新
                                    if sheet_map and md_pos is not None:
                                        # 正規出力時のみ既存のsheet_mapオフセットを更新
                                        if getattr(self, '_in_canonical_emit', False):
                                            for k, v in list(sheet_map.items()):
                                                try:
                                                    if v > (md_pos if md_pos is not None else -1):
                                                        # 使用したばかりのアンカーは更新しない
                                                        if k != (row_num if row_num in sheet_map else None):
                                                            # マッピングを新しいインデックスに更新
                                                            self._mark_sheet_map(sheet.title, k, v + 2 * len(imgs_for_row))
                                                except Exception as e:
                                                    pass  # XML解析エラーは無視
                                        else:
                                            debug_print(f"[TRACE] Skipping sheet_map offset updates in non-canonical pass for sheet={sheet.title}")

                                # すべての画像を使用済みとしてマーク
                                self._sheet_shape_next_idx[sheet.title] = len(imgs)
                                # このシートの最終挿入マッピングをログ出力（存在する場合）
                                try:
                                        if md_index_map:
                                            print(f"[INFO][_final_img_map] sheet={sheet.title} insert_mappings={md_index_map}")
                                except (ValueError, TypeError) as e:
                                    debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")
                        except (ValueError, TypeError):
                            # 何か失敗した場合、以前の単純な挿入にフォールバック
                            try:
                                if insert_index is not None:
                                    insert_at = insert_index
                                    for item in imgs:
                                        # itemはファイル名（str）または(row, filename)ペアの可能性あり
                                        if isinstance(item, (list, tuple)) and len(item) >= 2:
                                            img_fn = str(item[1])
                                        else:
                                            img_fn = str(item)
                                        ref = f"images/{img_fn}"
                                        already = any(ref in (ln or '') for ln in self.markdown_lines)
                                        if already:
                                            continue
                                        md = f"![{sheet.title}](images/{img_fn})"
                                        try:
                                            new_at = self._insert_markdown_image(insert_at, md, img_fn)
                                            try:
                                                insert_at = new_at
                                            except (ValueError, TypeError) as e:
                                                debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")
                                        except Exception:
                                            try:
                                                self.markdown_lines.append(md)
                                                self.markdown_lines.append("")
                                                self._mark_image_emitted(img_fn)
                                            except Exception as e:
                                                print(f"[WARNING] ファイル操作エラー: {e}")
                                    # 保存された画像（ファイル名）の数として次のインデックスを記録
                                    try:
                                        self._sheet_shape_next_idx[sheet.title] = len(imgs)
                                    except Exception as e:
                                        print(f"[WARNING] ファイル操作エラー: {e}")
                            except Exception:
                                self._sheet_shape_next_idx[sheet.title] = len(imgs)
                    else:
                        print(f"[WARNING] フォールバックレンダリングが実行されませんでした（外部ツール未検出など）。")
                except Exception as e:
                    print(f"[WARNING] フォールバックレンダリング中にエラーが発生しました: {e}")
                    
        except Exception as e:
            print(f"[WARNING] 画像処理エラー: {e}")
            return False
        return True
    
    def _process_excel_image(self, image, sheet_name: str) -> Optional[str]:
        """Excel画像を処理"""
        try:
            self.image_counter += 1
            
            # 画像データを取得
            image_data = None
            # 利用可能な場合はimage._data()を優先。ただし、openpyxlは
            # 親のZipFileが閉じられると無効になるZipExtFileを作成している
            # 可能性があります。PILはValueError: I/O operation on closed fileを
            # 発生させます。それを検出し、可能な場合はimage.ref（パスまたは
            # ベース名）を使用してXLSX zipからメディアバイトを直接読み取る
            # フォールバックを行います。
            if hasattr(image, '_data') and callable(getattr(image, '_data')):
                try:
                    image_data = image._data()
                    debug_print(f"[DEBUG] image._data() succeeded for image #{self.image_counter} on sheet '{sheet_name}'")
                except ValueError:
                    # 閉じられたZipExtFileの可能性が高い。zipベースのフォールバックにフォールスルー。
                    image_data = None
                except (ValueError, TypeError):
                    image_data = None

            if image_data is None:
                # パスのように見える場合、image.refを使用してワークブックZIPから読み込みを試行
                try:
                    ref = getattr(image, 'ref', None)
                    if isinstance(ref, bytes):
                        try:
                            ref = ref.decode('utf-8')
                        except Exception:
                            ref = None
                    if isinstance(ref, str) and ref:
                        ref_path = ref.lstrip('/') if ref.startswith('/') else ref
                        try:
                            z = zipfile.ZipFile(self.excel_file, 'r')
                            try:
                                # まず直接一致を試行
                                if ref_path in z.namelist():
                                    image_data = z.read(ref_path)
                                else:
                                    # ベース名で一致を試行
                                    import os as _os
                                    b = _os.path.basename(ref_path)
                                    for nm in z.namelist():
                                        if nm.endswith('/' + b) or nm == b:
                                            image_data = z.read(nm)
                                            break
                            finally:
                                try:
                                    z.close()
                                except Exception as e:
                                    print(f"[WARNING] ファイル操作エラー: {e}")
                        except Exception:
                            image_data = None
                    # refがstrでない場合はimage_dataはNone
                    debug_print(f"[DEBUG] image.ref-based extraction succeeded for image #{self.image_counter} on sheet '{sheet_name}'")
                except (ValueError, TypeError):
                    image_data = None

            if not image_data:
                print("[WARNING] 画像データを取得できませんでした")
                return
            
            # 画像形式を判定
            extension = self._detect_image_format(image_data)
            
            # ファイル名を生成（安全化）
            safe_sheet_name = self._sanitize_filename(sheet_name)
            # 画像バイトに基づく決定論的なファイル名を使用して、同じワークブックの
            # 繰り返し変換で新しいファイルが生成されないようにします。
            # 画像バイトの短いSHA1を計算します。
            try:
                import hashlib
                h = hashlib.sha1(image_data).hexdigest()[:8]
                image_filename = f"{self.base_name}_{safe_sheet_name}_image_{h}{extension}"
            except Exception:
                # シートレベルの安定した名前にフォールバック
                image_filename = f"{self.base_name}_{safe_sheet_name}_image{extension}"
            image_path = os.path.join(self.images_dir, image_filename)
            
            # 画像を保存
            # このコンテンツハッシュを持つファイルが既に存在する場合、再書き込みを回避します。
            try:
                if os.path.exists(image_path):
                    # 既存のファイル内容が一致するか確認。一致する場合は再利用。
                    try:
                        with open(image_path, 'rb') as ef:
                            existing = ef.read()
                        if existing == image_data:
                            # reuse
                            debug_print(f"[DEBUG] 既存の画像ファイルを再利用: {image_filename}")
                        else:
                            # 衝突は稀。一意のサフィックスにフォールバック
                            import time
                            alt = f"_{int(time.time())}"
                            image_filename = f"{self.base_name}_{safe_sheet_name}_image_{h}{alt}{extension}"
                            image_path = os.path.join(self.images_dir, image_filename)
                            with open(image_path, 'wb') as f:
                                f.write(image_data)
                    except (OSError, IOError, FileNotFoundError):
                        with open(image_path, 'wb') as f:
                            f.write(image_data)
                else:
                    with open(image_path, 'wb') as f:
                        f.write(image_data)
            except (OSError, IOError, FileNotFoundError):
                # 最後の手段として書き込み
                with open(image_path, 'wb') as f:
                    f.write(image_data)
            
            # 画像位置情報を取得
            position_info = self._get_image_position(image)
            
            # 保存された画像ファイル名（ベース名）を返します。呼び出し元は
            # この具体的なファイル名を使用してmarkdownを生成し、リンクが
            # 常にディスク上の既存ファイルを指すようにします。
            print(f"[SUCCESS] 画像を処理: {image_filename}")
            return os.path.basename(image_filename)
        except Exception as e:
            try:
                import traceback
                tb = traceback.format_exc()
                print(f"[ERROR] Excel画像処理エラー: {e}\n{tb}")
            except (ValueError, TypeError):
                print(f"[ERROR] Excel画像処理エラー: {e}")
            return None

    def _deduplicate_image_files(self):
        """output/images のファイルを内容でグループ化し、同一内容の複数ファイルを1つに集約する。

        - 同一ハッシュのファイル群から最も短い名前（またはアルファベット順先頭）を正規名とする
        - markdown_lines 内の参照を正規名に置換する
        - output/sorted_events.txt 内の image 行を正規名で置換する
        - 重複ファイルは削除する
        """
        try:
            import hashlib
            import collections
            img_dir = self.images_dir
            if not os.path.isdir(img_dir):
                return

            # デバッグ: グループ構築前に画像ファイル名とそのSHA256をログ出力
            try:
                import hashlib as _hashlib
                debug_print('[DEBUG][_dedupe] listing images and computing sha256 before dedupe:')
                for _fn in sorted(os.listdir(img_dir)):
                    _fp = os.path.join(img_dir, _fn)
                    if not os.path.isfile(_fp):
                        continue
                    try:
                        _h = _hashlib.sha256()
                        with open(_fp, 'rb') as _f:
                            for _chunk in iter(lambda: _f.read(8192), b''):
                                _h.update(_chunk)
                        debug_print(f"[DEBUG][_dedupe] pre-sha {_fn} = {_h.hexdigest()}")
                    except Exception as _e:
                        debug_print(f"[DEBUG][_dedupe] pre-sha {_fn} FAILED: {_e}")
            except (OSError, IOError, FileNotFoundError):
                # 致命的ではない。通常の重複排除を続行
                pass

            # ハッシュ -> [ファイル] を構築
            groups = collections.defaultdict(list)
            for fn in os.listdir(img_dir):
                fp = os.path.join(img_dir, fn)
                if not os.path.isfile(fp):
                    continue
                try:
                    with open(fp, 'rb') as f:
                        data = f.read()
                    h = hashlib.sha256(data).hexdigest()
                    groups[h].append((fn, data))
                except (OSError, IOError, FileNotFoundError):
                    continue

            # 1つ以上のファイルを持つ各グループについて、正規名を選択し参照を更新
            for h, items in groups.items():
                if len(items) <= 1:
                    continue
                # 正規ファイル名を選択: 最短を優先、次に辞書順
                items_sorted = sorted(items, key=lambda it: (len(it[0]), it[0]))
                canonical = items_sorted[0][0]
                duplicate_names = [it[0] for it in items_sorted[1:]]
                if not duplicate_names:
                    continue

                # このハッシュグループ内のすべてのファイルが同じワークブック
                # （self.base_name）から発生しているかどうかを判定。そうでない場合、
                # 異なるExcelファイルからの画像を別々に扱うというユーザーの
                # 要件を尊重するため、このグループの重複排除をスキップします。
                try:
                    bases = set([fn.split('_', 1)[0] if '_' in fn else fn for fn, _ in items_sorted])
                    if len(bases) != 1 or (self.base_name not in bases):
                        debug_print(f"[DEBUG][_dedupe] skipping cross-workbook dedupe for hash {h}: bases={bases}")
                        # このグループのファイルは削除しない。そのまま残す
                        continue
                except (ValueError, TypeError):
                    # 発生元の判定に失敗した場合、保守的にスキップ
                    debug_print(f"[DEBUG][_dedupe] skipping dedupe for hash {h} due to error determining origins")
                    continue

                # markdown_lines参照を更新（このワークブックに属するファイルのみ）
                try:
                    import re
                    new_lines = []
                    for ln in self.markdown_lines:
                        if not isinstance(ln, str):
                            new_lines.append(ln)
                            continue
                        s = ln
                        for dup in duplicate_names:
                            s = re.sub(r"!\[(.*?)\]\(images/" + re.escape(dup) + r"\)", r"![\1](images/" + canonical + r")", s)
                        new_lines.append(s)
                    self.markdown_lines = ExcelToMarkdownConverter._LoggingList(self)
                    self.markdown_lines += new_lines
                except Exception as e:
                    print(f"[WARNING] ファイル操作エラー: {e}")

                # 重複ファイルを削除（正規ファイルを保持）
                for dup in duplicate_names:
                    try:
                        p = os.path.join(img_dir, dup)
                        if os.path.exists(p):
                            os.remove(p)
                            debug_print(f"[DEBUG][_dedupe] removed duplicate image: {dup} -> canonical: {canonical}")
                    except (ValueError, TypeError):
                        pass  # ファイル操作失敗は無視

            # 最終ファイル名を反映するためにemitted imagesセットも再構築
            try:
                self._emitted_images = set()
                for ln in self.markdown_lines:
                    try:
                        import re
                        m = re.search(r"!\[.*?\]\(images/([^\)]+)\)", ln or "")
                        if m:
                            self._emitted_images.add(m.group(1))
                    except Exception:
                        continue
            except Exception as e:
                print(f"[WARNING] ファイル操作エラー: {e}")
        except Exception as e:
            print(f"[WARNING] ファイル操作エラー: {e}")

    # ========================================================================
    # Phase 1: 画像・図形処理の共通基盤メソッド群
    # ========================================================================

    def _get_drawing_xml_and_metadata(self, sheet) -> Optional[Dict[str, Any]]:
        """シートのdrawing.xmlとメタデータを取得
        
        ExcelファイルをZIPとして開き、指定シートのdrawing.xmlを取得します。
        3つのレンダリングメソッドで重複していた処理を統合。
        
        Args:
            sheet: 対象シート
        
        Returns:
            {
                'zip': ZipFile,
                'drawing_xml': ET.Element,
                'drawing_path': str,
                'sheet_index': int
            }
            または None (drawing.xmlが存在しない場合)
        
        Note:
            呼び出し側で返されたzip_fileをcloseする責任があります
        """
        try:
            zpath = self.excel_file
            z = zipfile.ZipFile(zpath, 'r')
            sheet_index = self.workbook.sheetnames.index(sheet.title)
            rels_path = f"xl/worksheets/_rels/sheet{sheet_index+1}.xml.rels"
            
            rels_xml = get_xml_from_zip(z, rels_path)
            if rels_xml is None:
                z.close()
                return None
            drawing_target = None
            for rel in rels_xml.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                if rel.attrib.get('Type', '').endswith('/drawing'):
                    drawing_target = rel.attrib.get('Target')
                    break
            
            if not drawing_target:
                z.close()
                return None
            
            # drawing_pathの正規化
            drawing_path = normalize_excel_path(drawing_target)
            if drawing_path not in z.namelist():
                drawing_path = drawing_path.replace('worksheets', 'drawings')
                if drawing_path not in z.namelist():
                    z.close()
                    return None
            
            drawing_xml_bytes = z.read(drawing_path)
            drawing_xml = ET.fromstring(drawing_xml_bytes)
            
            return {
                'zip': z,
                'drawing_xml': drawing_xml,
                'drawing_path': drawing_path,
                'sheet_index': sheet_index
            }
        
        except Exception as e:
            print(f"[WARNING] Drawing XML取得失敗: {e}")
            try:
                z.close()
            except:
                pass  # データ構造操作失敗は無視
            return None

    def _parse_theme_colors(self, zip_file: zipfile.ZipFile) -> Tuple[Dict[str, str], Dict[str, Any]]:
        """theme1.xmlからカラースキームとline参照を抽出
        
        ExcelのテーマXMLを解析し、色スキーム(schemeClr -> srgbClr)と
        lnRef(線のスタイル参照)のマッピングを取得します。
        2つのレンダリングメソッドで重複していた処理を統合。
        
        Args:
            zip_file: 開かれたZipFileオブジェクト
        
        Returns:
            (theme_color_map, ln_ref_map)
            - theme_color_map: {color_name: hex_value}
            - ln_ref_map: {index: ln_element}
        """
        theme_color_map = {}
        ln_ref_map = {}
        
        try:
            if 'xl/theme/theme1.xml' not in zip_file.namelist():
                return theme_color_map, ln_ref_map
            
            theme_xml = ET.fromstring(zip_file.read('xl/theme/theme1.xml'))
            a_ns = 'http://schemas.openxmlformats.org/drawingml/2006/main'
            
            # カラースキームの抽出
            clr = None
            for node in theme_xml.iter():
                if node.tag.split('}')[-1].lower() == 'clrscheme':
                    clr = node
                    break
            
            if clr is not None:
                for child in list(clr):
                    name = child.tag.split('}')[-1]
                    hexval = None
                    for sub in child.iter():
                        tag_name = sub.tag.split('}')[-1].lower()
                        if tag_name == 'srgbclr':
                            hexval = sub.attrib.get('val')
                            break
                        if tag_name == 'sysclr':
                            hexval = sub.attrib.get('lastclr') or sub.attrib.get('lastClr')
                            break
                    if hexval:
                        theme_color_map[name.lower()] = hexval
            
            # lnStyleLstの抽出
            try:
                ns = {'a': a_ns}
                ln_style_lst = theme_xml.find('.//{http://schemas.openxmlformats.org/drawingml/2006/main}lnStyleLst')
                if ln_style_lst is None:
                    ln_style_lst = theme_xml.find('.//a:lnStyleLst', ns)
                
                if ln_style_lst is not None:
                    import copy as _copy
                    for i, ln_el in enumerate(list(ln_style_lst)):
                        try:
                            # 属性も含めてディープコピー
                            ln_ref_map[str(i)] = _copy.deepcopy(ln_el)
                        except (ValueError, TypeError):
                            ln_ref_map[str(i)] = None
            except (ValueError, TypeError):
                pass  # データ構造操作失敗は無視
        
        except Exception as e:
            print(f"[WARNING] テーマカラー解析失敗: {e}")
        
        return theme_color_map, ln_ref_map

    def _resolve_connector_references(
        self,
        drawing_xml: ET.Element,
        anchors: List[ET.Element],
        keep_cnvpr_ids: Set[str]
    ) -> Set[str]:
        """
        BFSを使用してコネクタ参照を解決し、保持するアンカーIDの完全なセットを決定します。
        
        Args:
            drawing_xml: 描画XMLのルート要素
            anchors: フィルタリングされたアンカー要素のリスト
            keep_cnvpr_ids: 保持するcNvPr IDの初期セット
        
        Returns:
            保持するcNvPr IDの完全なセット（コネクタとエンドポイントを含む）
        """
        from collections import deque
        
        # 参照マッピングを構築
        refs = {}
        reverse_refs = {}
        
        for orig in list(drawing_xml):
            lname = orig.tag.split('}')[-1].lower()
            if lname not in ('twocellanchor', 'onecellanchor'):
                continue
            cid = extract_anchor_id(orig, allow_idx=False)
            if cid is None:
                continue
            
            # 参照されているIDを検索
            rset = set()
            for sub in orig.iter():
                st = sub.tag.split('}')[-1].lower()
                if st in ('stcxn', 'endcxn', 'stcxnpr', 'endcxnpr'):
                    vid = sub.attrib.get('id') or sub.attrib.get('idx')
                    if vid is not None:
                        rset.add(str(vid))
            if rset:
                refs[cid] = rset
                for rid in rset:
                    reverse_refs.setdefault(rid, set()).add(cid)
        
        # 行マッピングを構築
        id_to_row = {}
        all_id_to_row = {}
        ns_xdr = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
        
        for an in anchors:
            a_cid = extract_anchor_id(an, allow_idx=True)
            if a_cid is None:
                continue
            fr = an.find('{%s}from' % ns_xdr)
            if fr is not None:
                r = fr.find('{%s}row' % ns_xdr)
                if r is not None and r.text is not None:
                    try:
                        id_to_row[str(a_cid)] = int(r.text)
                    except (ValueError, TypeError) as e:
                        debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")
        
        # すべてのアンカーからのフォールバックマッピング
        for orig_an in list(drawing_xml):
            lname2 = orig_an.tag.split('}')[-1].lower()
            if lname2 not in ('twocellanchor', 'onecellanchor'):
                continue
            a_cid2 = extract_anchor_id(orig_an, allow_idx=True)
            if a_cid2 is None:
                continue
            fr2 = orig_an.find('{%s}from' % ns_xdr)
            if fr2 is not None:
                r2 = fr2.find('{%s}row' % ns_xdr)
                if r2 is not None and r2.text is not None:
                    try:
                        all_id_to_row[str(a_cid2)] = int(r2.text)
                    except (ValueError, TypeError) as e:
                        debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")
        
        # グループの行範囲を決定
        group_rows = set()
        for cid in keep_cnvpr_ids:
            rowval = id_to_row.get(str(cid))
            if rowval is not None:
                group_rows.add(int(rowval))
        
        # BFS展開
        preserve = set(keep_cnvpr_ids)
        q = deque(keep_cnvpr_ids)
        
        while q:
            current = q.popleft()
            for fwd in refs.get(current, set()):
                if fwd not in preserve:
                    preserve.add(fwd)
                    q.append(fwd)
            for rev in reverse_refs.get(current, set()):
                if rev not in preserve:
                    preserve.add(rev)
                    q.append(rev)
        
        # グループ行のコネクタのみのアンカーを含める（±1の許容範囲）
        for cid in list(all_id_to_row.keys()):
            scid = str(cid)
            if scid in preserve:
                continue
            rowc = id_to_row.get(scid) or all_id_to_row.get(scid)
            if rowc is not None and group_rows:
                if rowc in group_rows or any(abs(int(rowc) - int(gr)) <= 1 for gr in group_rows):
                    preserve.add(scid)
        
        debug_print(f"[DEBUG][_resolve_connector] keep={len(keep_cnvpr_ids)} → preserve={len(preserve)} (rows={sorted(list(group_rows))})")
        return preserve

    def _prune_drawing_anchors(
        self,
        drawing_relpath: str,
        keep_cnvpr_ids: Set[str],
        referenced_ids: Set[str],
        cell_range: Optional[Tuple[int, int, int, int]],
        group_rows: Set[int]
    ) -> None:
        """
        指定されたアンカーのみを保持するように描画XMLを刈り込みます。
        
        Args:
            drawing_relpath: 描画XMLファイルへのパス
            keep_cnvpr_ids: 保持するcNvPr IDのセット
            referenced_ids: コネクタによって参照されるIDのセット
            cell_range: オプションのセル範囲 (s_col, e_col, s_row, e_row)
            group_rows: グループの範囲内の行番号のセット
        """
        try:
            def node_contains_referenced_id(n):
                try:
                    vid = extract_anchor_id(n, allow_idx=True)
                    if vid is not None and str(vid) in referenced_ids:
                        return True
                    for sub in n.iter():
                        lname = sub.tag.split('}')[-1].lower()
                        if lname in ('stcxn', 'endcxn', 'stcxnpr', 'endcxnpr'):
                            vid = sub.attrib.get('id') or sub.attrib.get('idx')
                            if vid is not None and str(vid) in referenced_ids:
                                return True
                except (ValueError, TypeError) as e:
                    debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")
                return False
            
            tree = ET.parse(drawing_relpath)
            root = tree.getroot()
            
            removed_count = 0
            kept_count = 0
            
            for node in list(root):
                lname = node.tag.split('}')[-1].lower()
                if lname in ('twocellanchor', 'onecellanchor'):
                    this_cid = extract_anchor_id(node, allow_idx=True)
                    
                    if this_cid is not None and str(this_cid) in keep_cnvpr_ids:
                        kept_count += 1
                        debug_print(f"[DEBUG][_prune] KEEP anchor id={this_cid}")
                        continue
                    
                    try:
                        if node_contains_referenced_id(node):
                            continue
                    except (ValueError, TypeError) as e:
                        debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")
                    
                    try:
                        if (not keep_cnvpr_ids) and group_rows:
                            ns_xdr = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
                            fr = node.find('{%s}from' % ns_xdr)
                            if fr is not None:
                                r = fr.find('{%s}row' % ns_xdr)
                                if r is not None and r.text is not None:
                                    try:
                                        from_row = int(r.text)
                                        if from_row in group_rows or any(abs(from_row - gr) <= 1 for gr in group_rows):
                                            continue
                                    except (ValueError, TypeError) as e:
                                        debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")
                    except (ValueError, TypeError) as e:
                        debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")
                    
                    try:
                        root.remove(node)
                        removed_count += 1
                        debug_print(f"[DEBUG][_prune] REMOVE anchor id={this_cid}")
                    except (ValueError, TypeError):
                        try:
                            root.remove(node)
                            removed_count += 1
                            debug_print(f"[DEBUG][_prune] REMOVE anchor id={this_cid} (retry)")
                        except (ValueError, TypeError) as e:
                            debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")
            
            debug_print(f"[DEBUG][_prune] Summary: kept={kept_count}, removed={removed_count}, total={kept_count+removed_count}")
            tree.write(drawing_relpath, encoding='utf-8', xml_declaration=True)
        except Exception as e:
            debug_print(f"[DEBUG][_prune_drawing_anchors] Error: {e}")

    def _convert_excel_to_pdf(self, xlsx_path: str, tmpdir: str, apply_fit_to_page: bool = True) -> Optional[str]:
        """ExcelファイルをPDFに変換
        
        LibreOfficeを使用してExcelファイルをPDF形式に変換します。
        2つのレンダリングメソッドで重複していた処理を統合。
        
        Args:
            xlsx_path: 変換するExcelファイルのパス
            tmpdir: PDF出力先ディレクトリ
            apply_fit_to_page: 1ページに収める設定を適用するか
        
        Returns:
            生成されたPDFファイルのパス または None
        """
        try:
            # 元のファイルを上書きしないように一時コピーを作成
            tmp_xlsx = os.path.join(tmpdir, os.path.basename(xlsx_path))
            shutil.copyfile(xlsx_path, tmp_xlsx)
            
            # PDF変換前に縦横1ページ設定を適用
            if apply_fit_to_page:
                self._set_excel_fit_to_one_page(tmp_xlsx)
            
            # LibreOfficeでPDF変換
            cmd = [LIBREOFFICE_PATH, '--headless', '--convert-to', 'pdf', '--outdir', tmpdir, tmp_xlsx]
            debug_print(f"[DEBUG] LibreOffice export command: {' '.join(cmd)}")
            proc = subprocess.run(cmd, capture_output=True, text=True, timeout=90)
            
            if proc.returncode != 0:
                print(f"[WARNING] LibreOffice PDF 変換失敗: {proc.stderr}")
                return None
            
            # 生成されたPDFを探す
            pdf_name = f"{self.base_name}.pdf"
            pdf_path = os.path.join(tmpdir, pdf_name)
            
            if not os.path.exists(pdf_path):
                # LibreOfficeが異なる名前で出力した可能性
                candidates = [os.path.join(tmpdir, f) for f in os.listdir(tmpdir) if f.lower().endswith('.pdf')]
                if not candidates:
                    print("[WARNING] LibreOffice がPDFを出力しませんでした")
                    return None
                pdf_path = candidates[0]
            
            return pdf_path
        
        except Exception as e:
            print(f"[WARNING] Excel→PDF変換失敗: {e}")
            return None

    def _convert_pdf_page_to_png(self, pdf_path: str, page_index: int, dpi: int,
                                  output_dir: str, filename_prefix: str) -> Optional[str]:
        """PDFの指定ページをPNG画像に変換（PyMuPDF使用）
        
        PyMuPDFを使用してPDFの特定ページをPNG形式に変換します。
        複数のレンダリングメソッドで重複していた処理を統合。
        
        Args:
            pdf_path: 変換するPDFファイルのパス
            page_index: ページ番号(0始まり)
            dpi: 解像度
            output_dir: PNG出力先ディレクトリ
            filename_prefix: 出力ファイル名のプレフィックス
        
        Returns:
            生成されたPNGファイル名(相対パス) または None
        """
        try:
            png_filename = f"{filename_prefix}.png"
            png_path = os.path.join(output_dir, png_filename)
            
            debug_print(f"[DEBUG] PyMuPDFでPDF→PNG変換実行 (ページ {page_index}, DPI: {dpi})...")
            
            doc = fitz.open(pdf_path)
            if page_index >= len(doc):
                print(f"[WARNING] ページ{page_index}が存在しません（全{len(doc)}ページ）")
                doc.close()
                return None
            
            page = doc[page_index]
            
            mat = fitz.Matrix(dpi/72, dpi/72)
            pix = page.get_pixmap(matrix=mat, alpha=False)
            
            img_data = pix.tobytes("png")
            pix = None
            doc.close()
            
            img = Image.open(io.BytesIO(img_data))
            
            if img.mode == 'RGBA':
                background = Image.new('RGB', img.size, (255, 255, 255))
                background.paste(img, mask=img.split()[3] if len(img.split()) > 3 else None)
                img = background
            elif img.mode != 'RGB':
                img = img.convert('RGB')
            
            img.save(png_path, 'PNG', quality=100)
            
            print(f"[INFO] PNG変換完了: {png_path} (サイズ: {img.size[0]}x{img.size[1]})")
            
            return png_filename
        
        except Exception as e:
            print(f"[WARNING] PDF→PNG変換失敗: {e}")
            import traceback
            traceback.print_exc()
            return None

    def _convert_excel_to_svg(self, xlsx_path: str, tmpdir: str, apply_fit_to_page: bool = True) -> Optional[str]:
        """ExcelファイルをSVGに変換
        
        LibreOfficeを使用してExcelファイルをSVG形式に直接変換します。
        
        Args:
            xlsx_path: 変換するExcelファイルのパス
            tmpdir: SVG出力先ディレクトリ
            apply_fit_to_page: 1ページに収める設定を適用するか
        
        Returns:
            生成されたSVGファイルのパス または None
        """
        try:
            # 元のファイルを上書きしないように一時コピーを作成
            tmp_xlsx = os.path.join(tmpdir, os.path.basename(xlsx_path))
            shutil.copyfile(xlsx_path, tmp_xlsx)
            
            # SVG変換前に縦横1ページ設定を適用
            if apply_fit_to_page:
                self._set_excel_fit_to_one_page(tmp_xlsx)
            
            # LibreOfficeでSVG変換
            cmd = [LIBREOFFICE_PATH, '--headless', '--convert-to', 'svg', '--outdir', tmpdir, tmp_xlsx]
            debug_print(f"[DEBUG] LibreOffice SVG export command: {' '.join(cmd)}")
            proc = subprocess.run(cmd, capture_output=True, text=True, timeout=90)
            
            if proc.returncode != 0:
                print(f"[WARNING] LibreOffice SVG 変換失敗: {proc.stderr}")
                return None
            
            # 生成されたSVGを探す
            svg_name = f"{self.base_name}.svg"
            svg_path = os.path.join(tmpdir, svg_name)
            
            if not os.path.exists(svg_path):
                # LibreOfficeが異なる名前で出力した可能性
                candidates = [os.path.join(tmpdir, f) for f in os.listdir(tmpdir) if f.lower().endswith('.svg')]
                if not candidates:
                    print("[WARNING] LibreOffice がSVGを出力しませんでした")
                    return None
                svg_path = candidates[0]
            
            return svg_path
        
        except Exception as e:
            print(f"[WARNING] Excel→SVG変換失敗: {e}")
            return None

    def _convert_pdf_page_to_svg(self, pdf_path: str, page_index: int,
                                  output_dir: str, filename_prefix: str) -> Optional[str]:
        """PDFの指定ページをSVG画像に変換（PyMuPDF使用）
        
        PyMuPDFを使用してPDFの特定ページをSVG形式に変換します。
        
        Args:
            pdf_path: 変換するPDFファイルのパス
            page_index: ページ番号(0始まり)
            output_dir: SVG出力先ディレクトリ
            filename_prefix: 出力ファイル名のプレフィックス
        
        Returns:
            生成されたSVGファイル名(相対パス) または None
        """
        try:
            svg_filename = f"{filename_prefix}.svg"
            svg_path = os.path.join(output_dir, svg_filename)
            
            debug_print(f"[DEBUG] PyMuPDFでPDF→SVG変換実行 (ページ {page_index})...")
            
            doc = fitz.open(pdf_path)
            if page_index >= len(doc):
                print(f"[WARNING] ページ{page_index}が存在しません（全{len(doc)}ページ）")
                doc.close()
                return None
            
            page = doc[page_index]
            
            # SVGとして出力
            svg_content = page.get_svg_image()
            doc.close()
            
            # SVGファイルに書き込み
            with open(svg_path, 'w', encoding='utf-8') as f:
                f.write(svg_content)
            
            print(f"[INFO] SVG変換完了: {svg_path}")
            
            return svg_filename
        
        except Exception as e:
            print(f"[WARNING] PDF→SVG変換失敗: {e}")
            import traceback
            traceback.print_exc()
            return None

    def _convert_page_to_image(self, pdf_path: str, page_index: int, dpi: int,
                                output_dir: str, filename_prefix: str) -> Optional[str]:
        """PDFの指定ページを画像に変換（出力形式に応じてPNGまたはSVG）
        
        self.output_formatに基づいて適切な形式で出力します。
        
        Args:
            pdf_path: 変換するPDFファイルのパス
            page_index: ページ番号(0始まり)
            dpi: 解像度（PNG時のみ使用）
            output_dir: 出力先ディレクトリ
            filename_prefix: 出力ファイル名のプレフィックス
        
        Returns:
            生成された画像ファイル名(相対パス) または None
        """
        if self.output_format == 'svg':
            return self._convert_pdf_page_to_svg(pdf_path, page_index, output_dir, filename_prefix)
        else:
            return self._convert_pdf_page_to_png(pdf_path, page_index, dpi, output_dir, filename_prefix)

    # ========================================================================

    def _render_sheet_fallback(self, sheet, dpi: int = 600, insert_index: Optional[int] = None, insert_images: bool = True) -> bool:
        """シート全体を1枚の画像にレンダリング(真のフォールバック)
        
        isolated-groupレンダリングが行われない場合、または失敗した場合の最終手段として、
        シート全体を1枚の画像として出力します。出力形式はself.output_formatに従います。
        
        注意:
            isolated-groupレンダリングは_process_sheet_imagesで実行されるため、
            このメソッドでは単純にシート全体を画像化するのみです。
        
        Args:
            sheet: 対象シート
            dpi: DPI設定(デフォルト: 600、PNG時のみ使用)
            insert_index: Markdown挿入位置(None=末尾)
            insert_images: True=即座に挿入、False=登録のみ
        
        Returns:
            成功時True、失敗時False
        """
        tmpdir = None
        try:
            # 一時ディレクトリを作成
            tmpdir = tempfile.mkdtemp(prefix='xls2md_render_')
            
            # 1. Excel→PDF変換
            debug_print(f"[DEBUG] Fallback rendering for sheet: {sheet.title}")
            pdf_path = self._convert_excel_to_pdf(self.excel_file, tmpdir, apply_fit_to_page=True)
            if pdf_path is None:
                print("[WARNING] LibreOffice がPDFを出力しませんでした")
                return False
            
            # 2. シートのページインデックスを取得
            try:
                page_index = int(self.workbook.sheetnames.index(sheet.title))
            except (ValueError, TypeError):
                page_index = 0
            
            # 3. PDF→画像変換（出力形式に応じてPNGまたはSVG）
            safe_sheet = self._sanitize_filename(sheet.title)
            result_filename = self._convert_page_to_image(
                pdf_path,
                page_index,
                dpi,
                self.images_dir,
                f"{self.base_name}_{safe_sheet}_sheet"
            )
            
            if result_filename is None:
                fmt_name = self.output_format.upper()
                print(f"[WARNING] {fmt_name} 変換が失敗しました")
                return False
            
            # 4. 画像をMarkdownに登録または挿入
            if insert_images:
                # 即座にMarkdownに挿入
                md_line = f"![{sheet.title}](images/{result_filename})"
                try:
                    self._insert_markdown_image(insert_index, md_line, result_filename, sheet=sheet)
                    print(f"[SUCCESS] シート全体の画像を挿入: {result_filename}")
                except Exception as e:
                    print(f"[WARNING] 画像挿入失敗: {e}")
                    # フォールバック: markdown_linesに直接追加
                    self.markdown_lines.append(md_line)
                    self.markdown_lines.append("")
                    self._mark_image_emitted(result_filename)
            else:
                # 後で挿入するために登録のみ
                self._sheet_shape_images.setdefault(sheet.title, [])
                self._sheet_shape_images[sheet.title].append((1, result_filename))
            
            return True
            
        except Exception as e:
            print(f"[WARNING] フォールバックレンダリングエラー: {e}")
            import traceback
            traceback.print_exc()
            return False
            
        finally:
            # 一時ディレクトリをクリーンアップ
            if tmpdir and os.path.isdir(tmpdir):
                try:
                    shutil.rmtree(tmpdir, ignore_errors=True)
                except (ValueError, TypeError):
                    pass  # 一時ディレクトリ削除失敗は無視

    def _detect_image_format(self, image_data: bytes) -> str:
        """先頭バイトから一般的な画像形式を検出し、拡張子を返します。

        不明な場合は.pngにフォールバックします。
        """
        try:
            if not image_data or len(image_data) < 4:
                return '.png'
            # JPEG形式
            if image_data.startswith(b'\xff\xd8'):
                return '.jpg'
            # PNG形式
            if image_data.startswith(b'\x89PNG'):
                return '.png'
            # GIF形式
            if image_data.startswith(b'GIF87a') or image_data.startswith(b'GIF89a'):
                return '.gif'
            # BMP形式
            if image_data.startswith(b'BM'):
                return '.bmp'
            return '.png'
        except Exception:
            return '.png'

    def _sanitize_filename(self, s: str) -> str:
        """ファイルシステム上で安全なファイル名に正規化する。

        - Unicode 正規化 (NFKC)
        - 連続空白はアンダースコアに置換
        - ファイル名に使えない記号 (/\\:*?"<>|) を除去
        - 複数アンダースコアを単一に、先頭/末尾のアンダースコアを除去
        """
        import unicodedata, re
        if s is None:
            return 'image'
        txt = unicodedata.normalize('NFKC', str(s))
        # 空白をアンダースコアに置換
        txt = re.sub(r"\s+", '_', txt)
        # ファイル名で問題となる文字を削除
        txt = re.sub(r'[/\\:*?"<>|]', '', txt)
        # 複数のアンダースコアを1つに統合
        txt = re.sub(r'_+', '_', txt)
        # 先頭/末尾のアンダースコアを削除
        txt = txt.strip('_')
        if not txt:
            return 'image'
        return txt

    def _get_drawing_max_col_row(self, sheet):
        """図形が参照する最大の列・行番号を取得する。
        
        Returns:
            (max_col, max_row): 図形が参照する最大の列・行番号のタプル。
                               図形が存在しない場合は (None, None) を返す。
        """
        try:
            metadata = self._get_drawing_xml_and_metadata(sheet)
            if metadata is None:
                return None, None
            
            drawing_xml = metadata['drawing_xml']
            ns = {'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'}
            
            max_col = None
            max_row = None
            
            for node in drawing_xml:
                lname = node.tag.split('}')[-1].lower()
                if lname not in ('twocellanchor', 'onecellanchor'):
                    continue
                
                if lname == 'twocellanchor':
                    fr = node.find('xdr:from', ns)
                    to = node.find('xdr:to', ns)
                    if fr is not None:
                        try:
                            col = int(fr.find('xdr:col', ns).text)
                            row = int(fr.find('xdr:row', ns).text)
                            if max_col is None or col > max_col:
                                max_col = col
                            if max_row is None or row > max_row:
                                max_row = row
                        except Exception:
                            pass
                    if to is not None:
                        try:
                            col = int(to.find('xdr:col', ns).text)
                            row = int(to.find('xdr:row', ns).text)
                            if max_col is None or col > max_col:
                                max_col = col
                            if max_row is None or row > max_row:
                                max_row = row
                        except Exception:
                            pass
                elif lname == 'onecellanchor':
                    fr = node.find('xdr:from', ns)
                    if fr is not None:
                        try:
                            col = int(fr.find('xdr:col', ns).text)
                            row = int(fr.find('xdr:row', ns).text)
                            if max_col is None or col > max_col:
                                max_col = col
                            if max_row is None or row > max_row:
                                max_row = row
                        except Exception:
                            pass
            
            return max_col, max_row
        except Exception:
            return None, None

    def _compute_sheet_cell_pixel_map(self, sheet, DPI=300, min_cols=None, min_rows=None):
        """列の右端と行の下端のおおよそのピクセル位置を計算します。

        col_x[0] == 0 で col_x[i] は列i（1始まり）の右端を返します。
        row_y も同様に行に対応します。
        
        図形が参照する列・行がシートの max_column/max_row より大きい場合は、
        図形の範囲まで計算を拡張します。
        """
        try:
            max_col = sheet.max_column
            max_row = sheet.max_row
            
            drawing_max_col, drawing_max_row = self._get_drawing_max_col_row(sheet)
            if drawing_max_col is not None:
                drawing_max_col_1based = drawing_max_col + 1
                if drawing_max_col_1based > max_col:
                    max_col = drawing_max_col_1based
            if drawing_max_row is not None:
                drawing_max_row_1based = drawing_max_row + 1
                if drawing_max_row_1based > max_row:
                    max_row = drawing_max_row_1based
            
            if min_cols is not None:
                max_col = max(max_col, min_cols)
            if min_rows is not None:
                max_row = max(max_row, min_rows)
            
            col_pixels = []
            from openpyxl.utils import get_column_letter
            for c in range(1, max_col+1):
                cd = sheet.column_dimensions.get(get_column_letter(c))
                # Excelの列幅は文字単位です。Microsoftのドキュメントに基づく
                # より正確な変換を使用します:
                # pixels = floor(((256*W + floor(128/MAX_DIGIT_WIDTH))/256) * MAX_DIGIT_WIDTH)
                # MAX_DIGIT_WIDTHはワークブックのデフォルトフォントでの最大数字幅
                # （ピクセル単位）を近似します。Calibri/Arialのデフォルトサイズで
                # 一般的な7を保守的なデフォルト値として使用します。
                width = getattr(cd, 'width', None) if cd is not None else None
                if width is None:
                    try:
                        from openpyxl.utils import units as _units
                        width = getattr(sheet.sheet_format, 'defaultColWidth', None) or _units.DEFAULT_COLUMN_WIDTH
                    except Exception:
                        width = 8.43
                try:
                    import math
                    # 標準画面DPI（96）での基本ピクセル幅を計算します。その後、
                    # 要求されたDPIにスケーリングして、EMUオフセット（後で
                    # ターゲットラスタライズDPIを使用して変換される）が
                    # 生成されたPDF/PNGピクセルと一致するようにします。
                    # これにより、描画EMU変換と他で使用される列ピクセルマップ間の
                    # 不一致を軽減します。
                    MAX_DIGIT_WIDTH = 7
                    base_px = int(math.floor(((256.0 * float(width) + math.floor(128.0 / MAX_DIGIT_WIDTH)) / 256.0) * MAX_DIGIT_WIDTH))
                    if base_px < 1:
                        base_px = 1
                    # 96 DPI（一般的な画面）からターゲットDPIにスケーリング
                    scale = float(DPI) / 96.0 if DPI and DPI > 0 else 1.0
                    px = max(1, int(round(base_px * scale)))
                except (ValueError, TypeError):
                    # フォールバックヒューリスティック、DPIでもスケーリング
                    try:
                        base = max(1, int(float(width) * 7 + 5))
                        px = max(1, int(round(base * (float(DPI) / 96.0))))
                    except (ValueError, TypeError):
                        px = max(1, int(float(width) * 7 + 5))
                col_pixels.append(px)

            row_pixels = []
            for r in range(1, max_row+1):
                rd = sheet.row_dimensions.get(r)
                hpts = getattr(rd, 'height', None) if rd is not None else None
                if hpts is None:
                    try:
                        from openpyxl.utils import units as _units
                        hpts = _units.DEFAULT_ROW_HEIGHT
                    except Exception:
                        hpts = 15
                # 行の高さはポイント単位。ターゲットDPIでピクセルに変換
                try:
                    px = max(1, int(float(hpts) * DPI / 72.0))
                except (ValueError, TypeError):
                    px = max(1, int(hpts * DPI / 72))
                row_pixels.append(px)

            col_x = [0]
            for v in col_pixels:
                col_x.append(col_x[-1] + v)
            row_y = [0]
            for v in row_pixels:
                row_y.append(row_y[-1] + v)

            return col_x, row_y
        except Exception:
            return [0], [0]

    def _to_positive(self, value, orig_ext, orig_ch_ext, target_px):
        """EMU範囲が正であることを確認します。

        正の範囲を選択する優先順位:
        1. 'value'が既に正の場合はそれを保持
        2. orig_ext（提供され、>0の場合）にフォールバック
        3. orig_ch_ext（提供され、>0の場合）にフォールバック
        4. target_px -> EMU変換にフォールバック（最低1ピクセル）
        5. 最終的に絶対安全最小値として1 EMUを返す

        このヘルパーは防御的で、エラーを発生させません。
        常に > 0 の int を返します。
        """
        try:
            v = int(value) if value is not None else 0
        except (ValueError, TypeError):
            v = 0
        if v and v > 0:
            return v
        try:
            if orig_ext is not None:
                oe = int(orig_ext)
                if oe > 0:
                    return oe
        except (ValueError, TypeError):
            pass  # 型変換失敗は無視
        try:
            if orig_ch_ext is not None:
                oc = int(orig_ch_ext)
                if oc > 0:
                    return oc
        except (ValueError, TypeError):
            pass  # 型変換失敗は無視
        try:
            # target_pxはピクセル単位。利用可能な場合はオブジェクトのdpiを使用してEMUに変換
            DPI = int(getattr(self, 'dpi', 300) or 300)
            EMU_PER_INCH = 914400
            emu_per_pixel = EMU_PER_INCH / float(DPI) if DPI and DPI > 0 else EMU_PER_INCH / 300.0
            px = float(target_px) if target_px is not None else 1.0
            emu = int(round(max(1.0, px) * emu_per_pixel))
            if emu and emu > 0:
                return emu
        except (ValueError, TypeError):
            pass  # 型変換失敗は無視
        # 絶対フォールバック
        return 1

    def _snap_box_to_cell_bounds(self, box, col_x, row_y, DPI=300):
        """ピクセルボックス(l,t,r,b)を、提供されたcol_xとrow_y配列を使用して
        最も近い囲むセル境界にスナップします。整数ピクセルボックスを返します。
        """
        try:
            l, t, r, btm = box
            # 開始列を検索: col_x[c] >= l となる最小のc（小さな許容範囲を許可）
            # tolはDPIに応じてスケーリングし、DPIが異なる場合の以前の動作を保持
            try:
                tol = max(1, int(DPI / 300.0 * 3))  # DPIに依存する数ピクセルの許容範囲
            except (ValueError, TypeError):
                tol = 3
            start_col = None
            for c in range(1, len(col_x)):
                if col_x[c] >= l - tol:
                    start_col = c
                    break
            if start_col is None:
                start_col = max(1, len(col_x)-1)

            # 終了列を検索: col_x[c] >= r となる最小のc（小さな許容範囲を許可）
            end_col = None
            for c in range(1, len(col_x)):
                if col_x[c] >= r + tol:
                    end_col = c
                    break
            if end_col is None:
                end_col = max(1, len(col_x)-1)

            # 行
            start_row = None
            for rr in range(1, len(row_y)):
                if row_y[rr] >= t - tol:
                    start_row = rr
                    break
            if start_row is None:
                start_row = max(1, len(row_y)-1)

            end_row = None
            for rr in range(1, len(row_y)):
                if row_y[rr] >= btm + tol:
                    end_row = rr
                    break
            if end_row is None:
                end_row = max(1, len(row_y)-1)

            left_px = max(0, int(col_x[start_col-1]))
            top_px = max(0, int(row_y[start_row-1]))
            right_px = int(col_x[end_col]) if end_col < len(col_x) else int(col_x[-1])
            bottom_px = int(row_y[end_row]) if end_row < len(row_y) else int(row_y[-1])

            return left_px, top_px, right_px, bottom_px
        except (ValueError, TypeError):
            return int(box[0]), int(box[1]), int(box[2]), int(box[3])

    def _find_content_bbox(self, pil_image, white_thresh: int = 245):
        """PIL画像内の非白色コンテンツのバウンディングボックスを検索します。

        white_thresh: ピクセル輝度閾値（0-255）。すべてのチャンネルが
        >= white_threshのピクセルは背景/白と見なされます。
        コンテンツが検出されない場合は(l,t,r,b)またはNoneを返します。
        """
        try:
            if pil_image.mode not in ('RGB', 'RGBA'):
                img = pil_image.convert('RGB')
            else:
                img = pil_image
            pixels = img.load()
            w, h = img.size
            left = w; top = h; right = 0; bottom = 0
            found = False
            for y in range(h):
                for x in range(w):
                    r, g, b = pixels[x, y][:3]
                    if r < white_thresh or g < white_thresh or b < white_thresh:
                        found = True
                        if x < left: left = x
                        if x > right: right = x
                        if y < top: top = y
                        if y > bottom: bottom = y
            if not found:
                return None
            # right/bottomを包含的に -> 一般的なクロップ座標(r+1,b+1)に変換
            return (left, top, right + 1, bottom + 1)
        except Exception:
            return None

    def _crop_image_preserving_connectors(self, image_path: str, dpi: int = 300, white_thresh: int = 245):
        """image_pathの画像を開き、非白色のbboxを検索してパディング付きでクロップします。

        コネクタ/矢印の先端を切り取らないように、小さなパディング（DPIに依存）を追加します。
        クロップ結果で元のファイルを上書きします。
        """
        try:
            from PIL import Image
            if not os.path.exists(image_path):
                return False
            im = Image.open(image_path)
            bbox = self._find_content_bbox(im, white_thresh=white_thresh)
            if not bbox:
                # クロップするものがない
                im.close()
                return True
            l, t, r, b = bbox
            # 細い矢印の先端を切り取らないためのパディング。DPIでスケーリング
            base_pad = max(6, int(dpi / 300.0 * 12))
            # 尾部/矢じりがクリップされないように、下部パディングをやや大きくバイアス
            pad_top = base_pad
            pad_left = base_pad
            pad_right = base_pad
            pad_bottom = max(base_pad, int(base_pad * 1.25))
            l = max(0, l - pad_left)
            t = max(0, t - pad_top)
            r = min(im.width, r + pad_right)
            b = min(im.height, b + pad_bottom)
            # クロップを実行して保存（モードを保持）
            try:
                cropped = im.crop((l, t, r, b))
                cropped.save(image_path)
                cropped.close()
            except (ValueError, TypeError):
                # フォールバック: クロップが失敗した場合は上書きしない
                pass
            im.close()
            return True
        except Exception:
            return False

    def _get_pdf_page_box_points(self, pdf_path: str):
        """PDFのCropBoxまたはMediaBoxを使用して、最初のページの(width_points, height_points)を返します。

        これはPDFバイト内の'/CropBox'または'/MediaBox'配列を検索する軽量パーサーです。
        失敗時はNoneを返します。
        """
        try:
            with open(pdf_path, 'rb') as f:
                data = f.read()
            # まずCropBoxを検索し、次にMediaBoxを検索
            import re
            pat = re.compile(rb"/CropBox\s*\[\s*([0-9.+\-eE]+)\s+([0-9.+\-eE]+)\s+([0-9.+\-eE]+)\s+([0-9.+\-eE]+)\s*\]")
            m = pat.search(data)
            if not m:
                pat2 = re.compile(rb"/MediaBox\s*\[\s*([0-9.+\-eE]+)\s+([0-9.+\-eE]+)\s+([0-9.+\-eE]+)\s+([0-9.+\-eE]+)\s*\]")
                m = pat2.search(data)
            if not m:
                return None
            a = float(m.group(1))
            b = float(m.group(2))
            c = float(m.group(3))
            d = float(m.group(4))
            width_pts = abs(c - a)
            height_pts = abs(d - b)
            return (width_pts, height_pts)
        except (ValueError, TypeError):
            return None

    def _extract_drawing_cell_ranges(self, sheet) -> List[Tuple[int,int,int,int]]:
        """各描画可能アンカーの描画セル範囲(start_col, end_col, start_row, end_row)を抽出します。

        利用可能な場合は描画XMLを使用します。他の抽出器で使用されるアンカー順序に揃えたリストを返します。
        """
        print(f"[INFO] シート図形セル範囲抽出: {sheet.title}")
        ranges = []
        try:
            # Phase 1基盤メソッドを使用して描画XMLを取得
            metadata = self._get_drawing_xml_and_metadata(sheet)
            if metadata is None:
                return ranges
            
            z = metadata['zip']
            drawing_xml = metadata['drawing_xml']

            # oneCellのext変換用にピクセルマップを準備
            # EMUオフセットをピクセルに変換する際に一貫したDPIを使用
            DPI = 300
            try:
                DPI = int(getattr(self, 'dpi', DPI) or DPI)
            except (ValueError, TypeError):
                pass  # データ構造操作失敗は無視
            try:
                DPI = int(getattr(self, 'dpi', DPI) or DPI)
            except (ValueError, TypeError):
                DPI = DPI
            col_x, row_y = self._compute_sheet_cell_pixel_map(sheet, DPI=DPI)
            EMU_PER_INCH = 914400
            try:
                EMU_PER_PIXEL = EMU_PER_INCH / float(DPI)
            except (ValueError, TypeError):
                EMU_PER_PIXEL = EMU_PER_INCH / float(DPI)

            ns = {'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'}
            for node in drawing_xml:
                lname = node.tag.split('}')[-1].lower()
                if lname not in ('twocellanchor', 'onecellanchor'):
                    continue
                # セルインデックスを決定
                if lname == 'twocellanchor':
                    fr = node.find('xdr:from', ns)
                    to = node.find('xdr:to', ns)
                    if fr is None or to is None:
                        continue
                    try:
                        col = int(fr.find('xdr:col', ns).text)
                        row = int(fr.find('xdr:row', ns).text)
                        to_col = int(to.find('xdr:col', ns).text)
                        to_row = int(to.find('xdr:row', ns).text)
                    except (ValueError, TypeError):
                        continue
                    start_col = col + 1
                    end_col = to_col + 1
                    start_row = row + 1
                    end_row = to_row + 1

                    try:
                        ns_a = {'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'}
                        sp = node.find('.//xdr:sp', ns)
                        if sp is not None:
                            prst_geom = sp.find('.//a:prstGeom', ns_a)
                            if prst_geom is not None:
                                prst = prst_geom.get('prst', '')
                                if 'callout' in prst.lower():
                                    # これはコールアウト図形。調整値を確認
                                    av_lst = prst_geom.find('a:avLst', ns_a)
                                    if av_lst is not None:
                                        adj1_elem = av_lst.find('a:gd[@name="adj1"]', ns_a)
                                        adj2_elem = av_lst.find('a:gd[@name="adj2"]', ns_a)
                                        
                                        adj1 = 0
                                        adj2 = 0
                                        if adj1_elem is not None:
                                            fmla = adj1_elem.get('fmla', '')
                                            if fmla.startswith('val '):
                                                try:
                                                    adj1 = int(fmla.split()[1])
                                                except (ValueError, IndexError):
                                                    pass
                                        if adj2_elem is not None:
                                            fmla = adj2_elem.get('fmla', '')
                                            if fmla.startswith('val '):
                                                try:
                                                    adj2 = int(fmla.split()[1])
                                                except (ValueError, IndexError):
                                                    pass
                                        
                                        if adj1 < 0 or adj2 < 0:
                                            if start_row > 1:
                                                start_row -= 1
                    except Exception:
                        pass

                    ranges.append((start_col, end_col, start_row, end_row))
                else:
                    # oneCellAnchor: from.col/from.rowとext cx/cyを使用して終了セルを導出
                    fr = node.find('xdr:from', ns)
                    ext = node.find('xdr:ext', ns)
                    if fr is None or ext is None:
                        continue
                    try:
                        col = int(fr.find('xdr:col', ns).text)
                        row = int(fr.find('xdr:row', ns).text)
                        colOff = int(fr.find('xdr:colOff', ns).text)
                    except (ValueError, TypeError):
                        continue
                    cx = int(ext.attrib.get('cx', '0'))
                    cy = int(ext.attrib.get('cy', '0'))
                    left_px = col_x[col] + (colOff / EMU_PER_PIXEL) if col < len(col_x) else col_x[-1]
                    right_px = left_px + (cx / EMU_PER_PIXEL)
                    top_px = row_y[row] if row < len(row_y) else row_y[-1]
                    bottom_px = top_px + (cy / EMU_PER_PIXEL)
                    # ピクセルをセルインデックスにマップ
                    # start_colインデックスを検索
                    start_col = 1
                    for ci in range(1, len(col_x)):
                        if col_x[ci] >= left_px:
                            start_col = ci
                            break
                    end_col = len(col_x)-1
                    for ci in range(1, len(col_x)):
                        if col_x[ci] >= right_px:
                            end_col = ci
                            break
                    start_row = 1
                    for ri in range(1, len(row_y)):
                        if row_y[ri] >= top_px:
                            start_row = ri
                            break
                    end_row = len(row_y)-1
                    for ri in range(1, len(row_y)):
                        if row_y[ri] >= bottom_px:
                            end_row = ri
                            break
                    ranges.append((start_col, end_col, start_row, end_row))
        except Exception:
            pass  # データ構造操作失敗は無視
        print(f"[INFO] 抽出されたセル範囲: {ranges}")
        return ranges

    def _anchor_is_connector_only(self, sheet, anchor_idx) -> bool:
        """シートの描画内のanchor_idxにあるアンカーがコネクタのみのアンカー
        （つまり、コネクタエンドポイント参照を含むが、描画可能な画像/図形要素を
        含まない）であるかどうかを返します。保守的: 情報が判定できない場合はFalseを返します。
        """
        try:
            z = zipfile.ZipFile(self.excel_file)
            sheet_index = self.workbook.sheetnames.index(sheet.title)
            rels_path = f"xl/worksheets/_rels/sheet{sheet_index+1}.xml.rels"
            rels_xml = get_xml_from_zip(z, rels_path)
            if rels_xml is None:
                return False
            drawing_target = None
            for rel in rels_xml.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                t = rel.attrib.get('Type','')
                if t.endswith('/drawing'):
                    drawing_target = rel.attrib.get('Target')
                    break
            if not drawing_target:
                return False
            drawing_path = normalize_excel_path(drawing_target)
            if drawing_path not in z.namelist():
                drawing_path = drawing_path.replace('worksheets', 'drawings')
                if drawing_path not in z.namelist():
                    return False
            drawing_xml = get_xml_from_zip(z, drawing_path)
            if drawing_xml is None:
                return False
            # locate the requested anchor node
            idx = 0
            for node in drawing_xml:
                lname = node.tag.split('}')[-1].lower()
                if lname not in ('twocellanchor', 'onecellanchor'):
                    continue
                if idx == anchor_idx:
                    # ノードの子を検査して描画可能タイプとコネクタ参照を比較
                    has_drawable = False
                    has_connector_ref = False
                    for desc in node.iter():
                        t = desc.tag.split('}')[-1].lower()
                        if t in ('pic', 'sp', 'graphicframe', 'grpsp'):
                            has_drawable = True
                        if t in ('stcxn', 'endcxn', 'stcxnpr', 'endcxnpr'):
                            has_connector_ref = True
                        for k in desc.attrib.keys():
                            if k.lower() == 'id' and desc.tag.split('}')[-1].lower() != 'cnvpr':
                                has_connector_ref = True
                    return (has_connector_ref and not has_drawable)
                idx += 1
        except Exception:
            return False
        return False

    def _sheet_has_drawings(self, sheet) -> bool:
        """シートが描画可能要素（pic/sp/graphicFrame）を含む描画XMLを指す
        描画リレーションシップを持つ場合にTrueを返します。"""
        try:
            z = zipfile.ZipFile(self.excel_file)
            sheet_index = self.workbook.sheetnames.index(sheet.title)
            rels_path = f"xl/worksheets/_rels/sheet{sheet_index+1}.xml.rels"
            rels_xml = get_xml_from_zip(z, rels_path)
            if rels_xml is None:
                return False
            drawing_target = None
            for rel in rels_xml.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                t = rel.attrib.get('Type','')
                if t.endswith('/drawing'):
                    drawing_target = rel.attrib.get('Target')
                    break
            if not drawing_target:
                return False
            drawing_path = normalize_excel_path(drawing_target)
            if drawing_path not in z.namelist():
                drawing_path = drawing_path.replace('worksheets', 'drawings')
                if drawing_path not in z.namelist():
                    return False
            drawing_xml = get_xml_from_zip(z, drawing_path)
            if drawing_xml is None:
                return False
            # 描画可能な子孫要素を検索
            for node in drawing_xml.iter():
                lname = node.tag.split('}')[-1].lower()
                if lname in ('pic', 'sp', 'graphicframe', 'graphic', 'grpsp'):
                    return True
            return False
        except (ET.ParseError, KeyError, AttributeError):
            return False

    def _extract_shape_metadata_from_anchor(self, anchor, sheet) -> Dict[str, Any]:
        """drawing anchorから図形のメタデータを抽出
        
        Args:
            anchor: XML anchor element
            sheet: 対象シート
            
        Returns:
            図形のメタデータを含む辞書
        """
        metadata = {
            'type': 'unknown',
            'name': '',
            'position': {},
            'size': {},
            'text_content': [],
            'connector_type': None,
            'shape_properties': {}
        }
        
        try:
            ns_xdr = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
            ns_a = 'http://schemas.openxmlformats.org/drawingml/2006/main'
            
            anchor_type = anchor.tag.split('}')[-1].lower()
            metadata['anchor_type'] = anchor_type
            
            metadata['id'] = extract_anchor_id(anchor, allow_idx=False) or ''
            for sub in anchor.iter():
                if sub.tag.split('}')[-1].lower() == 'cnvpr':
                    metadata['name'] = sub.attrib.get('name', '')
                    metadata['description'] = sub.attrib.get('descr', '')
                    break
            
            if anchor_type == 'twocellanchor':
                fr = anchor.find('{%s}from' % ns_xdr)
                to = anchor.find('{%s}to' % ns_xdr)
                if fr is not None:
                    col_elem = fr.find('{%s}col' % ns_xdr)
                    row_elem = fr.find('{%s}row' % ns_xdr)
                    if col_elem is not None and row_elem is not None:
                        metadata['position']['from_col'] = int(col_elem.text) + 1
                        metadata['position']['from_row'] = int(row_elem.text) + 1
                if to is not None:
                    col_elem = to.find('{%s}col' % ns_xdr)
                    row_elem = to.find('{%s}row' % ns_xdr)
                    if col_elem is not None and row_elem is not None:
                        metadata['position']['to_col'] = int(col_elem.text) + 1
                        metadata['position']['to_row'] = int(row_elem.text) + 1
            
            elif anchor_type == 'onecellanchor':
                fr = anchor.find('{%s}from' % ns_xdr)
                ext = anchor.find('{%s}ext' % ns_xdr)
                if fr is not None:
                    col_elem = fr.find('{%s}col' % ns_xdr)
                    row_elem = fr.find('{%s}row' % ns_xdr)
                    if col_elem is not None and row_elem is not None:
                        metadata['position']['from_col'] = int(col_elem.text) + 1
                        metadata['position']['from_row'] = int(row_elem.text) + 1
                if ext is not None:
                    cx = ext.attrib.get('cx', '0')
                    cy = ext.attrib.get('cy', '0')
                    metadata['size']['width_emu'] = int(cx)
                    metadata['size']['height_emu'] = int(cy)
            
            for desc in anchor.iter():
                tag_name = desc.tag.split('}')[-1].lower()
                
                if tag_name == 'pic':
                    metadata['type'] = 'picture'
                elif tag_name == 'sp':
                    metadata['type'] = 'shape'
                    for prstgeom in desc.iter():
                        if prstgeom.tag.split('}')[-1].lower() == 'prstgeom':
                            prst = prstgeom.attrib.get('prst', '')
                            metadata['shape_properties']['preset'] = prst
                            break
                elif tag_name == 'cxnsp':
                    metadata['type'] = 'connector'
                    for prstgeom in desc.iter():
                        if prstgeom.tag.split('}')[-1].lower() == 'prstgeom':
                            prst = prstgeom.attrib.get('prst', '')
                            metadata['connector_type'] = prst
                            break
                elif tag_name == 'grpsp':
                    metadata['type'] = 'group'
                elif tag_name == 'graphicframe':
                    metadata['type'] = 'graphic_frame'
                
                if tag_name == 't' and desc.text and desc.text.strip():
                    metadata['text_content'].append(desc.text.strip())
            
            return metadata
            
        except Exception as e:
            print(f"[WARNING] Shape metadata extraction failed: {e}")
            return metadata

    def _extract_all_shapes_metadata(self, sheet, filter_ids: Optional[Set[str]] = None) -> List[Dict[str, Any]]:
        """シート内の全図形のメタデータを抽出
        
        Args:
            sheet: 対象シート
            filter_ids: 抽出対象の図形IDセット（Noneの場合は全て抽出）
            
        Returns:
            図形メタデータのリスト
        """
        shapes_metadata = []
        
        try:
            metadata = self._get_drawing_xml_and_metadata(sheet)
            if metadata is None:
                return shapes_metadata
            
            drawing_xml = metadata['drawing_xml']
            
            for anchor in drawing_xml:
                anchor_type = anchor.tag.split('}')[-1].lower()
                if anchor_type in ('twocellanchor', 'onecellanchor'):
                    if self._anchor_has_drawable(anchor):
                        if filter_ids is not None:
                            shape_id = extract_anchor_id(anchor, allow_idx=False) or ''
                            if shape_id and shape_id not in filter_ids:
                                continue
                        
                        shape_meta = self._extract_shape_metadata_from_anchor(anchor, sheet)
                        shapes_metadata.append(shape_meta)
            
            try:
                metadata['zip'].close()
            except:
                pass
                
        except Exception as e:
            print(f"[WARNING] Failed to extract shapes metadata: {e}")
        
        return shapes_metadata

    def _format_shape_metadata_as_text(self, shapes_metadata: List[Dict[str, Any]]) -> str:
        """図形メタデータを人間が読みやすいテキスト形式に整形
        
        Args:
            shapes_metadata: 図形メタデータのリスト
            
        Returns:
            整形されたテキスト
        """
        if not shapes_metadata:
            return ""
        
        lines = []
        lines.append("### 図形情報")
        lines.append("")
        
        for idx, meta in enumerate(shapes_metadata, 1):
            shape_type = meta.get('type', 'unknown')
            shape_name = meta.get('name', f'図形{idx}')
            
            type_map = {
                'picture': '画像',
                'shape': '図形',
                'connector': 'コネクタ',
                'group': 'グループ',
                'graphic_frame': 'グラフィックフレーム',
                'unknown': '不明'
            }
            type_ja = type_map.get(shape_type, shape_type)
            
            lines.append(f"**{shape_name}** ({type_ja})")
            
            pos = meta.get('position', {})
            if 'from_col' in pos and 'from_row' in pos:
                from_cell = f"{col_letter(pos['from_col'])}{pos['from_row']}"
                if 'to_col' in pos and 'to_row' in pos:
                    to_cell = f"{col_letter(pos['to_col'])}{pos['to_row']}"
                    lines.append(f"- 位置: {from_cell} ～ {to_cell}")
                else:
                    lines.append(f"- 位置: {from_cell} から")
            
            if shape_type == 'shape':
                preset = meta.get('shape_properties', {}).get('preset', '')
                if preset:
                    lines.append(f"- 図形タイプ: {preset}")
            
            if shape_type == 'connector':
                conn_type = meta.get('connector_type', '')
                if conn_type:
                    lines.append(f"- コネクタタイプ: {conn_type}")
            
            text_content = meta.get('text_content', [])
            if text_content:
                lines.append(f"- テキスト: {' / '.join(text_content)}")
            
            description = meta.get('description', '')
            if description:
                lines.append(f"- 説明: {description}")
            
            lines.append("")
        
        return '\n'.join(lines)

    def _format_shape_metadata_as_json(self, shapes_metadata: List[Dict[str, Any]]) -> str:
        """図形メタデータをJSON形式に整形
        
        Args:
            shapes_metadata: 図形メタデータのリスト
            
        Returns:
            JSON文字列
        """
        if not shapes_metadata:
            return "{}"
        
        import json
        
        output_data = {
            'shapes': [],
            'total_count': len(shapes_metadata)
        }
        
        for meta in shapes_metadata:
            shape_data = {
                'name': meta.get('name', ''),
                'type': meta.get('type', 'unknown'),
                'anchor_type': meta.get('anchor_type', ''),
                'position': meta.get('position', {}),
                'size': meta.get('size', {}),
                'text_content': meta.get('text_content', []),
                'properties': {}
            }
            
            if meta.get('type') == 'shape':
                shape_data['properties'] = meta.get('shape_properties', {})
            elif meta.get('type') == 'connector':
                shape_data['properties']['connector_type'] = meta.get('connector_type', '')
            
            if meta.get('description'):
                shape_data['description'] = meta.get('description')
            
            if meta.get('id'):
                shape_data['id'] = meta.get('id')
            
            output_data['shapes'].append(shape_data)
        
        return json.dumps(output_data, ensure_ascii=False, indent=2)

    def _anchor_has_drawable(self, a) -> bool:
        """共有ヘルパー: 描画アンカーが描画可能なコンテンツ（画像、図形、
        graphicFrame、またはコネクタ参照）を含むかどうかを判定します。
        この中央実装により、抽出とトリミングのロジックが一貫し、
        クラスタリングインデックスがアンカーと整合します。
        """
        try:
            # 単一の変換実行中に同じアンカーを複数回再評価することを避けるため、
            # インスタンスにキャッシュ辞書を作成します。利用可能な場合は最も近い
            # cNvPr/@id属性を安定したキーとして使用し、IDがない場合は
            # アンカーXMLの短いハッシュにフォールバックします。
            try:
                cache = getattr(self, '_anchor_drawable_cache')
            except Exception:
                cache = {}
                try:
                    setattr(self, '_anchor_drawable_cache', cache)
                except Exception as e:
                    pass  # XML解析エラーは無視

            key = None
            try:
                cid = extract_anchor_id(a, allow_idx=True)
                if cid is not None:
                    key = f"cnvpr:{cid}"
            except Exception:
                key = None

            if key is None:
                try:
                    # フォールバック: アンカーXMLの小さな安定したフィンガープリント
                    import hashlib
                    raw = ET.tostring(a) if hasattr(ET, 'tostring') else None
                    if raw:
                        key = 'hash:' + hashlib.sha1(raw).hexdigest()[:8]
                    else:
                        key = 'anon'
                except Exception:
                    key = 'anon'

            # キャッシュされた結果があれば返す
            try:
                if key in cache:
                    return cache[key]
            except Exception as e:
                pass  # XML解析エラーは無視

            drawable_types = []
            has_text = False
            has_connector_ref = False
            for desc in a.iter():
                lname = desc.tag.split('}')[-1].lower()
                # テキストコンテンツを検出
                if lname == 't' and (desc.text and desc.text.strip()):
                    has_text = True
                # 明示的な画像/図形タイプ（コネクタ図形を含む）
                if lname in ('pic', 'sp', 'graphicframe', 'grpsp', 'cxnsp'):
                    # 最も近いcNvPr子要素の非表示フラグを確認
                    if anchor_is_hidden(desc):
                        continue
                    drawable_types.append(lname)
                # コネクタエンドポイント参照を検出
                if lname in ('stcxn', 'endcxn', 'stcxnpr', 'endcxnpr'):
                    has_connector_ref = True
                # id属性を公開する非cNvPr要素を検出（ヒューリスティック）
                for k in desc.attrib.keys():
                    if k.lower() == 'id' and desc.tag.split('}')[-1].lower() != 'cnvpr':
                        has_connector_ref = True

            result = False
            if drawable_types:
                debug_print(f"[DEBUG] Anchor has drawable elements: {drawable_types}")
                result = True
            elif has_connector_ref:
                debug_print(f"[DEBUG] Anchor has connector references; treating as drawable")
                result = True
            elif has_text:
                debug_print(f"[DEBUG] Anchor contains only text; treating as non-drawable")
                result = False
            else:
                debug_print(f"[DEBUG] Anchor has no drawable elements")
                result = False

            # キャッシュして返す
            cache[key] = result
            return result
        except (ValueError, TypeError):
            return False
    
    def _cluster_shapes_common(self, sheet, shapes, cell_ranges=None, max_groups=2):
        """cell_rangesが利用可能な場合、整数行ギャップによる集中クラスタリング。

        (clusters, debug_dict)を返します。clustersはグループ（インデックスのリスト）のリストです。
        debug_dictには分割決定の追跡に役立つ診断情報が含まれます。
        cell_rangesが提供されないか不十分な場合、重心クラスタリングにフォールバックします。
        """
        try:
            debug = {'method': 'row_gap', 'clusters': None, 'indices_sorted': None, 'chosen_split': None, 'reason': None}
            if not cell_ranges or len(cell_ranges) < len(shapes):
                debug['reason'] = 'no_cell_ranges'
                clusters = self._cluster_shape_indices(shapes, max_groups=max_groups)
                debug['clusters'] = clusters
                return clusters, debug

            # 垂直中点で中心を構築してソート
            row_centers = [(((cr[2] + cr[3]) / 2.0) if (cr[2] is not None and cr[3] is not None) else 0.0, idx) for idx, cr in enumerate(cell_ranges)]
            row_centers.sort(key=lambda x: x[0])
            indices_sorted = [idx for _, idx in row_centers]
            debug['indices_sorted'] = indices_sorted

            # インデックスごとの開始/終了行を計算
            s_rows = []
            e_rows = []
            for idx in indices_sorted:
                try:
                    cr = cell_ranges[idx]
                    s_rows.append(int(cr[2]))
                    e_rows.append(int(cr[3]))
                except (ValueError, TypeError):
                    s_rows.append(None); e_rows.append(None)

            # カバーされた行のセットを構築
            all_covered = set()
            for cr in cell_ranges:
                try:
                    rs = int(cr[2]); re_ = int(cr[3])
                except (ValueError, TypeError):
                    continue
                for rr in range(rs, re_ + 1):
                    all_covered.add(rr)
            debug['all_covered_count'] = len(all_covered)

            # 支配的な大きなスパンをチェック（相対的）
            try:
                row_spans = []
                s_list = [int(cr[2]) for cr in cell_ranges if cr[2] is not None]
                e_list = [int(cr[3]) for cr in cell_ranges if cr[3] is not None]
                for _, idx in row_centers:
                    cr = cell_ranges[idx]
                    row_spans.append(int(cr[3]) - int(cr[2]))
                total_row_span = max(e_list) - min(s_list) if e_list and s_list else 0
                rel_row_span_thresh = 0.75
                dominating = []
                if total_row_span > 0:
                    dominating = [rs for rs in row_spans if (rs / float(total_row_span) >= rel_row_span_thresh)]
                if dominating:
                    debug['reason'] = 'dominating_span'
                    clusters = [indices_sorted]
                    debug['clusters'] = clusters
                    debug['chosen_split'] = None
                    return clusters, debug
            except (ValueError, TypeError) as e:
                debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")

            # 整数の空行が存在する隣接ペア分割を試行
            split_at = None
            chosen_row = None
            total_rows = None
            try:
                # sheet.max_rowの代わりにcell_rangesからtotal_rowsを計算
                e_list = [int(cr[3]) for cr in cell_ranges if cr[3] is not None]
                total_rows = max(e_list) if e_list else None
            except (ValueError, TypeError):
                total_rows = None

            for gi in range(len(indices_sorted) - 1):
                try:
                    left_max = max([int(cell_ranges[i][3]) for i in indices_sorted[:gi+1] if cell_ranges[i][3] is not None])
                    right_min = min([int(cell_ranges[i][2]) for i in indices_sorted[gi+1:] if cell_ranges[i][2] is not None])
                except (ValueError, TypeError):
                    left_max = None; right_min = None
                if left_max is None or right_min is None:
                    continue
                if right_min - left_max >= 2:
                    candidate = left_max + 1
                    if candidate not in all_covered:
                        split_at = gi + 1
                        chosen_row = candidate
                        break

            # フォールバック: 最大の未カバー内部ギャップを検索
            if split_at is None:
                try:
                    if total_rows:
                        uncovered = [r for r in range(1, total_rows+1) if r not in all_covered]
                        if uncovered:
                            # 連続するギャップを構築
                            gaps = []
                            start = uncovered[0]; prev = uncovered[0]
                            for r in uncovered[1:]:
                                if r == prev + 1:
                                    prev = r
                                else:
                                    gaps.append((start, prev)); start = r; prev = r
                            gaps.append((start, prev))
                            gaps.sort(key=lambda x: x[1] - x[0], reverse=True)
                            for gap_start, gap_end in gaps:
                                if gap_start == 1 or gap_end == total_rows:
                                    continue
                                gap_len = (gap_end - gap_start + 1)
                                if gap_len >= 2:
                                    left = []; right = []
                                    for idx in indices_sorted:
                                        try:
                                            s_r = int(cell_ranges[idx][2]); e_r = int(cell_ranges[idx][3])
                                        except (ValueError, TypeError):
                                            s_r = None; e_r = None
                                        if e_r is not None and e_r < gap_start:
                                            left.append(idx)
                                        elif s_r is not None and s_r > gap_end:
                                            right.append(idx)
                                    if left and right:
                                        clusters = [left, right]
                                        debug['chosen_split'] = ('gap', gap_start, gap_end)
                                        debug['clusters'] = clusters
                                        return clusters, debug
                except (ValueError, TypeError) as e:
                    debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")

            if split_at is not None:
                clusters = [indices_sorted[:split_at], indices_sorted[split_at:]]
                debug['chosen_split'] = ('adjacent', chosen_row)
                debug['clusters'] = clusters
                return clusters, debug

            # 有効な整数行分割が見つからない場合、すべての図形を単一クラスタとして返す
            clusters = [list(range(len(shapes)))]
            debug['reason'] = 'no_row_split'
            debug['clusters'] = clusters
            return clusters, debug
        except Exception as e:
            debug_print(f"[DEBUG] クラスタリングエラー: {e}")
            return [[i for i in range(len(shapes))]], {'reason': 'fatal'}
    
    def _get_image_position(self, image):
        """画像の位置情報を取得

        戻り値:
          - 成功時: {'col': int or None, 'row': int} の辞書（1-based 行/列インデックス）
          - 失敗時/不明: 文字列メッセージ（既存ログとの互換性維持）

        呼び出し側は dict を期待しており、dict の場合は 'row' を使って
        画像の代表開始行を決めるロジックがあります。ここで構造化して
        返すことで start_row が 1 固定になる問題を修正します。
        """
        try:
            if hasattr(image, 'anchor'):
                anchor = image.anchor
                # openpyxlアンカーは0ベースのcol/rowを持つ_from属性を公開する場合がある
                if hasattr(anchor, '_from'):
                    try:
                        col_idx = getattr(anchor._from, 'col', None)
                        row_idx = getattr(anchor._from, 'row', None)
                        # 存在する場合は1ベースのインデックスに変換
                        col_val = int(col_idx) + 1 if col_idx is not None else None
                        row_val = int(row_idx) + 1 if row_idx is not None else None
                        if row_val is not None:
                            return {'col': col_val, 'row': row_val}
                    except (ValueError, TypeError):
                        # 文字列フォールバックに移行
                        pass
            return "位置情報なし"
        except (ValueError, TypeError):
            return "位置情報不明"

    def _extract_drawing_shapes(self, sheet) -> List[Tuple[int,int,int,int]]:
        """ワークブックの描画XMLから図形のバウンディングボックスを抽出し、
        ラスタライズに使用されるDPIに一致するピクセル単位に座標を変換します。
        (left, top, right, bottom)タプルのリストを返します。
        """
        try:
            # Phase 1基盤メソッドを使用して描画XMLを取得
            metadata = self._get_drawing_xml_and_metadata(sheet)
            if metadata is None:
                return []
            
            drawing_xml = metadata['drawing_xml']
            # ランタイムDPIを使用して単純な列/行ピクセルマッピングを準備
            DPI = 300
            try:
                DPI = int(getattr(self, 'dpi', DPI) or DPI)
            except (ValueError, TypeError):
                DPI = DPI
            EMU_PER_INCH = 914400
            try:
                EMU_PER_PIXEL = EMU_PER_INCH / float(DPI)
            except (ValueError, TypeError):
                EMU_PER_PIXEL = EMU_PER_INCH / float(DPI)

            max_col = max(sheet.max_column, 100)  # 図形が範囲外にある可能性を考慮
            max_row = max(sheet.max_row, 200)
            col_pixels = []
            for c in range(1, max_col+1):
                cd = sheet.column_dimensions.get(get_column_letter(c))
                width = getattr(cd, 'width', None) if cd is not None else None
                if width is None:
                    width = 8.43
                px = max(1, int(width * 7 + 5))
                col_pixels.append(px)
            row_pixels = []
            for r in range(1, max_row+1):
                rd = sheet.row_dimensions.get(r)
                hpts = getattr(rd, 'height', None) if rd is not None else None
                if hpts is None:
                    hpts = 15
                px = max(1, int(hpts * DPI / 72))
                row_pixels.append(px)
            col_x = [0]
            for v in col_pixels:
                col_x.append(col_x[-1] + v)
            row_y = [0]
            for v in row_pixels:
                row_y.append(row_y[-1] + v)

            ns = {'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'}
            bboxes = []
            # アンカーが描画可能要素を含むかチェックするヘルパー
            # 一貫性のため中央ヘルパーに委譲
            def anchor_has_drawable(a):
                return self._anchor_has_drawable(a)

            # ドキュメント順序でトップレベルの描画子要素を反復処理し、
            # 分離図形トリミングパスで構築されたアンカーリストと順序を一致させます。
            # これにより、図形中心によるクラスタリング時のインデックス/アンカーの
            # 不一致を防ぎます。
            for node in list(drawing_xml):
                lname = node.tag.split('}')[-1].lower()
                if lname not in ('twocellanchor', 'onecellanchor'):
                    continue
                # 描画可能なコンテンツを持つアンカーのみを考慮
                if not anchor_has_drawable(node):
                    continue
                if lname == 'twocellanchor':
                    fr = node.find('xdr:from', ns)
                    to = node.find('xdr:to', ns)
                    if fr is None or to is None:
                        continue
                    try:
                        col = int(fr.find('xdr:col', ns).text)
                        colOff = int(fr.find('xdr:colOff', ns).text)
                        row = int(fr.find('xdr:row', ns).text)
                        rowOff = int(fr.find('xdr:rowOff', ns).text)
                        to_col = int(to.find('xdr:col', ns).text)
                        to_colOff = int(to.find('xdr:colOff', ns).text)
                        to_row = int(to.find('xdr:row', ns).text)
                        to_rowOff = int(to.find('xdr:rowOff', ns).text)
                    except (ValueError, TypeError):
                        continue
                    # 安全な配列アクセス(範囲外チェック)
                    if col < 0 or col >= len(col_x) or row < 0 or row >= len(row_y):
                        continue
                    if to_col < 0 or to_col >= len(col_x) or to_row < 0 or to_row >= len(row_y):
                        continue
                    left = col_x[col] + (colOff / EMU_PER_PIXEL)
                    top = row_y[row] + (rowOff / EMU_PER_PIXEL)
                    right = col_x[to_col] + (to_colOff / EMU_PER_PIXEL)
                    bottom = row_y[to_row] + (to_rowOff / EMU_PER_PIXEL)
                else:
                    fr = node.find('xdr:from', ns)
                    ext = node.find('xdr:ext', ns)
                    if fr is None or ext is None:
                        continue
                    try:
                        col = int(fr.find('xdr:col', ns).text)
                        colOff = int(fr.find('xdr:colOff', ns).text)
                        row = int(fr.find('xdr:row', ns).text)
                        rowOff = int(fr.find('xdr:rowOff', ns).text)
                        cx = int(ext.attrib.get('cx', '0'))
                        cy = int(ext.attrib.get('cy', '0'))
                    except (ValueError, TypeError):
                        continue
                    # 安全な配列アクセス(範囲外チェック)
                    if col < 0 or col >= len(col_x) or row < 0 or row >= len(row_y):
                        continue
                    left = col_x[col] + (colOff / EMU_PER_PIXEL)
                    top = row_y[row] + (rowOff / EMU_PER_PIXEL)
                    right = left + (cx / EMU_PER_PIXEL)
                    bottom = top + (cy / EMU_PER_PIXEL)
                # ページの大部分をカバーするボックスを除外（小さな描画ではない可能性が高い）
                page_w = col_x[-1]
                page_h = row_y[-1]
                try:
                    box_area = max(0, right-left) * max(0, bottom-top)
                    page_area = max(1, page_w * page_h)
                    if box_area / page_area > 0.85:
                        continue
                except Exception as e:
                    print(f"[WARNING] ファイル操作エラー: {e}")
                bboxes.append((left, top, right, bottom))

            # bboxesを返す（ピクセル単位の(left, top, right, bottom)のリスト）
            debug_print(f"[DEBUG] _extract_drawing_shapes found {len(bboxes)} bboxes")
            return bboxes
        except Exception as e:
            print(f"[WARNING] _extract_drawing_shapes exception: {e}")
            import traceback
            traceback.print_exc()
            return []

    def _render_sheet_isolated_group_v2(self, sheet, shape_indices: List[int], dpi: int = 600, cell_range: Optional[Tuple[int,int,int,int]] = None) -> Optional[Tuple[str, int]]:
        """
        Render a group of shape indices as a single isolated image (refactored version).
        
        **EXPERIMENTAL - NOT RECOMMENDED FOR PRODUCTION USE**
        
        This is a streamlined implementation that uses extracted helper methods for:
        - Connector reference resolution (_resolve_connector_references)
        - Drawing anchor pruning (_prune_drawing_anchors)
        
        **MISSING**: Connector cosmetic processing (~600 lines)
        - connector_children_by_id construction
        - Theme color resolution for schemeClr -> srgbClr
        - Line style materialization from lnRef
        - Arrow head/tail preservation
        - Duplicate cosmetic element deduplication
        
        **Result**: Images generated by this method may have:
        - Missing or incorrect connector line styles
        - Wrong connector colors
        - Missing arrow heads/tails
        
        **RECOMMENDATION**: Use the original _render_sheet_isolated_group method for production.
        This v2 method is kept for educational purposes and as a foundation for future refactoring.
        
        Args:
            sheet: Worksheet to render
            shape_indices: List of shape indices to include
            dpi: DPI for rendering (default: 600)
            cell_range: Optional tuple (s_col, e_col, s_row, e_row) to constrain the output
        
        Returns:
            Generated filename (relative to images_dir) or None on failure
        """
        import warnings
        warnings.warn(
            "_render_sheet_isolated_group_v2 is experimental and missing connector cosmetic processing. "
            "Use the original _render_sheet_isolated_group for production.",
            FutureWarning,
            stacklevel=2
        )
        
        try:
            # 保持されたIDマーカーをリセット
            try:
                self._last_iso_preserved_ids = set()
            except Exception:
                pass  # ファイルクローズ失敗は無視
            
            # Excelファイルを開いて描画を検索
            zpath = self.excel_file
            with zipfile.ZipFile(zpath, 'r') as z:
                sheet_index = self.workbook.sheetnames.index(sheet.title)
                rels_path = f"xl/worksheets/_rels/sheet{sheet_index+1}.xml.rels"
                
                # 描画リレーションシップを検索
                rels_xml = get_xml_from_zip(z, rels_path)
                if rels_xml is None:
                    debug_print(f"[DEBUG][_iso_v2] sheet={sheet.title} missing rels")
                    return None
                drawing_target = None
                for rel in rels_xml.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                    if rel.attrib.get('Type', '').endswith('/drawing'):
                        drawing_target = rel.attrib.get('Target')
                        break
                
                if not drawing_target:
                    debug_print(f"[DEBUG][_iso_v2] sheet={sheet.title} no drawing relationship")
                    return None
                
                # 描画パスを正規化
                drawing_path = normalize_excel_path(drawing_target)
                
                if drawing_path not in z.namelist():
                    drawing_path = drawing_path.replace('worksheets', 'drawings')
                    if drawing_path not in z.namelist():
                        debug_print(f"[DEBUG][_iso_v2] drawing_path not found: {drawing_path}")
                        return None
                
                # 描画XMLを解析
                drawing_xml_bytes = z.read(drawing_path)
                drawing_xml = ET.fromstring(drawing_xml_bytes)
            
            # 描画可能要素のみにアンカーをフィルタリング
            anchors = []
            for node in drawing_xml:
                lname = node.tag.split('}')[-1].lower()
                if lname in ('twocellanchor', 'onecellanchor') and self._anchor_has_drawable(node):
                    anchors.append(node)
            
            if not anchors:
                debug_print(f"[DEBUG][_iso_v2] no drawable anchors found")
                return None
            
                # cell_rangeが提供されていない場合は計算
            # このクラスタの最小行も追跡（マークダウン順序付けに使用）
            cluster_min_row = 1  # デフォルトフォールバック
            if cell_range is None and shape_indices:
                try:
                    all_ranges = self._extract_drawing_cell_ranges(sheet)
                    picked = [all_ranges[idx] for idx in shape_indices if 0 <= idx < len(all_ranges)]
                    if picked:
                        s_col = min(r[0] for r in picked)
                        e_col = max(r[1] for r in picked)
                        s_row = min(r[2] for r in picked)
                        e_row = max(r[3] for r in picked)
                        
                        # マークダウン順序付けで後で使用するためにクラスタ最小行を保存
                        cluster_min_row = s_row
                        
                        # 図形が完全に表示されるように10%のパディングを追加
                        # 一部の図形はアンカーポイントを超えて境界線やコネクタが延びる場合がある
                        col_padding = max(2, int((e_col - s_col) * 0.1))
                        row_padding = max(2, int((e_row - s_row) * 0.1))
                        s_col = max(1, s_col - col_padding)
                        e_col = e_col + col_padding
                        s_row = max(1, s_row - row_padding)
                        e_row = e_row + row_padding
                        
                        cell_range = (s_col, e_col, s_row, e_row)
                        debug_print(f"[DEBUG][_iso_v2] Computed cell_range from shapes: cols {s_col}-{e_col}, rows {s_row}-{e_row} (with padding)")
                        debug_print(f"[DEBUG][_iso_v2] Original shape ranges: {picked}")
                except Exception as e:
                    debug_print(f"[DEBUG][_iso_v2] Failed to compute cell_range: {e}")            # shape_indicesからkeep_cnvpr_idsを構築
            keep_cnvpr_ids = set()
            for si in shape_indices:
                if 0 <= si < len(anchors):
                    cid = extract_anchor_id(anchors[si], allow_idx=False)
                    if cid:
                        keep_cnvpr_ids.add(str(cid))
            
            debug_print(f"[DEBUG][_iso_v2] anchors={len(anchors)} keep_ids={sorted(list(keep_cnvpr_ids))}")
            
            # ヘルパーメソッドを使用してコネクタ参照を解決
            referenced_ids = self._resolve_connector_references(
                drawing_xml=drawing_xml,
                anchors=anchors,
                keep_cnvpr_ids=keep_cnvpr_ids
            )
            
            # 呼び出し元のために保持されたIDを公開
            try:
                self._last_iso_preserved_ids = set(referenced_ids)
            except Exception as e:
                print(f"[WARNING] ファイル操作エラー: {e}")
            
            # 一時ディレクトリを作成してワークブックを展開
            tmp_base = tempfile.mkdtemp(prefix='xls2md_iso_v2_base_')
            tmpdir = tempfile.mkdtemp(prefix='xls2md_iso_v2_', dir=tmp_base)
            try:
                with zipfile.ZipFile(zpath, 'r') as zin:
                    zin.extractall(tmpdir)
                
                # 出力に無関係なシートが含まれないように、対象シート以外のすべてのシートを削除
                # これにより、生成されたExcelファイルにはトリミングされた対象シートのみが含まれます
                # 適切な参照を維持するために対象シートの描画ファイルを保持
                try:
                    # 対象シートの描画ファイル名を取得して保持（存在する場合）
                    target_sheet_drawing = None
                    target_sheet_rels_path = os.path.join(tmpdir, f'xl/worksheets/_rels/sheet{sheet_index+1}.xml.rels')
                    if os.path.exists(target_sheet_rels_path):
                        try:
                            rels_tree = ET.parse(target_sheet_rels_path)
                            rels_root = rels_tree.getroot()
                            for rel in rels_root:
                                rel_type = rel.attrib.get('Type', '')
                                if '/drawing' in rel_type:
                                    target_drawing = rel.attrib.get('Target', '')
                                    if target_drawing:
                                        # パスを正規化: ../drawings/drawing1.xml -> drawing1.xml
                                        target_sheet_drawing = os.path.basename(target_drawing)
                                        break
                        except (ET.ParseError, KeyError, AttributeError) as e:
                            debug_print(f"[DEBUG] XML解析エラー（無視）: {type(e).__name__}")
                    
                    # workbook.xmlを解析してシートリレーションシップを取得
                    wb_path = os.path.join(tmpdir, 'xl/workbook.xml')
                    wb_rels_path = os.path.join(tmpdir, 'xl/_rels/workbook.xml.rels')
                    
                    if os.path.exists(wb_path):
                        wb_tree = ET.parse(wb_path)
                        wb_root = wb_tree.getroot()
                        wb_ns = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
                        rel_ns = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
                        
                        # すべてのシートを検索し、対象シートのみを保持
                        sheets_el = wb_root.find(f'{{{wb_ns}}}sheets')
                        if sheets_el is not None:
                            target_sheet_rid = None
                            sheets_to_remove = []
                            
                            for idx, sheet_el in enumerate(list(sheets_el)):
                                if idx == sheet_index:
                                    # これは対象シート - 保持してリレーションシップIDを取得
                                    target_sheet_rid = sheet_el.attrib.get(f'{{{rel_ns}}}id')
                                else:
                                    # 削除対象としてマーク
                                    sheets_to_remove.append((idx, sheet_el))
                            
                            # workbook.xmlから非対象シートを削除
                            for _, sheet_el in sheets_to_remove:
                                sheets_el.remove(sheet_el)
                            
                            # 対象シートをシート1に再番号付け（sheetId="1"）
                            # これにより、Excel/LibreOfficeが最初のシートとして正しく認識します
                            if sheets_el is not None:
                                for sheet_el in list(sheets_el):
                                    # sheetIdを1に設定（最初のシート）
                                    sheet_el.set('sheetId', '1')
                                    # リレーションシップIDをrId1に更新
                                    sheet_el.set(f'{{{rel_ns}}}id', 'rId1')
                            
                            # 変更されたworkbook.xmlを書き戻す
                            wb_tree.write(wb_path, encoding='utf-8', xml_declaration=True)
                            
                            # workbook.xml.relsを解析して保持するリレーションシップIDを検索
                            if os.path.exists(wb_rels_path):
                                rels_tree = ET.parse(wb_rels_path)
                                rels_root = rels_tree.getroot()
                                pkg_rel_ns = 'http://schemas.openxmlformats.org/package/2006/relationships'
                                
                                # 削除するシートリレーションシップターゲットを検索
                                rels_to_remove = []
                                target_sheet_rel = None
                                for rel in list(rels_root):
                                    rid = rel.attrib.get('Id')
                                    target = rel.attrib.get('Target')
                                    rel_type = rel.attrib.get('Type', '')
                                    
                                    # 対象シートのリレーションシップを保持し、他を削除
                                    if rel_type.endswith('/worksheet'):
                                        if rid == target_sheet_rid:
                                            target_sheet_rel = rel
                                        else:
                                            rels_to_remove.append(rel)
                                
                                # 非対象シートのリレーションシップを削除
                                for rel in rels_to_remove:
                                    rels_root.remove(rel)
                                
                                # 対象シートのリレーションシップをrId1に再番号付け
                                if target_sheet_rel is not None:
                                    target_sheet_rel.set('Id', 'rId1')
                                
                                # 変更されたrelsを書き戻す
                                rels_tree.write(wb_rels_path, encoding='utf-8', xml_declaration=True)
                            
                            # 非対象シートの物理シートファイルを削除
                            for idx, _ in sheets_to_remove:
                                # シートXMLファイルを削除
                                sheet_file = os.path.join(tmpdir, f'xl/worksheets/sheet{idx+1}.xml')
                                if os.path.exists(sheet_file):
                                    os.remove(sheet_file)
                                
                                # シートrelsファイルを削除
                                sheet_rels = os.path.join(tmpdir, f'xl/worksheets/_rels/sheet{idx+1}.xml.rels')
                                if os.path.exists(sheet_rels):
                                    os.remove(sheet_rels)
                            
                            # 対象シートの描画以外のすべての描画ファイルを削除
                            # これにより、孤立した描画参照によるエラーを防ぎます
                            drawings_dir = os.path.join(tmpdir, 'xl/drawings')
                            if os.path.exists(drawings_dir):
                                for fname in os.listdir(drawings_dir):
                                    # 対象シートの描画ファイルをスキップ
                                    if target_sheet_drawing and fname == target_sheet_drawing:
                                        continue
                                    
                                    # 他の描画XMLファイルを削除
                                    if fname.endswith('.xml') and not fname.startswith('_rels'):
                                        drawing_file = os.path.join(drawings_dir, fname)
                                        try:

                                            os.remove(p)

                                        except (OSError, FileNotFoundError):

                                            pass  # ファイル削除失敗は無視
                                
                                # 対象シートに属さない描画relsを削除
                                rels_dir = os.path.join(drawings_dir, '_rels')
                                if os.path.exists(rels_dir) and target_sheet_drawing:
                                    target_rels = target_sheet_drawing.replace('.xml', '.xml.rels')
                                    for fname in os.listdir(rels_dir):
                                        if fname != target_rels and fname.endswith('.rels'):
                                            try:
                                                os.remove(os.path.join(rels_dir, fname))
                                            except Exception:
                                                pass  # 一時ファイルの削除失敗は無視
                            
                            debug_print(f"[DEBUG][_iso_v2] Removed {len(sheets_to_remove)} non-target sheets from workbook (kept drawing: {target_sheet_drawing or 'none'})")
                
                except Exception as e:
                    if getattr(self, 'verbose', False):
                        print(f"[WARN][_iso_v2] Failed to remove non-target sheets: {e}")
                        import traceback
                        traceback.print_exc()
                
                # cell_rangeからgroup_rowsを計算
                group_rows = set()
                if cell_range:
                    try:
                        s_col, e_col, s_row, e_row = cell_range
                        group_rows = set(range(s_row, e_row + 1))
                    except (ValueError, TypeError) as e:
                        debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")
                
                # ヘルパーメソッドを使用して描画アンカーを剪定
                drawing_relpath = os.path.join(tmpdir, drawing_path)
                self._prune_drawing_anchors(
                    drawing_relpath=drawing_relpath,
                    keep_cnvpr_ids=keep_cnvpr_ids,
                    referenced_ids=referenced_ids,
                    cell_range=cell_range,
                    group_rows=group_rows
                )
                
                # 重要: 描画座標を調整しない
                # 元のdrawing.xml座標をそのまま保持
                # LibreOfficeは図形を正しくレンダリングするために元の座標が必要
                # セルデータのみをトリミングし、描画位置は変更しない
                debug_print(f"[DEBUG][_iso_v2] Preserving original drawing coordinates (no adjustment)")
                if cell_range:
                    s_col, e_col, s_row, e_row = cell_range
                    debug_print(f"[DEBUG][_iso_v2] Cell range for data trimming: cols {s_col}-{e_col}, rows {s_row}-{e_row}")
                
                # ワークシートを再構築しない - すべての元データを保持
                # これにより、図形が正しく参照できるように元のセル位置が保持されます
                # 描画アンカーのみを剪定し、セルデータは変更しない
                sheet_rel = os.path.join(tmpdir, f"xl/worksheets/sheet{sheet_index+1}.xml")
                
                # ただし、scale=25の縮小を防ぐためにpageSetupを修正する必要がある
                # これはワークシート再構築とは別に行われる
                if os.path.exists(sheet_rel):
                    try:
                        stree = ET.parse(sheet_rel)
                        sroot = stree.getroot()
                        ns = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
                        
                        # 適切なスケーリングでpageSetupを設定
                        # 重要: 既存のpageSetupを削除し、scale=100で新しいものを作成
                        # fitToHeight/fitToWidthは図形を極小サイズに縮小する可能性がある
                        # 既存のすべてのpageSetup要素を削除
                        for old_ps in list(sroot.findall(f'.//{{{ns}}}pageSetup')):
                            sroot.remove(old_ps)
                        
                        # 通常の100%スケールで新しいpageSetupを作成
                        ps = ET.Element(f'{{{ns}}}pageSetup')
                        ps.set('scale', '100')
                        ps.set('paperSize', '1')  # レター（標準）
                        ps.set('orientation', 'portrait')
                        ps.set('pageOrder', 'downThenOver')
                        ps.set('blackAndWhite', 'false')
                        ps.set('draft', 'false')
                        ps.set('cellComments', 'none')
                        ps.set('horizontalDpi', '300')
                        ps.set('verticalDpi', '300')
                        ps.set('copies', '1')
                        # シートの末尾に追加
                        sroot.append(ps)
                        
                        # 変更されたシートを書き戻す
                        stree.write(sheet_rel, encoding='utf-8', xml_declaration=True)
                        debug_print(f"[DEBUG][_iso_v2] Set pageSetup to scale=100 (normal size) to preserve shapes")
                    except Exception as e:
                        if getattr(self, 'verbose', False):
                            print(f"[WARN][_iso_v2] Failed to fix pageSetup: {e}")
                
                # ワークシート再構築コード（無効 - 元のシートデータを保持）
                if False and os.path.exists(sheet_rel) and cell_range:
                    try:
                        s_col, e_col, s_row, e_row = cell_range
                        stree = ET.parse(sheet_rel)
                        sroot = stree.getroot()
                        ns = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
                        
                        # セル値を取得するためにソースExcelファイルから元のsheet.xmlを読み取る
                        with zipfile.ZipFile(self.excel_file, 'r') as src_z:
                            src_sheet_path = f"xl/worksheets/sheet{sheet_index+1}.xml"
                            src_sheet_xml = get_xml_from_zip(src_z, src_sheet_path)
                            if src_sheet_xml is not None:
                                src_sheet_data = src_sheet_xml.find(f'{{{ns}}}sheetData')
                            else:
                                src_sheet_data = None
                        
                        # 範囲内の行/列のみを含むようにsheetDataを再構築
                        # 元の行/列番号を保持（1から再番号付けしない）
                        sheet_data_tag = f'{{{ns}}}sheetData'
                        sheet_data = sroot.find(sheet_data_tag)
                        if sheet_data is not None and src_sheet_data is not None:
                            new_sheet_data = ET.Element(sheet_data_tag)
                            src_rows = src_sheet_data.findall(f'{{{ns}}}row')
                            debug_print(f"[DEBUG][_iso_v2] Found {len(src_rows)} rows in source sheet.xml")
                            cells_copied = 0
                            
                            # 範囲内の行をコピーし、元の行番号を保持
                            for row_el in src_rows:
                                try:
                                    rnum = int(row_el.attrib.get('r', '0'))
                                except (ValueError, TypeError):
                                    continue
                                if rnum < s_row or rnum > e_row:
                                    continue
                                
                                # 元の行番号で新しい行を作成
                                new_row = ET.Element(f'{{{ns}}}row')
                                new_row.set('r', str(rnum))  # 元の行番号を保持
                                
                                # 行属性をコピー
                                for attr in ('ht', 'hidden', 'customHeight'):
                                    if attr in row_el.attrib:
                                        new_row.set(attr, row_el.attrib.get(attr))
                                
                                # 列範囲内のセルをコピーし、元の列文字を保持
                                for c in list(row_el):
                                    if c.tag.split('}')[-1] != 'c':
                                        continue
                                    cell_r = c.attrib.get('r', '')
                                    col_letters = ''.join([ch for ch in cell_r if ch.isalpha()]) if cell_r else None
                                    if not col_letters:
                                        continue
                                    
                                    # 列文字をインデックスに変換
                                    col_idx = 0
                                    for ch in col_letters:
                                        col_idx = col_idx * 26 + (ord(ch.upper()) - 64)
                                    if col_idx < s_col or col_idx > e_col:
                                        continue
                                    
                                    # 元のセル参照でセルをコピー（例: "D17"）
                                    import copy
                                    new_cell = copy.deepcopy(c)
                                    new_row.append(new_cell)
                                    cells_copied += 1
                                
                                if len(new_row) > 0:  # セルがある場合のみ行を追加
                                    new_sheet_data.append(new_row)
                            
                            debug_print(f"[DEBUG][_iso_v2] Copied {cells_copied} cells with original row/col numbers")
                            
                            # 古いsheetDataを新しいものに置き換え
                            for child in list(sroot):
                                if child.tag == sheet_data_tag:
                                    sroot.remove(child)
                            sroot.append(new_sheet_data)
                            
                            # 元の行/列番号でdimension要素を更新
                            dim_tag = f'{{{ns}}}dimension'
                            dim = sroot.find(dim_tag)
                            if dim is None:
                                dim = ET.Element(dim_tag)
                                sroot.insert(0, dim)
                            # 元の行/列番号を使用
                            start_addr = f"{col_letter(s_col)}{s_row}"
                            end_addr = f"{col_letter(e_col)}{e_row}"
                            dim.set('ref', f"{start_addr}:{end_addr}")
                        
                        # 元の列番号でcols要素を再構築
                        cols_tag = f'{{{ns}}}cols'
                        col_tag = f'{{{ns}}}col'
                        for child in list(sroot):
                            if child.tag == cols_tag:
                                try:
                                    sroot.remove(child)
                                except Exception:
                                    pass  # 一時ファイルの削除失敗は無視
                        
                        cols_el = ET.Element(cols_tag)
                        try:
                            from openpyxl.utils import get_column_letter
                            default_col_w = getattr(sheet.sheet_format, 'defaultColWidth', None) or 8.43
                            for c in range(s_col, e_col + 1):
                                cd = sheet.column_dimensions.get(get_column_letter(c))
                                width = getattr(cd, 'width', None) if cd else None
                                hidden = getattr(cd, 'hidden', None) if cd else None
                                if width is None:
                                    width = default_col_w
                                
                                col_el = ET.Element(col_tag)
                                # 元の列インデックスを使用（再番号付けしない）
                                col_el.set('min', str(c))
                                col_el.set('max', str(c))
                                col_el.set('width', str(float(width)))
                                if cd and getattr(cd, 'width', None) is not None:
                                    col_el.set('customWidth', '1')
                                if hidden:
                                    col_el.set('hidden', '1')
                                cols_el.append(col_el)
                        except (ValueError, TypeError):
                            # フォールバック: 元の列番号でデフォルト幅を設定
                            for c in range(s_col, e_col + 1):
                                col_el = ET.Element(col_tag)
                                col_el.set('min', str(c))
                                col_el.set('max', str(c))
                                col_el.set('width', '8.43')
                                cols_el.append(col_el)
                        
                        # sheetDataの前にcolsを挿入
                        sd_idx = list(sroot).index(new_sheet_data)
                        sroot.insert(sd_idx, cols_el)
                        
                        # ページマージンをゼロに設定（元のメソッドと同じ）
                        # pageSetupPr fitToPage属性でsheetPrを追加または変更
                        sheet_pr = sroot.find(f'.//{{{ns}}}sheetPr')
                        if sheet_pr is None:
                            sheet_pr = ET.Element(f'{{{ns}}}sheetPr')
                            sroot.insert(0, sheet_pr)
                        page_setup_pr = sheet_pr.find(f'{{{ns}}}pageSetUpPr')
                        if page_setup_pr is None:
                            page_setup_pr = ET.SubElement(sheet_pr, f'{{{ns}}}pageSetUpPr')
                        page_setup_pr.set('fitToPage', '1')
                        
                        # printOptionsを追加または変更
                        print_opts = sroot.find(f'.//{{{ns}}}printOptions')
                        if print_opts is None:
                            print_opts = ET.Element(f'{{{ns}}}printOptions')
                            sroot.append(print_opts)
                        print_opts.set('horizontalCentered', '1')
                        print_opts.set('verticalCentered', '1')
                        
                        # 適切なスケーリングでpageSetupを設定
                        # 重要: 既存のpageSetupを削除し、scale=100で新しいものを作成
                        # fitToHeight/fitToWidthは図形を極小サイズに縮小する可能性がある
                        # 既存のすべてのpageSetup要素を削除
                        for old_ps in list(sroot.findall(f'.//{{{ns}}}pageSetup')):
                            sroot.remove(old_ps)
                        
                        # 通常の100%スケールで新しいpageSetupを作成
                        ps = ET.Element(f'{{{ns}}}pageSetup')
                        ps.set('scale', '100')
                        ps.set('paperSize', '1')  # レター（標準）
                        ps.set('orientation', 'portrait')
                        ps.set('pageOrder', 'downThenOver')
                        ps.set('blackAndWhite', 'false')
                        ps.set('draft', 'false')
                        ps.set('cellComments', 'none')
                        ps.set('horizontalDpi', '300')
                        ps.set('verticalDpi', '300')
                        ps.set('copies', '1')
                        # シートの末尾に追加
                        sroot.append(ps)
                        debug_print(f"[DEBUG][_iso_v2] Set pageSetup to scale=100 (normal size) to preserve shapes")
                        
                        # ページマージンを設定（属性として、標準Excelフォーマット）
                        pm_tag = f'{{{ns}}}pageMargins'
                        pm = sroot.find(pm_tag)
                        if pm is None:
                            pm = ET.Element(pm_tag)
                            sroot.append(pm)
                        pm.set('left', '0.25')
                        pm.set('right', '0.25')
                        pm.set('top', '0.75')
                        pm.set('bottom', '0.75')
                        pm.set('header', '0.3')
                        pm.set('footer', '0.3')
                        
                        # ヘッダー/フッター要素を削除
                        hf_tag = f'{{{ns}}}headerFooter'
                        for hf in list(sroot.findall(hf_tag)):
                            sroot.remove(hf)
                        
                        stree.write(sheet_rel, encoding='utf-8', xml_declaration=True)
                        debug_print(f"[DEBUG][_iso_v2] Reconstructed sheet data: kept original rows {s_row}-{e_row}, cols {s_col}-{e_col}")
                    except Exception as e:
                        if getattr(self, 'verbose', False):
                            print(f"[WARN][_iso_v2] Failed to reconstruct worksheet: {e}")

                # 重要: すべての図形が表示されるようにPrint_Areaを完全に削除
                # 印刷範囲は表示領域を制限し、定義された範囲外の図形を非表示にする可能性がある
                # 完全なシート構造を保持しているため、Print_Areaは不要
                try:
                    wb_rel = os.path.join(tmpdir, 'xl/workbook.xml')
                    if os.path.exists(wb_rel):
                        wtree = ET.parse(wb_rel)
                        wroot = wtree.getroot()
                        ns = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
                        
                        # definedNames要素を検索
                        dn_tag = f'{{{ns}}}definedNames'
                        dn = wroot.find(dn_tag)
                        
                        # 表示の問題を防ぐためにすべての定義名（Print_Areaを含む）を削除
                        if dn is not None:
                            wroot.remove(dn)
                            debug_print(f"[DEBUG][_iso_v2] Removed Print_Area and all defined names to ensure shapes are visible")
                        
                        wtree.write(wb_rel, encoding='utf-8', xml_declaration=True)
                except Exception as e:
                    if getattr(self, 'verbose', False):
                        print(f"[WARN][_iso_v2] Failed to remove Print_Area: {e}")

                # デバッグ用にトリミングされたワークブックZIPを作成（出力ディレクトリに保存）
                debug_xlsx_filename = f"{self.base_name}_{sheet.title}_group_{shape_indices[0] if shape_indices else 0}_debug.xlsx"
                debug_xlsx_path = os.path.join(self.output_dir, debug_xlsx_filename)
                debug_zip_base = os.path.join(self.output_dir, f"{self.base_name}_{sheet.title}_group_{shape_indices[0] if shape_indices else 0}_debug")
                
                try:
                    # 古いファイルが存在する場合は削除
                    if os.path.exists(debug_xlsx_path):
                        os.remove(debug_xlsx_path)
                    if os.path.exists(debug_zip_base + '.zip'):
                        os.remove(debug_zip_base + '.zip')
                    
                    shutil.make_archive(debug_zip_base, 'zip', tmpdir)
                    shutil.move(debug_zip_base + '.zip', debug_xlsx_path)
                    debug_print(f"[DEBUG][_iso_v2] Saved debug workbook: {debug_xlsx_path}")
                except Exception as e:
                    if getattr(self, 'verbose', False):
                        print(f"[WARN][_iso_v2] Failed to create trimmed workbook: {e}")
                    return None

                # PDFとPNGに変換（デバッグ用にPDFを保存）
                try:
                    # fit-to-pageを適用しない - 図形を25%に縮小して見えなくなる
                    # pageSetupは上記のワークシートXMLで既に適切に設定されている
                    # self._set_excel_fit_to_one_page(debug_xlsx_path)  # 無効
                    
                    # PDFに変換（xlsxと同じディレクトリに出力）
                    cmd = [LIBREOFFICE_PATH, '--headless', '--convert-to', 'pdf', '--outdir', self.output_dir, debug_xlsx_path]
                    debug_print(f"[DEBUG][_iso_v2] LibreOffice command: {' '.join(cmd)}")
                    proc = subprocess.run(cmd, capture_output=True, text=True, timeout=90)
                    
                    if proc.returncode != 0:
                        if getattr(self, 'verbose', False):
                            print(f"[WARN][_iso_v2] LibreOffice PDF conversion failed: {proc.stderr}")
                        return None
                    
                    # 生成されたPDFを検索
                    debug_pdf_filename = debug_xlsx_filename.replace('.xlsx', '.pdf')
                    pdf_path = os.path.join(self.output_dir, debug_pdf_filename)
                    
                    if not os.path.exists(pdf_path):
                        # 作成されたPDFを検索
                        pdf_candidates = [f for f in os.listdir(self.output_dir) 
                                        if f.lower().endswith('.pdf') and 'group' in f and sheet.title in f]
                        if not pdf_candidates:
                            if getattr(self, 'verbose', False):
                                print("[WARN][_iso_v2] PDF conversion failed - no output")
                            return None
                        pdf_path = os.path.join(self.output_dir, pdf_candidates[0])
                    
                    debug_print(f"[DEBUG][_iso_v2] Saved debug PDF: {pdf_path}")
                    
                    # PDFをPNGに変換（最終出力はimagesディレクトリ）
                    png_filename = f"{self.base_name}_{sheet.title}_group_{shape_indices[0] if shape_indices else 0}.png"
                    png_path = os.path.join(self.images_dir, png_filename)
                    
                    # 古いPNGが存在する場合は削除
                    if os.path.exists(png_path):
                        os.remove(png_path)
                    
                    # ImageMagick: 透明/黒い領域を防ぐために-background white -flattenを使用
                    # -flattenはすべてのレイヤーを白い背景に合成
                    cmd = ['convert', '-density', str(dpi), f'{pdf_path}[0]', 
                           '-background', 'white', '-flatten',
                           '-colorspace', 'sRGB', '-quality', str(IMAGE_QUALITY), png_path]
                    debug_print(f"[DEBUG][_iso_v2] ImageMagick command: {' '.join(cmd)}")
                    proc = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
                    
                    if proc.returncode != 0 or not os.path.exists(png_path):
                        if getattr(self, 'verbose', False):
                            print(f"[WARN][_iso_v2] ImageMagick PNG conversion failed: {proc.stderr}")
                        return None
                    
                    debug_print(f"[DEBUG][_iso_v2] Successfully rendered group: {png_filename}")
                    debug_print(f"[DEBUG][_iso_v2] Debug files: {debug_xlsx_filename}, {debug_pdf_filename}")
                    
                    # コネクタを保持しながら余分な空白を削除するために画像をクロップ
                    # 分離グループにはより厳密なクロップを使用（より積極的なwhite_thresh=250）
                    try:
                        # より高い白しきい値でより積極的なクロップ
                        from PIL import Image
                        if os.path.exists(png_path):
                            im = Image.open(png_path)
                            bbox = self._find_content_bbox(im, white_thresh=250)
                            if bbox:
                                l, t, r, b = bbox
                                # 分離グループの最小パディング（図形は既に適切なマージンを持っている）
                                pad = max(4, int(dpi / 300.0 * 6))  # 通常のパディングの半分
                                l = max(0, l - pad)
                                t = max(0, t - pad)
                                r = min(im.width, r + pad)
                                b = min(im.height, b + pad)
                                cropped = im.crop((l, t, r, b))
                                cropped.save(png_path)
                                cropped.close()
                                debug_print(f"[DEBUG][_iso_v2] Cropped image: {im.size} → {cropped.size}")
                            im.close()
                    except Exception as crop_err:
                        if getattr(self, 'verbose', False):
                            print(f"[WARN][_iso_v2] Failed to crop image: {crop_err}")
                    
                    # タプルを返す: (ファイル名, クラスタの最小行)
                    debug_print(f"[DEBUG][_iso_v2] Returning: filename={png_filename}, cluster_min_row={cluster_min_row}")
                    return (png_filename, cluster_min_row)
                    
                except Exception as e:
                    if getattr(self, 'verbose', False):
                        print(f"[ERROR][_iso_v2] Conversion failed: {e}")
                    return None
                
            finally:
                try:
                    if tmpdir and os.path.exists(tmpdir):
                        shutil.rmtree(tmpdir, ignore_errors=True)
                    if tmp_base and os.path.exists(tmp_base):
                        shutil.rmtree(tmp_base, ignore_errors=True)
                except (ValueError, TypeError) as e:
                    debug_print(f"[DEBUG] 型変換エラー（無視）: {e}")
        
        except Exception as e:
            print(f"[ERROR][_iso_v2] Exception: {e}")
            import traceback
            traceback.print_exc()
            return None

    def _render_sheet_isolated_group(self, sheet, shape_indices: List[int], dpi: int = 600, cell_range: Optional[Tuple[int,int,int,int]] = None) -> Optional[Tuple[str, int]]:
        """Render a group of shape indices as a single isolated image.
        
        **PRODUCTION METHOD - RECOMMENDED FOR ALL USE CASES**
        
        This implementation delegates to IsolatedGroupRenderer class for better code organization.
        The original monolithic implementation has been refactored into separate phases.
        
        Features:
        - Complete connector cosmetic processing
        - Theme color resolution (schemeClr -> srgbClr)
        - Line style materialization from lnRef references
        - Arrow head/tail preservation
        - Robust handling of complex flowcharts
        
        Creates a temporary workbook containing only the specified drawing anchors
        and renders them together as a single PNG image using LibreOffice -> PDF -> ImageMagick.
        This preserves the spatial relationships between shapes, making it ideal for
        flowcharts and composite diagrams.
        
        Returns:
            Optional[Tuple[str, int]]: (filename, start_row) or None on failure
        """
        renderer = IsolatedGroupRenderer(self)
        return renderer.render(sheet, shape_indices, dpi, cell_range)
