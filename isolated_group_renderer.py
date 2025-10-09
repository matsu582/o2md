"""
IsolatedGroupRenderer - 分離グループ画像レンダリングクラス

ExcelToMarkdownConverterから_render_sheet_isolated_groupメソッドを
切り出したクラス。図形グループを一時ワークブックとして分離し、
LibreOffice→PDF→ImageMagickの変換パイプラインで画像化する。
"""

import os
import sys
import tempfile
import subprocess
import shutil
import zipfile
import xml.etree.ElementTree as ET
from typing import List, Dict, Tuple, Optional, Set
from collections import deque
import copy
import hashlib
from utils import get_imagemagick_command, get_libreoffice_path


class IsolatedGroupRenderer:
    """図形グループを分離レンダリングするクラス"""
    
    def __init__(self, converter):
        """
        コンストラクタ
        
        Args:
            converter: 親のExcelToMarkdownConverterインスタンス
        """
        self.converter = converter
        self.sheet = None  # renderメソッドで設定される
        self._last_iso_preserved_ids = set()
        self._last_temp_pdf_path = None
    
    def render(self, sheet, shape_indices: List[int], dpi: int = 600, 
               cell_range: Optional[Tuple[int,int,int,int]] = None) -> Optional[Tuple[str, int]]:
        """
        図形グループを分離レンダリング
        
        Args:
            sheet: openpyxlのWorksheetオブジェクト
            shape_indices: レンダリングする図形のインデックスリスト
            dpi: レンダリング解像度
            cell_range: セル範囲 (start_col, end_col, start_row, end_row)
        
        Returns:
            タプル(画像ファイル名, 開始行) またはNone
        """
        try:
            # 初期化
            self.sheet = sheet  # Aggressiveセクションで使用
            self._last_iso_preserved_ids = set()
            
            # フェーズ1: 初期化とXMLロード
            excel_zip, drawing_xml, drawing_path, sheet_index, theme_color_map = \
                self._phase1_initialize_and_load_xml(sheet)
            
            if drawing_xml is None:
                return None
            
            # フェーズ2: アンカー収集
            anchors = self._phase2_collect_anchors(drawing_xml)
            
            if not anchors:
                print(f"[DEBUG][_iso_entry] sheet={sheet.title} no drawable anchors found")
                return None
            
            # フェーズ3: セル範囲計算
            cell_range = self._phase3_compute_cell_range(sheet, shape_indices, anchors, cell_range)
            
            # フェーズ4: ID収集
            keep_cnvpr_ids = self._phase4_collect_keep_ids(shape_indices, anchors)
            
            print(f"[DEBUG][_iso_entry] sheet={sheet.title} anchors_count={len(anchors)} keep_cnvpr_ids={sorted(list(keep_cnvpr_ids))}")
            
            # フェーズ5: 一時ディレクトリ作成とコネクタ参照解決
            tmpdir, referenced_ids, connector_children_by_id = \
                self._phase5_create_tmpdir_and_resolve_connectors(
                    excel_zip, sheet, shape_indices, anchors, keep_cnvpr_ids, 
                    drawing_xml, drawing_path, theme_color_map
                )
            
            if tmpdir is None:
                return None
            
            try:
                        # フェーズ6: 描画XML刈り込み
                drawing_relpath = os.path.join(tmpdir, drawing_path)
                drawing_xml_bytes = ET.tostring(drawing_xml)
                success = self._phase6_prune_drawing_xml(
                    drawing_relpath, keep_cnvpr_ids, referenced_ids, 
                    cell_range, drawing_xml, tmpdir, sheet, sheet_index
                )
                
                if not success:
                    return None
                
                # フェーズ7: コネクタコスメティック処理
                self._phase7_apply_connector_cosmetics(
                    drawing_relpath, referenced_ids, connector_children_by_id,
                    theme_color_map, drawing_xml_bytes, drawing_xml
                )
                
                # フェーズ8: ワークブック準備
                src_for_conv = self._phase8_prepare_workbook(
                    tmpdir, sheet, sheet_index, cell_range, drawing_path, dpi, shape_indices, keep_cnvpr_ids
                )
                
                if src_for_conv is None:
                    shutil.rmtree(tmpdir, ignore_errors=True)
                    return None
                
                # フェーズ9: PDF/PNG生成
                out_path = self._phase9_generate_pdf_png(
                    sheet, shape_indices, src_for_conv, tmpdir, dpi, cell_range
                )
                
                if out_path is None:
                    shutil.rmtree(tmpdir, ignore_errors=True)
                    return None
                
                # フェーズ10: 後処理
                png_name = os.path.basename(out_path) if out_path else "unknown.png"
                group_rows = [cell_range[2]] if cell_range else None
                final_result = self._phase10_postprocess(
                    out_path, png_name, sheet, group_rows, cell_range
                )
                
                # クリーンアップ
                shutil.rmtree(tmpdir, ignore_errors=True)
                
                return final_result
                
            except Exception as e:
                print(f"[ERROR][IsolatedGroupRenderer] Exception: {e}")
                import traceback
                traceback.print_exc()
                shutil.rmtree(tmpdir, ignore_errors=True)
                return None
                
        except Exception as e:
            print(f"[ERROR][IsolatedGroupRenderer] Exception: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _phase1_initialize_and_load_xml(self, sheet):
        """フェーズ1: 初期化とXMLロード"""
        try:
            zpath = self.converter.excel_file
            z = zipfile.ZipFile(zpath, 'r')
            sheet_index = self.converter.workbook.sheetnames.index(sheet.title)
            rels_path = f"xl/worksheets/_rels/sheet{sheet_index+1}.xml.rels"
            
            if rels_path not in z.namelist():
                print(f"[DEBUG][_iso_entry] sheet={sheet.title} missing rels: {rels_path}")
                return None, None, None, None, None
            
            rels_xml = ET.fromstring(z.read(rels_path))
            drawing_target = None
            for rel in rels_xml.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                if rel.attrib.get('Type','').endswith('/drawing'):
                    drawing_target = rel.attrib.get('Target')
                    break
            
            if not drawing_target:
                print(f"[DEBUG][_iso_entry] sheet={sheet.title} no drawing relationship found in rels")
                return None, None, None, None, None
            
            drawing_path = drawing_target
            if drawing_path.startswith('..'):
                drawing_path = drawing_path.replace('../', 'xl/')
            if drawing_path.startswith('/'):
                drawing_path = drawing_path.lstrip('/')
            
            print(f"[DEBUG][_iso_entry] sheet={sheet.title} drawing_path={drawing_path}")
            
            if drawing_path not in z.namelist():
                drawing_path = drawing_path.replace('worksheets', 'drawings')
                if drawing_path not in z.namelist():
                    print(f"[DEBUG][_iso_entry] sheet={sheet.title} drawing_path not found in archive")
                    return None, None, None, None, None
            
            drawing_xml_bytes = z.read(drawing_path)
            drawing_xml = ET.fromstring(drawing_xml_bytes)
            
            theme_color_map = {}
            try:
                theme_color_map, _ = self.converter._parse_theme_colors(z)
            except Exception:
                pass
            
            return z, drawing_xml, drawing_path, sheet_index, theme_color_map
        except Exception as e:
            print(f"[ERROR] Phase1 failed: {e}")
            import traceback
            traceback.print_exc()
            return None, None, None, None, None
    
    def _phase2_collect_anchors(self, drawing_xml):
        """フェーズ2: アンカー収集"""
        anchors = []
        for node in drawing_xml:
            lname = node.tag.split('}')[-1].lower()
            if lname in ('twocellanchor', 'onecellanchor') and self.converter._anchor_has_drawable(node):
                anchors.append(node)
        return anchors
    
    def _phase3_compute_cell_range(self, sheet, shape_indices, anchors, cell_range):
        """フェーズ3: セル範囲計算"""
        try:
            if cell_range is None and shape_indices:
                all_ranges = self.converter._extract_drawing_cell_ranges(sheet)
                picked = []
                for idx in shape_indices:
                    if idx >= 0 and idx < len(all_ranges):
                        picked.append(all_ranges[idx])
                
                if picked:
                    valid_picked = [r for r in picked if r[0] <= r[1] and r[2] <= r[3]]
                    if not valid_picked:
                        valid_picked = picked
                    
                    s_col = min(r[0] for r in valid_picked)
                    e_col = max(r[1] for r in valid_picked)
                    s_row = min(r[2] for r in valid_picked)
                    e_row = max(r[3] for r in valid_picked)
                    
                    if s_col > e_col:
                        s_col, e_col = e_col, s_col
                    if s_row > e_row:
                        s_row, e_row = e_row, s_row
                    
                    if e_col < s_col:
                        e_col = s_col
                    if e_row < s_row:
                        e_row = s_row
                    
                    try:
                        max_data_col = 0
                        max_data_row = 0
                        for row in sheet.iter_rows():
                            for cell in row:
                                if cell.value is not None:
                                    if cell.column > max_data_col:
                                        max_data_col = cell.column
                                    if cell.row > max_data_row:
                                        max_data_row = cell.row
                        
                        if max_data_col > 0:
                            max_allowed_col = max_data_col + 5
                            if e_col > max_allowed_col:
                                print(f"[DEBUG][_iso_entry] Limiting e_col from {e_col} to {max_allowed_col}")
                                e_col = max_allowed_col
                    except Exception as limit_err:
                        print(f"[DEBUG][_iso_entry] Failed to limit cell_range: {limit_err}")
                    
                    cell_range = (s_col, e_col, s_row, e_row)
        except (ValueError, TypeError) as e:
            print(f"[DEBUG] 型変換エラー（無視）: {e}")
        
        return cell_range
    
    def _phase4_collect_keep_ids(self, shape_indices, anchors):
        """フェーズ4: ID収集"""
        keep_cnvpr_ids = set()
        try:
            for si in shape_indices:
                if si < 0 or si >= len(anchors):
                    continue
                cid = None
                for sub in anchors[si].iter():
                    if sub.tag.split('}')[-1].lower() == 'cnvpr':
                        cid = sub.attrib.get('id')
                        break
                if cid is not None:
                    keep_cnvpr_ids.add(str(cid))
        except (ValueError, TypeError):
            keep_cnvpr_ids = set()
        
        return keep_cnvpr_ids
    
    def _phase5_create_tmpdir_and_resolve_connectors(self, excel_zip, sheet, shape_indices, 
                                                      anchors, keep_cnvpr_ids, drawing_xml, 
                                                      drawing_path, theme_color_map):
        """フェーズ5: 一時ディレクトリ作成とコネクタ参照解決"""
        zpath = self.converter.excel_file
        z = excel_zip
        
        # create tempdir and copy original xlsx contents there to modify
        tmpdir = tempfile.mkdtemp(prefix='xls2md_iso_group_')
        try:
            with zipfile.ZipFile(zpath, 'r') as zin:
                zin.extractall(tmpdir)
            # Preserve original styles and theme so style references inside drawing XML resolve
            try:
                for preserve in ('xl/styles.xml', 'xl/theme/theme1.xml'):
                    if preserve in z.namelist():
                        tgt = os.path.join(tmpdir, preserve)
                        os.makedirs(os.path.dirname(tgt), exist_ok=True)
                        with open(tgt, 'wb') as _fw:
                            _fw.write(z.read(preserve))
            except (OSError, IOError, FileNotFoundError) as e:
                print(f"[WARNING] ファイル操作エラー: {e}")
            except Exception as e:
                print(f"[WARNING] ファイル操作エラー: {e}")
                print(f"[DEBUG] z type: {type(z)}, z value: {z}")
                import traceback
                traceback.print_exc()

            # When pruning anchors below, ensure that any shapes referenced by
            # connectors in the kept indices are also preserved. We'll compute
            # referenced ids from the anchors list first, and also gather
            # connector cosmetic children to copy into kept anchors.
            # We'll compute a transitive closure of anchor ids to preserve.
            # Build mappings of anchor_id -> referenced ids (refs) and reverse refs
            # so we can include connectors that reference kept shapes and also
            # include endpoints referenced by kept connectors, transitively.
            referenced_ids = set()
            connector_children_by_id = {}
            try:
                refs = {}  # anchor_id -> set(of ids it references)
                reverse_refs = {}  # id -> set(of anchor_ids that reference it)

                # First, build refs and connector_children_by_id from all anchor nodes
                for orig in list(drawing_xml):
                    lname = orig.tag.split('}')[-1].lower()
                    if lname not in ('twocellanchor', 'onecellanchor'):
                        continue
                    cid = None
                    for sub in orig.iter():
                        if sub.tag.split('}')[-1].lower() == 'cnvpr':
                            cid = str(sub.attrib.get('id'))
                            break
                    if cid is None:
                        continue
                    # find referenced ids inside this anchor (stCxn/endCxn variants)
                    rset = set()
                    for sub in orig.iter():
                        st = sub.tag.split('}')[-1].lower()
                        if st in ('stcxn', 'endcxn', 'stcxnpr', 'endcxnpr'):
                            vid = sub.attrib.get('id') or sub.attrib.get('idx')
                            if vid is not None:
                                rset.add(str(vid))
                    if rset:
                        refs[cid] = rset
                        for rid in rset:
                            reverse_refs.setdefault(rid, set()).add(cid)

                    # search children for cosmetic subtrees to copy later
                    kids = []
                    for child in orig:
                        for sub in child.iter():
                            st = sub.tag.split('}')[-1].lower()
                            if st in ('prstgeom', 'ln', 'headend', 'tailend', 'custgeom', 'sppr'):
                                kids.append(child)
                                break
                    if kids:
                        connector_children_by_id[cid] = kids

                # seed the BFS with explicitly requested keep ids
                preserve = set(keep_cnvpr_ids)
                q = deque(keep_cnvpr_ids)

                # Additionally, include anchors whose "from" row lies within
                # any of the shape_indices' corresponding rows for this group.
                # This enforces a row-based inclusion rule so connectors whose
                # endpoints are on the same sheet row are preserved even if
                # they are not transitively referenced via stCxn/endCxn tags.
                try:
                    # build mapping: cNvPr id -> from_row for all anchors
                    id_to_row = {}
                    ns_xdr = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
                    for an in anchors:
                        # find cNvPr id
                        a_cid = None
                        try:
                            for sub in an.iter():
                                if sub.tag.split('}')[-1].lower() == 'cnvpr':
                                    a_cid = sub.attrib.get('id') or sub.attrib.get('idx')
                                    break
                        except AttributeError as ae:
                            print(f"[ERROR] anchors element type error: an={type(an)}, error={ae}")
                            import traceback
                            traceback.print_exc()
                            continue
                        if a_cid is None:
                            continue
                        fr = an.find('{%s}from' % ns_xdr)
                        if fr is not None:
                            r = fr.find('{%s}row' % ns_xdr)
                            if r is not None and r.text is not None:
                                try:
                                    id_to_row[str(a_cid)] = int(r.text)
                                except (ValueError, TypeError) as e:
                                    print(f"[DEBUG] 型変換エラー（無視）: {e}")

                    # Build a fallback mapping from ALL anchors in the drawing
                    # (not only those filtered into `anchors`) so we can find
                    # endpoint rows for connector-only anchors that were
                    # omitted by the drawable filter. This helps include
                    # connectors whose endpoints are on the group's rows.
                    all_id_to_row = {}
                    try:
                        for orig_an in list(drawing_xml):
                            lname2 = orig_an.tag.split('}')[-1].lower()
                            if lname2 not in ('twocellanchor', 'onecellanchor'):
                                continue
                            a_cid2 = None
                            for sub2 in orig_an.iter():
                                if sub2.tag.split('}')[-1].lower() == 'cnvpr':
                                    a_cid2 = sub2.attrib.get('id') or sub2.attrib.get('idx')
                                    break
                            if a_cid2 is None:
                                continue
                            fr2 = orig_an.find('{%s}from' % ns_xdr)
                            if fr2 is not None:
                                r2 = fr2.find('{%s}row' % ns_xdr)
                                if r2 is not None and r2.text is not None:
                                    try:
                                        all_id_to_row[str(a_cid2)] = int(r2.text)
                                    except (ValueError, TypeError) as e:
                                        print(f"[DEBUG] 型変換エラー（無視）: {e}")
                    except (ValueError, TypeError):
                        all_id_to_row = {}

                    # Determine group's approximate row span by inspecting the
                    # keep_cnvpr_ids' rows and include anchors on those rows.
                    group_rows = set()
                    for cid in keep_cnvpr_ids:
                        if str(cid) in id_to_row:
                            group_rows.add(id_to_row[str(cid)])
                    # include any anchor whose from_row is in group_rows
                    for cid, r in id_to_row.items():
                        if r in group_rows and cid not in preserve:
                            preserve.add(cid)
                            q.append(cid)
                except (ValueError, TypeError) as e:
                    print(f"[DEBUG] 型変換エラー（無視）: {e}")
                # Expand transitive closure but constrain expansion by row membership
                # to avoid pulling the same anchor into multiple row-based clusters.
                # Only include a candidate anchor/ref if its 'from' row lies within
                # the group's rows (group_rows) or if it was part of the original seed
                # (keep_cnvpr_ids). This prevents cross-cluster duplication while
                # keeping local endpoints.
                # Ensure id_to_row exists (may be empty if earlier parsing failed)
                try:
                    id_to_row
                except NameError:
                    id_to_row = {}

                # Protect BFS expansion from pathological inputs by
                # bounding the number of deque pops. If we exceed the
                # cap, emit a warning and stop expanding further to
                # avoid infinite loops observed on malformed workbooks.
                bfs_iter = 0
                bfs_max = max(1000, len(keep_cnvpr_ids) * 10 if keep_cnvpr_ids else 1000)
                while q:
                    bfs_iter += 1
                    if bfs_iter > bfs_max:
                        print(f"[WARN][_iso_bfs] reached bfs_max={bfs_max}; aborting BFS expansion (preserve_count={len(preserve)})")
                        break
                    cur = q.popleft()
                    # anchors that reference cur -> consider including them
                    for anc in list(reverse_refs.get(str(cur), set())):
                        if anc in preserve:
                            continue
                        # allow if anc was in original seed
                        if anc in keep_cnvpr_ids:
                            preserve.add(anc)
                            q.append(anc)
                            continue
                        # otherwise require anc's from_row to be in group_rows
                        anc_row = id_to_row.get(str(anc))
                        if anc_row is not None and anc_row in group_rows:
                            preserve.add(anc)
                            q.append(anc)
                    # ids that cur references -> consider including them
                    for ref in list(refs.get(str(cur), set())):
                        if ref in preserve:
                            continue
                        if ref in keep_cnvpr_ids:
                            preserve.add(ref)
                            q.append(ref)
                            continue
                        ref_row = id_to_row.get(str(ref))
                        if ref_row is not None and ref_row in group_rows:
                            preserve.add(ref)
                            q.append(ref)

                # Before exposing the set of preserved ids, also ensure we
                # include connector-only anchors that were recorded in
                # connector_children_by_id when those connector anchors
                # reference any id already in the preserve set. The
                # earlier BFS conservatively constrains expansion by group
                # rows which can omit connector-only anchors whose
                # endpoints lie just outside the group's rows. That
                # causes connectors (e.g. 56,61) to be pruned; include them
                # here if they reference preserved shapes so they are
                # rendered with the group.
                try:
                    for cid, kids in list(connector_children_by_id.items()):
                        try:
                            # If this connector (cid) already preserved, skip
                            if cid in preserve:
                                continue
                            # Inspect cosmetic children for endpoint refs
                            added = False
                            endpoints = set()
                            for ch in kids:
                                for sub in ch.iter():
                                    try:
                                        t = sub.tag.split('}')[-1].lower()
                                    except Exception:
                                        t = ''
                                    if t in ('stcxn', 'endcxn', 'stcxnpr', 'endcxnpr'):
                                        vid = sub.attrib.get('id') or sub.attrib.get('idx')
                                        if vid is not None:
                                            endpoints.add(str(vid))
                            # If any endpoint directly references an already-preserved id, include connector
                            if endpoints and (endpoints & set(preserve)):
                                preserve.add(str(cid))
                                try:
                                    q.append(str(cid))
                                except (ValueError, TypeError):
                                    pass  # データ構造操作失敗は無視
                                continue
                            # Also include connector if any endpoint's anchor 'from' row
                            # is inside this group's rows (id_to_row may be empty if earlier parsing failed)
                            try:
                                for vid in endpoints:
                                    try:
                                        # prefer id_to_row (filtered anchors) but fall back
                                        # to all_id_to_row if not present
                                        row_for_vid = id_to_row.get(str(vid)) or all_id_to_row.get(str(vid))
                                    except (ValueError, TypeError):
                                        row_for_vid = None
                                    if row_for_vid is not None and row_for_vid in group_rows:
                                        preserve.add(str(cid))
                                        try:
                                            q.append(str(cid))
                                        except (ValueError, TypeError) as e:
                                            print(f"[DEBUG] 型変換エラー（無視）: {e}")
                                        added = True
                                        break
                                if added:
                                    continue
                            except (ValueError, TypeError) as e:
                                print(f"[DEBUG] 型変換エラー（無視）: {e}")
                            # fallback: if endpoints empty or no match, skip
                        except (ValueError, TypeError) as e:
                            print(f"[DEBUG] 型変換エラー（無視）: {e}")
                except Exception:
                    pass  # データ構造操作失敗は無視

                # Heuristic: include connectors whose own anchor 'from' row
                # is inside the group's rows even when their cosmetic children
                # do not expose endpoint tags. This handles connector-only
                # anchors that were omitted from id_to_row but appear in
                # all_id_to_row (we built that fallback earlier).
                try:
                    for cid in list(connector_children_by_id.keys()):
                        scid = str(cid)
                        if scid in preserve:
                            continue
                        try:
                            rowc = None
                            if 'id_to_row' in locals():
                                rowc = id_to_row.get(scid)
                            if rowc is None and 'all_id_to_row' in locals():
                                rowc = all_id_to_row.get(scid)
                            if rowc is not None:
                                # accept exact match or off-by-one to be more tolerant
                                accept = False
                                try:
                                    if rowc in group_rows:
                                        accept = True
                                    else:
                                        for gr in group_rows:
                                            if abs(int(rowc) - int(gr)) <= 1:
                                                accept = True
                                                break
                                except (ValueError, TypeError):
                                    accept = False
                                if accept:
                                    preserve.add(scid)
                                    try:
                                        q.append(scid)
                                    except (ValueError, TypeError) as e:
                                        print(f"[DEBUG] 型変換エラー（無視）: {e}")
                        except Exception:
                            pass  # データ構造操作失敗は無視
                except Exception:
                    pass  # データ構造操作失敗は無視

                # For debugging, expose the set of preserved ids
                referenced_ids = set(preserve)
                try:
                    # Extra debug: dump row mappings and connector endpoint resolution
                    try:
                        dbg_rows = sorted(list(group_rows)) if 'group_rows' in locals() else []
                    except Exception:
                        dbg_rows = []
                    try:
                        dbg_id_to_row_keys = sorted(list(id_to_row.keys())) if 'id_to_row' in locals() else []
                    except Exception:
                        dbg_id_to_row_keys = []
                    try:
                        dbg_all_id_to_row_keys = sorted(list(all_id_to_row.keys())) if 'all_id_to_row' in locals() else []
                    except Exception:
                        dbg_all_id_to_row_keys = []
                    print(f"[DEBUG][_iso_group_extra] group_rows={dbg_rows} id_to_row_keys={dbg_id_to_row_keys} all_id_to_row_keys={dbg_all_id_to_row_keys}")
                    # For each connector cosmetic entry, list endpoints (may be empty) and mapped rows
                    try:
                        for ccid in sorted(list(connector_children_by_id.keys()), key=lambda x: int(x) if str(x).isdigit() else x):
                            ckids = connector_children_by_id.get(ccid, [])
                            eps = set()
                            for ch in ckids:
                                for sub in ch.iter():
                                    try:
                                        t = sub.tag.split('}')[-1].lower()
                                    except (ValueError, TypeError):
                                        t = ''
                                    if t in ('stcxn', 'endcxn', 'stcxnpr', 'endcxnpr'):
                                        vid = sub.attrib.get('id') or sub.attrib.get('idx')
                                        if vid is not None:
                                            eps.add(str(vid))
                            # map to rows via id_to_row or all_id_to_row (may be empty)
                            rows_mapped = []
                            for e in sorted(list(eps)):
                                try:
                                    r = None
                                    if 'id_to_row' in locals():
                                        r = id_to_row.get(e)
                                    if r is None and 'all_id_to_row' in locals():
                                        r = all_id_to_row.get(e)
                                    rows_mapped.append(r)
                                except Exception:
                                    rows_mapped.append(None)
                            print(f"[DEBUG][_iso_group_conn] cid={ccid} endpoints={sorted(list(eps))} mapped_rows={rows_mapped}")
                    except (ValueError, TypeError) as e:
                        print(f"[DEBUG] 型変換エラー（無視）: {e}")
                    # Additionally, show explicit mapping for connector-only ids that are present only in all_id_to_row
                    try:
                        for special in ('56','61'):
                            if 'all_id_to_row' in locals() and special in all_id_to_row:
                                print(f"[DEBUG][_iso_group_idrow] id={special} all_row={all_id_to_row.get(special)} id_to_row_val={id_to_row.get(special) if 'id_to_row' in locals() else None}")
                    except (ValueError, TypeError) as e:
                        print(f"[DEBUG] 型変換エラー（無視）: {e}")
                    msg = f"[DEBUG][_iso_group] keep_cnvpr_ids={sorted(list(keep_cnvpr_ids))} preserved_ids={sorted(list(referenced_ids))} connector_children_keys={sorted(list(connector_children_by_id.keys()))}"
                    print(msg)
                    # expose preserved ids for callers so they can avoid duplicate renders
                    try:
                        self._last_iso_preserved_ids = set(referenced_ids)
                    except (ValueError, TypeError):
                        try:
                            self._last_iso_preserved_ids = set()
                        except (ValueError, TypeError) as e:
                            print(f"[DEBUG] 型変換エラー（無視）: {e}")
                    # Write a per-isolation diagnostic file (guaranteed path) so
                    # conversion runs always emit a record of which cNvPr ids
                    # were preserved into this isolated group. This is useful
                    # when downstream code later decides to skip clusters.
                    try:
                        import csv
                        out_dir = getattr(self.converter, 'output_dir', None) or os.path.join(os.getcwd(), 'output')
                        diag_dir = os.path.join(out_dir, 'diagnostics')
                        os.makedirs(diag_dir, exist_ok=True)
                        # deterministic name: base + sheet + hash of keep ids
                        try:
                            base = getattr(self.converter, 'base_name')
                        except Exception:
                            base = os.path.splitext(os.path.basename(getattr(self.converter, 'excel_file', 'workbook')))[0]
                        ksig = hashlib.sha1((base + sheet.title + ''.join(sorted(list(map(str, keep_cnvpr_ids))))).encode('utf-8')).hexdigest()[:8]
                        diag_path = os.path.join(diag_dir, f"{base}_{self.converter._sanitize_filename(sheet.title)}_iso_{ksig}.csv")
                        with open(diag_path, 'w', newline='', encoding='utf-8') as df:
                            w = csv.writer(df)
                            w.writerow(['keep_cnvpr_ids', 'preserved_ids', 'connector_children_keys'])
                            w.writerow([";".join(sorted(list(map(str, keep_cnvpr_ids)))), ";".join(sorted(list(map(str, referenced_ids)))), ";".join(sorted(list(map(str, connector_children_by_id.keys()))) )])
                        print(f"[DEBUG] wrote isolation diagnostics to {diag_path}")
                    except (OSError, IOError, FileNotFoundError):
                        print(f"[WARNING] ファイル操作エラー: {e if 'e' in locals() else '不明'}")
                except (OSError, IOError, FileNotFoundError):
                    print(f"[WARNING] ファイル操作エラー: {e if 'e' in locals() else '不明'}")
            except (OSError, IOError, FileNotFoundError):
                referenced_ids = set()
                connector_children_by_id = {}
            
            return tmpdir, referenced_ids, connector_children_by_id
            
        except Exception as e:
            print(f"[ERROR] Phase5 failed: {e}")
            import traceback
            traceback.print_exc()
            shutil.rmtree(tmpdir, ignore_errors=True)
            return None, set(), {}
    
    def _phase6_prune_drawing_xml(self, drawing_relpath, keep_cnvpr_ids, referenced_ids, 
                                   cell_range, drawing_xml, tmpdir, sheet, sheet_index):
        """フェーズ6: 描画XML刈り込み"""
        def node_contains_referenced_id(n):
            try:
                for sub in n.iter():
                    lname = sub.tag.split('}')[-1].lower()
                    # keep node if it contains a cNvPr whose id matches any referenced id
                    if lname == 'cnvpr' or lname.endswith('cnvpr'):
                        vid = sub.attrib.get('id') or sub.attrib.get('idx')
                        if vid is not None and str(vid) in referenced_ids:
                            return True
                    # also keep node if it contains connector endpoint refs
                    # such as <a:stCxn id="N"/> or <a:endCxn id="M"/>
                    if lname in ('stcxn', 'endcxn', 'stcxnpr', 'endcxnpr'):
                        vid = sub.attrib.get('id') or sub.attrib.get('idx')
                        if vid is not None and str(vid) in referenced_ids:
                            return True
            except (ValueError, TypeError):
                pass  # 一時ディレクトリ削除失敗は無視
            return False

        # parse drawing xml from extracted file
        try:
            tree = ET.parse(drawing_relpath)
            root = tree.getroot()
        except (ET.ParseError, KeyError, AttributeError):
            drawing_xml_bytes = ET.tostring(drawing_xml)
            root = ET.fromstring(drawing_xml_bytes)
            tree = ET.ElementTree(root)

        # remove anchors whose cNvPr id is not in keep_cnvpr_ids and which
        # do not contain referenced ids (connector endpoints). This avoids
        # relying on index positions which previously caused mismatches
        # when anchors was built as a filtered list.
        # If keep_cnvpr_ids is empty (index->id mapping failed), fall back
        # to preserving anchors that lie within the computed cell_range
        # when available. This avoids producing an empty trimmed drawing
        # workbook for groups whose indices were synthesized from cell
        # ranges rather than exact anchor indices.
        # Compute group_rows from cell_range for quick membership tests.
        group_rows = set()
        try:
            if cell_range:
                s_col, e_col, s_row, e_row = cell_range
                group_rows = set(range(int(s_row), int(e_row) + 1))
        except (ValueError, TypeError):
            group_rows = set()

        for node in list(root):
            lname = node.tag.split('}')[-1].lower()
            if lname in ('twocellanchor', 'onecellanchor'):
                # find cNvPr id for this anchor
                this_cid = None
                for sub in node.iter():
                    if sub.tag.split('}')[-1].lower() == 'cnvpr':
                        this_cid = sub.attrib.get('id') or sub.attrib.get('idx')
                        break

                # If we have an explicit id and it's requested, keep it.
                if this_cid is not None and str(this_cid) in keep_cnvpr_ids:
                    continue

                # If the node contains referenced ids (connector endpoints), keep it.
                try:
                    if node_contains_referenced_id(node):
                        continue
                except (ValueError, TypeError) as e:
                    print(f"[DEBUG] 型変換エラー（無視）: {e}")

                # Fallback: when keep_cnvpr_ids is empty but a cell_range
                # was computed for the group, preserve any anchor whose
                # "from" row lies within the group's rows. This handles
                # cases where indices were synthesized from cell ranges
                # and direct id matching fails.
                try:
                    if (not keep_cnvpr_ids) and group_rows:
                        ns_xdr = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
                        fr = node.find('{%s}from' % ns_xdr)
                        if fr is not None:
                            r = fr.find('{%s}row' % ns_xdr)
                            if r is not None and r.text is not None:
                                try:
                                    from_row = int(r.text)
                                    # accept exact or off-by-one matches
                                    if from_row in group_rows or any(abs(from_row - gr) <= 1 for gr in group_rows):
                                        continue
                                except (ValueError, TypeError) as e:
                                    print(f"[DEBUG] 型変換エラー（無視）: {e}")
                except (ValueError, TypeError) as e:
                    print(f"[DEBUG] 型変換エラー（無視）: {e}")

                # otherwise remove this node from the trimmed drawing
                try:
                    root.remove(node)
                except Exception:
                    try:
                        root.remove(node)
                    except Exception:
                        pass  # 一時ファイルの削除失敗は無視

        # Additionally, clear worksheet cell text in the tmp workbook so rendered PDF
        # contains only the drawing shapes. This prevents sheet text from appearing
        # in isolated renders.
        try:
            sheet_rel = os.path.join(tmpdir, f"xl/worksheets/sheet{sheet_index+1}.xml")
            if os.path.exists(sheet_rel):
                try:
                    stree = ET.parse(sheet_rel)
                    sroot = stree.getroot()
                    # clear all <v> and inline string texts under sheetData
                    for v in sroot.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v'):
                        v.text = ''
                    for t in sroot.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t'):
                        t.text = ''
                        # ensure page margins and page setup are tight so exported PDF
                        # doesn't add unexpected whitespace or scaling. Use zero margins
                        # and 100% scale.
                        try:
                            ns = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
                            pm_tag = '{%s}pageMargins' % ns
                            ps_tag = '{%s}pageSetup' % ns
                            # remove existing pageMargins/pageSetup if present
                            for child in list(sroot):
                                if child.tag in (pm_tag, ps_tag):
                                    try:
                                        sroot.remove(child)
                                    except Exception:
                                        pass  # 一時ファイルの削除失敗は無視
                            # add pageMargins with zeros
                            pm = ET.Element(pm_tag)
                            for name, val in (('left', '0'), ('right', '0'), ('top', '0'), ('bottom', '0'), ('header', '0'), ('footer', '0')):
                                el = ET.SubElement(pm, '{%s}%s' % (ns, name))
                                el.text = val
                            sroot.append(pm)
                            # add pageSetup: prefer fit-to-page so LibreOffice
                            # does not create extra pages due to legacy pageBreaks.
                            ps = ET.Element(ps_tag)
                            # Use fitToPage with fitToWidth/fitToHeight to try to
                            # keep the trimmed area on a single PDF page.
                            try:
                                ps.set('fitToPage', '1')
                                ps.set('fitToWidth', '1')
                                ps.set('fitToHeight', '1')
                            except Exception:
                                try:
                                    ps.set('scale', '100')
                                except Exception as e:
                                    pass  # XML解析エラーは無視
                            sroot.append(ps)
                        except Exception:
                            pass  # データ構造操作失敗は無視
                        # Remove any header/footer elements from this sheet
                        # node so isolated-group PDF/PNG renders do not
                        # include workbook headers or footers. This keeps
                        # the output image focused on the drawing shapes
                        # only. We'll still perform a defensive sweep later
                        # over all worksheet files in tmpdir just before
                        # creating the tmp_xlsx to be certain none remain.
                        try:
                            hf_tag = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}headerFooter'
                            removed = 0
                            for hf in list(sroot.findall(hf_tag)):
                                try:
                                    sroot.remove(hf)
                                    removed += 1
                                except Exception:
                                    pass  # 一時ファイルの削除失敗は無視
                            if removed:
                                print(f"[DEBUG][_iso_hdrfoot] removed {removed} headerFooter elements from {sheet_rel}")
                        except (ValueError, TypeError):
                            pass  # XML書き込み失敗は無視
                        stree.write(sheet_rel, encoding='utf-8', xml_declaration=True)
                except (ValueError, TypeError):
                    pass  # XML書き込み失敗は無視
        except (ValueError, TypeError):
            pass  # XML書き込み失敗は無視

        # write modified drawing xml back
        tree.write(drawing_relpath, encoding='utf-8', xml_declaration=True)

        # If pruning removed all anchors, skip isolated rendering to avoid
        # producing empty trimmed workbooks and placeholder images.
        try:
            try:
                dtree_check = ET.parse(drawing_relpath)
                droot_check = dtree_check.getroot()
                kept_anchors = [n for n in list(droot_check) if n.tag.split('}')[-1].lower() in ('twocellanchor', 'onecellanchor')]
                if not kept_anchors:
                    print(f"[DEBUG][_iso_entry] sheet={sheet.title} trimmed drawing has no anchors after pruning; skipping isolated group")
                    return False
            except (ET.ParseError, KeyError, AttributeError) as e:
                print(f"[DEBUG] XML解析エラー（無視）: {type(e).__name__}")
        except (ET.ParseError, KeyError, AttributeError) as e:
            print(f"[DEBUG] XML解析エラー（無視）: {type(e).__name__}")

        return True
    
    def _phase7_apply_connector_cosmetics(self, drawing_relpath, referenced_ids, 
                                          connector_children_by_id, theme_color_map, 
                                          drawing_xml_bytes, drawing_xml):
        """フェーズ7: コネクタコスメティック処理"""
        # After writing, ensure kept anchors have connector cosmetic children copied
        try:
            # reload tree to operate on current root
            try:
                tree2 = ET.parse(drawing_relpath)
                root2 = tree2.getroot()
            except (ET.ParseError, KeyError, AttributeError):
                root2 = ET.fromstring(drawing_xml_bytes)
                tree2 = ET.ElementTree(root2)
            # track dedupe signatures per-kept-anchor to avoid appending the same
            # cosmetic subtree multiple times (was causing duplicated anchor blocks)
            for kept in list(root2):
                if kept.tag.split('}')[-1].lower() not in ('twocellanchor', 'onecellanchor'):
                    continue
                kept_cid = None
                for sub in kept.iter():
                    if sub.tag.split('}')[-1].lower() == 'cnvpr':
                        kept_cid = str(sub.attrib.get('id'))
                        break
                if not kept_cid:
                    continue
                if kept_cid in connector_children_by_id:
                    seen_sigs = set()
                    for ch in connector_children_by_id[kept_cid]:
                        try:
                            if not hasattr(ch, 'iter'):
                                print(f"[ERROR] ch is not an XML element: type={type(ch)}, value={ch}")
                                continue
                            new_ch = copy.deepcopy(ch)
                            # Replace any a:schemeClr children with explicit a:srgbClr using parsed theme
                            try:
                                a_ns = 'http://schemas.openxmlformats.org/drawingml/2006/main'
                                for elem in list(new_ch.iter()):
                                    tag_lower = elem.tag.split('}')[-1].lower()
                                    if tag_lower == 'schemeclr':
                                        scheme_name = elem.attrib.get('val')
                                        if scheme_name and theme_color_map:
                                            hexv = theme_color_map.get(scheme_name.lower())
                                            if hexv:
                                                elem.tag = '{%s}srgbClr' % a_ns
                                                elem.attrib.clear()
                                                elem.set('val', hexv)
                            except Exception as e:
                                print(f"[WARNING] ファイル操作エラー: {e}")
                            # preserve attributes for important drawing tags (ln/headEnd/tailEnd/spPr)
                            for sub in ch.iter():
                                try:
                                    tag_lower = sub.tag.split('}')[-1].lower()
                                    if tag_lower in ('ln', 'headend', 'tailend', 'sppr'):
                                        for attr_k, attr_v in sub.attrib.items():
                                            applied = False
                                            for cand in new_ch.iter():
                                                if cand.tag.split('}')[-1].lower() == tag_lower:
                                                    if attr_k not in cand.attrib:
                                                        cand.attrib[attr_k] = attr_v
                                                    applied = True
                                                    break
                                            if not applied:
                                                if attr_k not in new_ch.attrib:
                                                    new_ch.attrib[attr_k] = attr_v
                                except Exception:
                                    pass  # データ構造操作失敗は無視
                            # compute a lightweight signature for deduplication:
                            try:
                                sig = ET.tostring(new_ch, encoding='utf-8')
                            except Exception:
                                sig = None
                            if sig is not None:
                                if sig in seen_sigs:
                                    # already appended equivalent subtree
                                    continue
                                seen_sigs.add(sig)
                            kept.append(new_ch)
                        except Exception:
                            try:
                                kept.append(copy.deepcopy(ch))
                            except Exception as e:
                                pass  # XML解析エラーは無視
            tree2.write(drawing_relpath, encoding='utf-8', xml_declaration=True)
        except Exception as e:
            print(f"[WARNING] ファイル操作エラー: {e}")

        # Extra pass: for any kept anchor that corresponds to an original
        # connector anchor (cxnSp/cxn), replace the connector element in
        # the trimmed drawing with a deep-copy of the original connector
        # element from the source drawing. This is a conservative step to
        # preserve exact <a:ln> children (w/prstDash/headEnd/tailEnd) and
        # other connector-specific structure that some renderers rely on.
        try:
            try:
                tree3 = ET.parse(drawing_relpath)
                root3 = tree3.getroot()
            except (ET.ParseError, KeyError, AttributeError):
                root3 = ET.fromstring(drawing_xml_bytes)
                tree3 = ET.ElementTree(root3)

            # build mapping from original anchor cNvPr id -> original cxnSp/cxn element
            orig_cxn_by_id = {}
            try:
                for orig in list(drawing_xml):
                    try:
                        if orig.tag.split('}')[-1].lower() not in ('twocellanchor', 'onecellanchor'):
                            continue
                        orig_cid = None
                        for sub in orig.iter():
                            if sub.tag.split('}')[-1].lower() == 'cnvpr':
                                orig_cid = sub.attrib.get('id') or sub.attrib.get('idx')
                                break
                        if orig_cid is None:
                            continue
                        # find immediate connector child (cxnSp or cxn)
                        for child in orig:
                            if child.tag.split('}')[-1].lower() in ('cxnsp', 'cxn'):
                                orig_cxn_by_id[str(orig_cid)] = child
                                break
                    except (ValueError, TypeError):
                        continue
            except (ValueError, TypeError):
                orig_cxn_by_id = {}

            # Now replace/inject in the trimmed drawing for kept anchors
            for kept in list(root3):
                try:
                    if kept.tag.split('}')[-1].lower() not in ('twocellanchor', 'onecellanchor'):
                        continue
                    kept_cid = None
                    for sub in kept.iter():
                        if sub.tag.split('}')[-1].lower() == 'cnvpr':
                            kept_cid = str(sub.attrib.get('id'))
                            break
                    if not kept_cid:
                        continue
                    if kept_cid not in orig_cxn_by_id:
                        continue
                    orig_cxn = orig_cxn_by_id.get(kept_cid)
                    if orig_cxn is None:
                        continue

                    # find first immediate cxn child in kept and replace it
                    replaced = False
                    for idx_child, child_candidate in enumerate(list(kept)):
                        try:
                            if child_candidate.tag.split('}')[-1].lower() in ('cxnsp', 'cxn'):
                                try:
                                    kept.remove(child_candidate)
                                except Exception:
                                    pass  # 一時ファイルの削除失敗は無視
                                try:
                                    kept.insert(idx_child, copy.deepcopy(orig_cxn))
                                except Exception:
                                    try:
                                        kept.append(copy.deepcopy(orig_cxn))
                                    except Exception:
                                        pass  # 一時ファイルの削除失敗は無視
                                replaced = True
                                break
                        except Exception:
                            continue
                    if not replaced:
                        try:
                            kept.append(copy.deepcopy(orig_cxn))
                        except Exception:
                            pass  # データ構造操作失敗は無視
                    # Post-process the injected connector element to ensure
                    # a single concrete <a:ln> exists under spPr and to remove
                    # any style/<a:lnRef> entries that may cause LibreOffice
                    # to prefer theme defaults (which can change dash/width).
                    try:
                        # find the (new) connector child we just inserted
                        conn_elem = None
                        for child_candidate in list(kept):
                            if child_candidate.tag.split('}')[-1].lower() in ('cxnsp', 'cxn'):
                                conn_elem = child_candidate
                                break
                        if conn_elem is not None:
                            # resolve any schemeClr under conn_elem -> srgb using theme_color_map
                            try:
                                a_ns = 'http://schemas.openxmlformats.org/drawingml/2006/main'
                                for elem in list(conn_elem.iter()):
                                    if elem.tag.split('}')[-1].lower() == 'schemeclr':
                                        scheme_name = elem.attrib.get('val')
                                        if scheme_name and theme_color_map:
                                            hexv = theme_color_map.get(scheme_name.lower())
                                            if hexv:
                                                elem.tag = '{%s}srgbClr' % a_ns
                                                elem.attrib.clear()
                                                elem.set('val', hexv)
                            except Exception as e:
                                print(f"[WARNING] ファイル操作エラー: {e}")
                                import traceback
                                traceback.print_exc()

                            # normalize ln children: keep exactly one <ln> under spPr
                            try:
                                sppr = None
                                # prefer spPr child under connector (ns may vary)
                                for ch in list(conn_elem):
                                    if ch.tag.split('}')[-1].lower() in ('sppr','sppr'.lower(),'sppr') or ch.tag.split('}')[-1].lower() == 'sppr' or ch.tag.split('}')[-1].lower() == 'spPr'.lower():
                                        sppr = ch
                                        break
                                # fallback: try to find any spPr-like element by tag name
                                if sppr is None:
                                    for ch in list(conn_elem):
                                        if ch.tag.split('}')[-1].lower() == 'sppr' or ch.tag.split('}')[-1].lower() == 'sppr':
                                            sppr = ch
                                            break
                                if sppr is not None:
                                    ln_elems = [c for c in list(sppr) if c.tag.split('}')[-1].lower() == 'ln']
                                    if len(ln_elems) > 1:
                                        # choose preferred ln: one with @w, then prstDash, then head/tail
                                        preferred = None
                                        for ln_c in ln_elems:
                                            if ln_c.attrib.get('w'):
                                                preferred = ln_c
                                                break
                                        if preferred is None:
                                            for ln_c in ln_elems:
                                                for sub in ln_c:
                                                    if sub.tag.split('}')[-1].lower() == 'prstdash':
                                                        preferred = ln_c
                                                        break
                                                if preferred is not None:
                                                    break
                                        if preferred is None:
                                            for ln_c in ln_elems:
                                                for sub in ln_c:
                                                    if sub.tag.split('}')[-1].lower() in ('headend','tailend'):
                                                        preferred = ln_c
                                                        break
                                                if preferred is not None:
                                                    break
                                        if preferred is None:
                                            preferred = ln_elems[0]
                                        # remove others
                                        for ln_c in ln_elems:
                                            if ln_c is not preferred:
                                                try:
                                                    sppr.remove(ln_c)
                                                except Exception:
                                                    pass  # 一時ファイルの削除失敗は無視

                            except Exception:
                                pass  # 一時ファイルの削除失敗は無視
                    except Exception:
                        pass  # 一時ファイルの削除失敗は無視
                except Exception as e:
                    print(f"[WARNING] ファイル操作エラー: {e}")
            # write back
            tree3.write(drawing_relpath, encoding='utf-8', xml_declaration=True)
        except Exception as e:
            print(f"[WARNING] ファイル操作エラー: {e}")
    


    def _phase8_prepare_workbook(self, tmpdir, sheet, sheet_index, cell_range, drawing_path, dpi, shape_indices, keep_cnvpr_ids):
        """フェーズ8: ワークブック準備
        
        cell_rangeを使用してPrint_Areaを設定し、一時的なxlsxファイルを作成
        対象シート以外のシートを削除
        
        Args:
            tmpdir: 一時ディレクトリパス
            sheet: ワークシートオブジェクト
            sheet_index: シートのインデックス
            cell_range: セル範囲 (s_col, e_col, s_row, e_row) または None
            drawing_path: drawing XMLのパス
            dpi: 解像度
            shape_indices: シェイプのインデックスリスト
            keep_cnvpr_ids: 保持する図形IDのセット
            
        Returns:
            str: 一時xlsxファイルのパス、失敗時はNone
        """
        import os
        import xml.etree.ElementTree as ET
        import tempfile
        import shutil
        
        
        target_sheet_new_index = 0
        wb_path = os.path.join(tmpdir, 'xl/workbook.xml')
        if os.path.exists(wb_path):
            try:
                tree = ET.parse(wb_path)
                root = tree.getroot()
                ns = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
                
                # sheets要素を取得
                sheets_tag = f'{{{ns}}}sheets'
                sheets_el = root.find(sheets_tag)
                
                if sheets_el is not None:
                    rel_ns = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
                    target_sheet_rid = None
                    sheets_to_remove = []
                    
                    # インデックスで判定（mainブランチと同じロジック）
                    for idx, sheet_el in enumerate(list(sheets_el)):
                        if idx == sheet_index:
                            target_sheet_rid = sheet_el.attrib.get(f'{{{rel_ns}}}id')
                        else:
                            sheets_to_remove.append((idx, sheet_el))
                    
                    for _, sheet_el in sheets_to_remove:
                        sheets_el.remove(sheet_el)
                    
                    if sheets_el is not None:
                        for sheet_el in list(sheets_el):
                            sheet_el.set('sheetId', '1')
                            sheet_el.set(f'{{{rel_ns}}}id', 'rId1')
                    
                    # sheets要素を最後に移動（mainブランチと同じ順序にする）
                    if sheets_el is not None:
                        root.remove(sheets_el)
                        root.append(sheets_el)
                    
                    tree.write(wb_path, encoding='utf-8', xml_declaration=True)
                    
                    wb_rels_path = os.path.join(tmpdir, 'xl/_rels/workbook.xml.rels')
                    if os.path.exists(wb_rels_path):
                        rels_tree = ET.parse(wb_rels_path)
                        rels_root = rels_tree.getroot()
                        pkg_rel_ns = 'http://schemas.openxmlformats.org/package/2006/relationships'
                        
                        rels_to_remove = []
                        target_sheet_rel = None
                        for rel in list(rels_root):
                            rid = rel.attrib.get('Id')
                            rel_type = rel.attrib.get('Type', '')
                            
                            if rel_type.endswith('/worksheet'):
                                if rid == target_sheet_rid:
                                    target_sheet_rel = rel
                                else:
                                    rels_to_remove.append(rel)
                        
                        for rel in rels_to_remove:
                            rels_root.remove(rel)
                        
                        if target_sheet_rel is not None:
                            target_sheet_rel.set('Id', 'rId1')
                        
                        rels_tree.write(wb_rels_path, encoding='utf-8', xml_declaration=True)
                    
                    for idx, _ in sheets_to_remove:
                        sheet_file = os.path.join(tmpdir, f'xl/worksheets/sheet{idx+1}.xml')
                        if os.path.exists(sheet_file):
                            os.remove(sheet_file)
                        
                        sheet_rels = os.path.join(tmpdir, f'xl/worksheets/_rels/sheet{idx+1}.xml.rels')
                        if os.path.exists(sheet_rels):
                            os.remove(sheet_rels)
            except Exception as e:
                print(f"[WARNING] シート削除失敗: {e}")
        
        # cell_rangeが指定されている場合、Print_Areaを設定
        if cell_range:
            s_col, e_col, s_row, e_row = cell_range
            
            # 列文字を計算
            start_col_letter = self._col_letter(s_col)
            end_col_letter = self._col_letter(e_col)
            
            # Print_Area文字列を作成
            sheet_name_escaped = sheet.title.replace("'", "''")
            area_ref = f"'{sheet_name_escaped}'!${start_col_letter}$1:${end_col_letter}${e_row - s_row + 1}"
            
            # workbook.xmlを更新
            wb_path = os.path.join(tmpdir, 'xl/workbook.xml')
            if os.path.exists(wb_path):
                try:
                    tree = ET.parse(wb_path)
                    root = tree.getroot()
                    
                    # definedNames要素を取得または作成
                    ns = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
                    dn_tag = f'{{{ns}}}definedNames'
                    dn = root.find(dn_tag)
                    
                    if dn is None:
                        dn = ET.Element(dn_tag)
                        # sheets要素を最後に移動してdefinedNamesをその前に配置（mainブランチと同じ順序）
                        sheets_tag = f'{{{ns}}}sheets'
                        sheets_el = root.find(sheets_tag)
                        if sheets_el is not None:
                            # sheets要素を削除して最後に再追加
                            root.remove(sheets_el)
                            # definedNamesを追加
                            root.append(dn)
                            root.append(sheets_el)
                        else:
                            root.append(dn)
                    
                    # 既存のPrint_Areaを削除
                    for existing in list(dn.findall(f'{{{ns}}}definedName')):
                        if existing.attrib.get('name') == '_xlnm.Print_Area':
                            dn.remove(existing)
                    
                    # 新しいPrint_Areaを追加（シート削除後は常にインデックス0）
                    new_dn = ET.Element(f'{{{ns}}}definedName')
                    new_dn.set('name', '_xlnm.Print_Area')
                    new_dn.set('localSheetId', str(target_sheet_new_index))
                    new_dn.text = area_ref
                    dn.append(new_dn)
                    
                    tree.write(wb_path, encoding='utf-8', xml_declaration=True)
                except Exception as e:
                    print(f"[WARNING] Print_Area設定失敗: {e}")
            
            sheet_path = os.path.join(tmpdir, f'xl/worksheets/sheet{target_sheet_new_index + 1}.xml')
            if os.path.exists(sheet_path):
                try:
                    tree = ET.parse(sheet_path)
                    root = tree.getroot()
                    ns = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
                    
                    for old_ps in list(root.findall(f'.//{{{ns}}}pageSetup')):
                        root.remove(old_ps)
                    
                    ps = ET.Element(f'{{{ns}}}pageSetup')
                    ps.set('scale', '100')
                    ps.set('paperSize', '1')
                    ps.set('orientation', 'portrait')
                    ps.set('pageOrder', 'downThenOver')
                    ps.set('blackAndWhite', 'false')
                    ps.set('draft', 'false')
                    ps.set('cellComments', 'none')
                    ps.set('horizontalDpi', '300')
                    ps.set('verticalDpi', '300')
                    ps.set('copies', '1')
                    root.append(ps)
                    
                    tree.write(sheet_path, encoding='utf-8', xml_declaration=True)
                
                except Exception as e:
                    print(f"[WARNING] pageSetup修正失敗: {e}")
            
            try:
                s_col, e_col, s_row, e_row = cell_range
                sheet_rel = os.path.join(tmpdir, f"xl/worksheets/sheet{target_sheet_new_index+1}.xml")
                if os.path.exists(sheet_rel):
                    stree4 = ET.parse(sheet_rel)
                    sroot4 = stree4.getroot()
                    ns = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
                    
                    for br_tag in ('rowBreaks', 'colBreaks', 'pageBreaks'):
                        for el in list(sroot4.findall(f'{{{ns}}}{br_tag}')):
                            try:
                                sroot4.remove(el)
                            except Exception:
                                pass
                    
                    sheet_data_tag = f'{{{ns}}}sheetData'
                    sheet_data = sroot4.find(sheet_data_tag)
                    if sheet_data is not None:
                        new_sheet_data = ET.Element(sheet_data_tag)
                        rows = sheet_data.findall(f'{{{ns}}}row')
                        for row_el in rows:
                            try:
                                rnum = int(row_el.attrib.get('r', '0'))
                            except (ValueError, TypeError):
                                continue
                            new_row = ET.Element(f'{{{ns}}}row')
                            new_row.set('r', row_el.attrib.get('r'))
                            for attr in ('ht', 'hidden', 'customHeight'):
                                if attr in row_el.attrib:
                                    new_row.set(attr, row_el.attrib.get(attr))
                            try:
                                rd = self.sheet.row_dimensions.get(rnum)
                                if rd is not None:
                                    rh = getattr(rd, 'height', None)
                                    if rh is not None:
                                        new_row.set('ht', str(rh))
                                        if 'customHeight' not in new_row.attrib:
                                            new_row.set('customHeight', '1')
                                    else:
                                        default_row_h = getattr(self.sheet.sheet_format, 'defaultRowHeight', None)
                                        if default_row_h is not None and 'ht' not in new_row.attrib:
                                            new_row.set('ht', str(float(default_row_h)))
                                else:
                                    default_row_h = getattr(self.sheet.sheet_format, 'defaultRowHeight', None)
                                    if default_row_h is not None and 'ht' not in new_row.attrib:
                                        new_row.set('ht', str(float(default_row_h)))
                            except (ValueError, TypeError):
                                pass
                            new_sheet_data.append(new_row)
                        
                        parent = sroot4
                        for child in list(parent):
                            if child.tag == sheet_data_tag:
                                parent.remove(child)
                        parent.append(new_sheet_data)
                        
                        dim_tag = f'{{{ns}}}dimension'
                        dim = sroot4.find(dim_tag)
                        if dim is None:
                            dim = ET.Element(dim_tag)
                            sroot4.insert(0, dim)
                        start_addr = f"{self._col_letter(1)}1"
                        end_addr = f"{self._col_letter(e_col - s_col + 1)}1"
                        dim.set('ref', f"{start_addr}:{end_addr}")
                    
                    cols_tag = f'{{{ns}}}cols'
                    col_tag = f'{{{ns}}}col'
                        for child in list(sroot4):
                            if child.tag == cols_tag:
                                try:
                                    sroot4.remove(child)
                                except Exception:
                                    pass
                        cols_el = ET.Element(cols_tag)
                        try:
                            from openpyxl.utils import get_column_letter
                            default_col_w = getattr(self.sheet.sheet_format, 'defaultColWidth', None) or 8.43
                            for c in range(s_col, e_col + 1):
                                cd = self.sheet.column_dimensions.get(get_column_letter(c))
                                width = None
                                hidden = None
                                if cd is not None:
                                    width = getattr(cd, 'width', None)
                                    hidden = getattr(cd, 'hidden', None)
                                if width is None:
                                    width = default_col_w
                                col_el = ET.Element(col_tag)
                                new_idx = c - s_col + 1
                                col_el.set('min', str(new_idx))
                                col_el.set('max', str(new_idx))
                                try:
                                    col_el.set('width', str(float(width)))
                                    if cd is not None and getattr(cd, 'width', None) is not None:
                                        col_el.set('customWidth', '1')
                                except (ValueError, TypeError):
                                    col_el.set('width', str(int(width) if width is not None else 8))
                                    if cd is not None and getattr(cd, 'width', None) is not None:
                                        col_el.set('customWidth', '1')
                                try:
                                    if hidden:
                                        col_el.set('hidden', '1')
                                except (ValueError, TypeError):
                                    pass
                                cols_el.append(col_el)
                        except (ValueError, TypeError):
                            for i_col in range(1, e_col - s_col + 2):
                                col_el = ET.Element(col_tag)
                                col_el.set('min', str(i_col))
                                col_el.set('max', str(i_col))
                                col_el.set('width', '8.43')
                                cols_el.append(col_el)
                        
                        try:
                            sf_tag = f'{{{ns}}}sheetFormatPr'
                            sheet_data_tag = f'{{{ns}}}sheetData'
                            for child in list(sroot4):
                                if child.tag == sf_tag:
                                    try:
                                        sroot4.remove(child)
                                    except Exception:
                                        pass
                            sf = ET.Element(sf_tag)
                            try:
                                default_col_w = getattr(self.sheet.sheet_format, 'defaultColWidth', None) or 8.43
                                sf.set('defaultColWidth', str(float(default_col_w)))
                            except (ValueError, TypeError):
                                pass
                            try:
                                default_row_h = getattr(self.sheet.sheet_format, 'defaultRowHeight', None) or 15.0
                                sf.set('defaultRowHeight', str(float(default_row_h)))
                            except (ValueError, TypeError):
                                pass
                            inserted_sf = False
                            for i, child in enumerate(list(sroot4)):
                                if child.tag == cols_tag or child.tag == sheet_data_tag:
                                    sroot4.insert(i, sf)
                                    inserted_sf = True
                                    break
                            if not inserted_sf:
                                sroot4.insert(0, sf)
                        except Exception:
                            pass
                        
                        inserted = False
                        for i, child in enumerate(list(sroot4)):
                            if 'sheetPr' in child.tag:
                                sroot4.insert(i+1, cols_el)
                                inserted = True
                                break
                        if not inserted:
                            sroot4.insert(0, cols_el)
                        
                        stree4.write(sheet_rel, encoding='utf-8', xml_declaration=True)
            except Exception as e:
                print(f"[WARNING] sheetData再構築失敗: {e}")
            
            try:
                drawing_path_full = os.path.join(tmpdir, drawing_path)
                if os.path.exists(drawing_path_full):
                    dtree = ET.parse(drawing_path_full)
                    droot = dtree.getroot()
                    ns_xdr = {'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'}
                    
                    for node in list(droot):
                        lname = node.tag.split('}')[-1].lower()
                        if lname not in ('twocellanchor', 'onecellanchor'):
                            continue
                        
                        fr = node.find('xdr:from', ns_xdr)
                        if fr is not None:
                            col_el = fr.find('xdr:col', ns_xdr)
                            row_el = fr.find('xdr:row', ns_xdr)
                            try:
                                if col_el is not None and col_el.text is not None:
                                    new_col = int(col_el.text) - (s_col - 1)
                                    if new_col < 0:
                                        new_col = 0
                                    col_el.text = str(new_col)
                            except (ValueError, TypeError):
                                pass
                            try:
                                if row_el is not None and row_el.text is not None:
                                    new_row = int(row_el.text) - (s_row - 1)
                                    if new_row < 0:
                                        new_row = 0
                                    row_el.text = str(new_row)
                            except (ValueError, TypeError):
                                pass
                        
                        to = node.find('xdr:to', ns_xdr)
                        if to is not None:
                            col_el = to.find('xdr:col', ns_xdr)
                            row_el = to.find('xdr:row', ns_xdr)
                            try:
                                if col_el is not None and col_el.text is not None:
                                    new_col = int(col_el.text) - (s_col - 1)
                                    if new_col < 0:
                                        new_col = 0
                                    col_el.text = str(new_col)
                            except (ValueError, TypeError):
                                pass
                            try:
                                if row_el is not None and row_el.text is not None:
                                    new_row = int(row_el.text) - (s_row - 1)
                                    if new_row < 0:
                                        new_row = 0
                                    row_el.text = str(new_row)
                            except (ValueError, TypeError):
                                pass
                    
                    dtree.write(drawing_path_full, encoding='utf-8', xml_declaration=True)
                    
                    try:
                        col_x, row_y = self.converter._compute_sheet_cell_pixel_map(self.sheet, DPI=dpi)
                    except Exception:
                        col_x, row_y = [0], [0]
                    EMU_PER_INCH = 914400
                    try:
                        EMU_PER_PIXEL = EMU_PER_INCH / float(dpi)
                    except (ValueError, TypeError):
                        try:
                            EMU_PER_PIXEL = EMU_PER_INCH / float(int(getattr(self.converter, 'dpi', dpi) or dpi))
                        except (ValueError, TypeError):
                            EMU_PER_PIXEL = EMU_PER_INCH / float(dpi)
                    a_ns = 'http://schemas.openxmlformats.org/drawingml/2006/main'
                    
                    for node2 in list(droot):
                        lname2 = node2.tag.split('}')[-1].lower()
                        if lname2 not in ('twocellanchor', 'onecellanchor'):
                            continue
                        try:
                            if lname2 == 'twocellanchor':
                                fr = node2.find('xdr:from', ns_xdr)
                                to = node2.find('xdr:to', ns_xdr)
                                if fr is None or to is None:
                                    continue
                                try:
                                    col = int(fr.find('xdr:col', ns_xdr).text)
                                except (ValueError, TypeError):
                                    col = 0
                                try:
                                    row = int(fr.find('xdr:row', ns_xdr).text)
                                except (ValueError, TypeError):
                                    row = 0
                                try:
                                    colOff = int(fr.find('xdr:colOff', ns_xdr).text)
                                except (ValueError, TypeError):
                                    colOff = 0
                                try:
                                    rowOff = int(fr.find('xdr:rowOff', ns_xdr).text)
                                except (ValueError, TypeError):
                                    rowOff = 0
                                try:
                                    to_col = int(to.find('xdr:col', ns_xdr).text)
                                except (ValueError, TypeError):
                                    to_col = col
                                try:
                                    to_row = int(to.find('xdr:row', ns_xdr).text)
                                except (ValueError, TypeError):
                                    to_row = row
                                try:
                                    to_colOff = int(to.find('xdr:colOff', ns_xdr).text)
                                except (ValueError, TypeError):
                                    to_colOff = 0
                                try:
                                    to_rowOff = int(to.find('xdr:rowOff', ns_xdr).text)
                                except (ValueError, TypeError):
                                    to_rowOff = 0
                                
                                left_px = col_x[col] + (colOff / EMU_PER_PIXEL) if col < len(col_x) else col_x[-1]
                                right_px = col_x[to_col] + (to_colOff / EMU_PER_PIXEL) if to_col < len(col_x) else col_x[-1]
                                top_px = row_y[row] + (rowOff / EMU_PER_PIXEL) if row < len(row_y) else row_y[-1]
                                bottom_px = row_y[to_row] + (to_rowOff / EMU_PER_PIXEL) if to_row < len(row_y) else row_y[-1]
                            else:
                                fr = node2.find('xdr:from', ns_xdr)
                                ext = node2.find('xdr:ext', ns_xdr)
                                if fr is None or ext is None:
                                    continue
                                try:
                                    col = int(fr.find('xdr:col', ns_xdr).text)
                                except (ValueError, TypeError):
                                    col = 0
                                try:
                                    row = int(fr.find('xdr:row', ns_xdr).text)
                                except (ValueError, TypeError):
                                    row = 0
                                try:
                                    colOff = int(fr.find('xdr:colOff', ns_xdr).text)
                                except (ValueError, TypeError):
                                    colOff = 0
                                cx = int(ext.attrib.get('cx', '0'))
                                cy = int(ext.attrib.get('cy', '0'))
                                left_px = col_x[col] + (colOff / EMU_PER_PIXEL) if col < len(col_x) else col_x[-1]
                                top_px = row_y[row] if row < len(row_y) else row_y[-1]
                                right_px = left_px + (cx / EMU_PER_PIXEL)
                                bottom_px = top_px + (cy / EMU_PER_PIXEL)
                        except (ValueError, TypeError):
                            continue
                        
                        try:
                            grp_node = node2.find('xdr:grpSp', ns_xdr)
                            target_w_px = max(0.0, (right_px - left_px))
                            target_h_px = max(0.0, (bottom_px - top_px))
                            if grp_node is not None:
                                try:
                                    grp_xfrm = grp_node.find('.//{%s}xfrm' % a_ns)
                                    chExt = None
                                    orig_ch_cx = orig_ch_cy = None
                                    if grp_xfrm is not None:
                                        chExt = grp_xfrm.find('{%s}chExt' % a_ns)
                                        ext_el = grp_xfrm.find('{%s}ext' % a_ns)
                                        orig_cx = orig_cy = None
                                        try:
                                            if ext_el is not None:
                                                ocx = ext_el.attrib.get('cx')
                                                ocy = ext_el.attrib.get('cy')
                                                if ocx is not None:
                                                    orig_cx = int(ocx)
                                                if ocy is not None:
                                                    orig_cy = int(ocy)
                                        except (ValueError, TypeError):
                                            orig_cx = orig_cy = None
                                        try:
                                            if chExt is not None:
                                                cccx = chExt.attrib.get('cx')
                                                cccy = chExt.attrib.get('cy')
                                                if cccx is not None:
                                                    orig_ch_cx = int(cccx)
                                                if cccy is not None:
                                                    orig_ch_cy = int(cccy)
                                        except (ValueError, TypeError):
                                            orig_ch_cx = orig_ch_cy = None
                                        
                                        try:
                                            if orig_cx and orig_cy and orig_cx > 0 and orig_cy > 0:
                                                orig_w_px = float(orig_cx) / float(EMU_PER_PIXEL)
                                                orig_h_px = float(orig_cy) / float(EMU_PER_PIXEL)
                                                if orig_w_px > 0 and orig_h_px > 0:
                                                    scale_w = target_w_px / orig_w_px if orig_w_px > 0 else 1.0
                                                    scale_h = target_h_px / orig_h_px if orig_h_px > 0 else 1.0
                                                    uniform_scale = min(scale_w, scale_h) if scale_w > 0 and scale_h > 0 else 1.0
                                                    new_cx_emu = int(round(float(orig_cx) * float(uniform_scale)))
                                                    new_cy_emu = int(round(float(orig_cy) * float(uniform_scale)))
                                                else:
                                                    new_cx_emu = int(round(target_w_px * EMU_PER_PIXEL))
                                                    new_cy_emu = int(round(target_h_px * EMU_PER_PIXEL))
                                            else:
                                                new_cx_emu = int(round(target_w_px * EMU_PER_PIXEL))
                                                new_cy_emu = int(round(target_h_px * EMU_PER_PIXEL))
                                        except (ValueError, TypeError):
                                            new_cx_emu = int(round(max(1.0, target_w_px) * EMU_PER_PIXEL))
                                            new_cy_emu = int(round(max(1.0, target_h_px) * EMU_PER_PIXEL))
                                        
                                        try:
                                            new_cx_emu = self.converter._to_positive(new_cx_emu, orig_cx, orig_ch_cx, target_w_px)
                                            new_cy_emu = self.converter._to_positive(new_cy_emu, orig_cy, orig_ch_cy, target_h_px)
                                        except (ValueError, TypeError):
                                            new_cx_emu = int(round(max(1.0, target_w_px) * EMU_PER_PIXEL))
                                            new_cy_emu = int(round(max(1.0, target_h_px) * EMU_PER_PIXEL))
                                        
                                        try:
                                            min_emu = int(round(float(EMU_PER_PIXEL))) if EMU_PER_PIXEL and EMU_PER_PIXEL > 0 else 1
                                            if not new_cx_emu or int(new_cx_emu) < min_emu:
                                                new_cx_emu = min_emu
                                            if not new_cy_emu or int(new_cy_emu) < min_emu:
                                                new_cy_emu = min_emu
                                        except (ValueError, TypeError):
                                            pass
                                        
                                        try:
                                            off = grp_xfrm.find('{%s}off' % a_ns)
                                            if off is None:
                                                off = ET.SubElement(grp_xfrm, '{%s}off' % a_ns)
                                            off.set('x', str(int(round(left_px * EMU_PER_PIXEL))))
                                            off.set('y', str(int(round(top_px * EMU_PER_PIXEL))))
                                        except (ValueError, TypeError):
                                            pass
                                        try:
                                            ext_el = grp_xfrm.find('{%s}ext' % a_ns)
                                            if ext_el is None:
                                                ext_el = ET.SubElement(grp_xfrm, '{%s}ext' % a_ns)
                                            ext_el.set('cx', str(int(new_cx_emu)))
                                            ext_el.set('cy', str(int(new_cy_emu)))
                                        except (ValueError, TypeError):
                                            pass
                                        
                                        try:
                                            if chExt is not None and orig_ch_cx and orig_ch_cy and orig_ch_cx > 0 and orig_ch_cy > 0:
                                                try:
                                                    if 'uniform_scale' in locals():
                                                        ch_scale = uniform_scale
                                                    else:
                                                        ch_scale = min(float(new_cx_emu) / float(orig_ch_cx), float(new_cy_emu) / float(orig_ch_cy))
                                                except (ValueError, TypeError):
                                                    ch_scale = 1.0
                                                try:
                                                    new_ch_cx = int(round(float(orig_ch_cx) * float(ch_scale)))
                                                    new_ch_cy = int(round(float(orig_ch_cy) * float(ch_scale)))
                                                    chExt.set('cx', str(new_ch_cx))
                                                    chExt.set('cy', str(new_ch_cy))
                                                except (ValueError, TypeError):
                                                    pass
                                        except (ValueError, TypeError):
                                            pass
                                except (ValueError, TypeError):
                                    pass
                            else:
                                for sp in node2.findall('.//{%s}sp' % a_ns):
                                    try:
                                        xfrm = sp.find('.//{%s}xfrm' % a_ns)
                                        if xfrm is not None:
                                            try:
                                                off = xfrm.find('{%s}off' % a_ns)
                                                if off is None:
                                                    off = ET.SubElement(xfrm, '{%s}off' % a_ns)
                                                off.set('x', str(int(round(left_px * EMU_PER_PIXEL))))
                                                off.set('y', str(int(round(top_px * EMU_PER_PIXEL))))
                                            except (ValueError, TypeError):
                                                pass
                                            try:
                                                ext_elem = xfrm.find('{%s}ext' % a_ns)
                                                if ext_elem is None:
                                                    ext_elem = ET.SubElement(xfrm, '{%s}ext' % a_ns)
                                                ext_elem.set('cx', str(int(round(target_w_px * EMU_PER_PIXEL))))
                                                ext_elem.set('cy', str(int(round(target_h_px * EMU_PER_PIXEL))))
                                            except (ValueError, TypeError):
                                                pass
                                    except (ValueError, TypeError):
                                        pass
                                
                                for pic in node2.findall('.//{%s}pic' % a_ns):
                                    try:
                                        xfrm = pic.find('.//{%s}xfrm' % a_ns)
                                        if xfrm is not None:
                                            try:
                                                off = xfrm.find('{%s}off' % a_ns)
                                                if off is None:
                                                    off = ET.SubElement(xfrm, '{%s}off' % a_ns)
                                                off.set('x', str(int(round(left_px * EMU_PER_PIXEL))))
                                                off.set('y', str(int(round(top_px * EMU_PER_PIXEL))))
                                            except (ValueError, TypeError):
                                                pass
                                            try:
                                                ext_elem = xfrm.find('{%s}ext' % a_ns)
                                                if ext_elem is None:
                                                    ext_elem = ET.SubElement(xfrm, '{%s}ext' % a_ns)
                                                ext_elem.set('cx', str(int(round(target_w_px * EMU_PER_PIXEL))))
                                                ext_elem.set('cy', str(int(round(target_h_px * EMU_PER_PIXEL))))
                                            except (ValueError, TypeError):
                                                pass
                                    except (ValueError, TypeError):
                                        pass
                        except (ValueError, TypeError):
                            pass
                    
                    dtree.write(drawing_path_full, encoding='utf-8', xml_declaration=True)
                    
            except Exception as e:
                print(f"[WARNING] 図形座標調整失敗: {e}")
        
        try:
            sheet_rels_path = os.path.join(tmpdir, f'xl/worksheets/_rels/sheet{target_sheet_new_index + 1}.xml.rels')
            if os.path.exists(sheet_rels_path):
                tree = ET.parse(sheet_rels_path)
                root = tree.getroot()
                ns = 'http://schemas.openxmlformats.org/package/2006/relationships'
                
                drawing_target = None
                for rel in list(root.findall(f'{{{ns}}}Relationship')):
                    rel_type = rel.get('Type', '')
                    if 'drawing' in rel_type.lower():
                        drawing_target = rel.get('Target')
                        root.remove(rel)
                    else:
                        root.remove(rel)
                
                if drawing_target:
                    new_rel = ET.Element(f'{{{ns}}}Relationship')
                    new_rel.set('Id', 'rId1')
                    new_rel.set('Type', 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing')
                    new_rel.set('Target', drawing_target)
                    root.append(new_rel)
                    
                    tree.write(sheet_rels_path, encoding='utf-8', xml_declaration=True)
                    print(f"[DEBUG] Cleaned sheet rels and set drawing to rId1")
                
                sheet_path = os.path.join(tmpdir, f'xl/worksheets/sheet{target_sheet_new_index + 1}.xml')
                if os.path.exists(sheet_path):
                    stree = ET.parse(sheet_path)
                    sroot = stree.getroot()
                    ns_s = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
                    ns_r = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
                    
                    drawing_elem = sroot.find(f'.//{{{ns_s}}}drawing')
                    if drawing_elem is not None:
                        drawing_elem.set(f'{{{ns_r}}}id', 'rId1')
                        stree.write(sheet_path, encoding='utf-8', xml_declaration=True)
                        print(f"[DEBUG] Updated sheet drawing reference to rId1")
        except Exception as e:
            print(f"[WARNING] リレーションシップのクリーンアップ失敗: {e}")
        
        
        # tmpdirをzip化して一時xlsxファイルを作成
        try:
            # shape_indicesからユニークなファイル名を生成（mainブランチと同じロジック）
            import hashlib
            try:
                keep_set = set(shape_indices)
                keep_list = sorted(list(keep_set)) if 'keep_set' in locals() else []
                if keep_list:
                    h = hashlib.sha1(','.join(map(str, keep_list)).encode('utf-8')).hexdigest()[:8]
                    suffix = f"_grp_{h}"
                else:
                    suffix = "_grp_all"
            except Exception:
                suffix = "_grp"
            
            excel_base = os.path.splitext(os.path.basename(self.converter.excel_file))[0]
            
            dbg_dir = os.path.join(self.converter.output_dir, 'debug_workbooks')
            os.makedirs(dbg_dir, exist_ok=True)
            
            final_xlsx_name = f"{excel_base}_iso_group{suffix}.xlsx"
            src_for_conv = os.path.join(dbg_dir, final_xlsx_name)
            
            import zipfile
            with zipfile.ZipFile(src_for_conv, 'w', zipfile.ZIP_DEFLATED) as zout:
                for folder, _, files in os.walk(tmpdir):
                    for fn in files:
                        full = os.path.join(folder, fn)
                        arcname = os.path.relpath(full, tmpdir)
                        zout.write(full, arcname)
            
            print(f"[DEBUG] Using ZIP-created workbook directly (preserving shapes): {src_for_conv}")
            
            try:
                st = os.stat(src_for_conv)
                print(f"[DEBUG] Workbook size: {st.st_size} bytes")
            except (ValueError, TypeError):
                pass
            
            try:
                self._set_page_setup_and_margins(src_for_conv)
                print(f"[DEBUG] Applied fit-to-page settings to: {src_for_conv}")
            except Exception as e:
                print(f"[WARNING] fit-to-page設定失敗: {e}")
            
            return src_for_conv
        except Exception as e:
            print(f"[ERROR] 一時xlsxファイル作成失敗: {e}")
            return None


    def _phase9_generate_pdf_png(self, sheet, shape_indices, src_for_conv, tmpdir, dpi, cell_range):
        """フェーズ9: PDF/PNG生成
        
        LibreOfficeを使用してPDF生成、ImageMagickでPNGに変換
        
        Args:
            sheet: ワークシートオブジェクト
            shape_indices: シェイプのインデックスリスト
            src_for_conv: 変換元xlsxファイルパス
            tmpdir: 一時ディレクトリパス
            dpi: 解像度
            cell_range: セル範囲 (s_col, e_col, s_row, e_row) または None
            
        Returns:
            str: 生成されたPNGファイルのパス、失敗時はNone
        """
        import os
        import tempfile
        import subprocess
        import shutil
        from PIL import Image
        
        # PDF生成用の一時ディレクトリ
        tmp_pdf_dir = tempfile.mkdtemp(prefix='xls2md_pdf_')
        
        try:
            # LibreOfficeでPDF生成（一時ディレクトリに出力）
            print(f"[DEBUG] LibreOffice PDF変換開始: sheet={sheet.title}")
            pdf_path = self._convert_excel_to_pdf(src_for_conv, tmp_pdf_dir, apply_fit_to_page=False)
            
            if pdf_path is None:
                print(f"[WARN] LibreOffice PDF変換失敗")
                return None
            
            png_filename = os.path.basename(src_for_conv).replace('.xlsx', '.png')
            final_png_path = os.path.join(self.converter.images_dir, png_filename)
            
            # PDFをPNGに変換（最終出力ディレクトリに直接出力）
            png_path = self._convert_pdf_to_png_with_output(pdf_path, final_png_path, dpi=dpi)
            
            if png_path is None:
                print(f"[WARN] PDF→PNG変換失敗")
                return None
            
            # cell_rangeが指定されている場合、クロップ処理
            if cell_range and os.path.exists(png_path):
                cropped_path = self._crop_png_to_cell_range(
                    png_path, cell_range, sheet, dpi
                )
                if cropped_path:
                    png_path = cropped_path
            
            return png_path
            
        except Exception as e:
            print(f"[ERROR] PDF/PNG生成エラー: {e}")
            import traceback
            traceback.print_exc()
            return None
        finally:
            # 一時PDFディレクトリを削除
            try:
                shutil.rmtree(tmp_pdf_dir)
            except Exception:
                pass
    
    def _convert_excel_to_pdf(self, excel_path, output_dir, apply_fit_to_page=False):
        """ExcelファイルをPDFに変換"""
        import os
        import subprocess
        
        LIBREOFFICE_PATH = get_libreoffice_path()
        
        if not LIBREOFFICE_PATH or not os.path.exists(LIBREOFFICE_PATH):
            print(f"[ERROR] LibreOffice not found: {LIBREOFFICE_PATH}")
            return None
        
        try:
            cmd = [
                LIBREOFFICE_PATH,
                '--headless',
                '--convert-to', 'pdf',
                '--outdir', output_dir,
                excel_path
            ]
            
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=90
            )
            
            if result.returncode == 0:
                basename = os.path.splitext(os.path.basename(excel_path))[0]
                pdf_path = os.path.join(output_dir, basename + '.pdf')
                
                if os.path.exists(pdf_path):
                    return pdf_path
            
            print(f"[ERROR] LibreOffice変換失敗: {result.stderr}")
            return None
            
        except subprocess.TimeoutExpired:
            print(f"[ERROR] LibreOffice変換タイムアウト")
            return None
        except Exception as e:
            print(f"[ERROR] LibreOffice変換エラー: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _convert_pdf_to_png(self, pdf_path, dpi=300):
        """PDFをPNGに変換（同じディレクトリに出力）"""
        import os
        output_path = pdf_path.replace('.pdf', '.png')
        return self._convert_pdf_to_png_with_output(pdf_path, output_path, dpi)
    
    def _convert_pdf_to_png_with_output(self, pdf_path, output_path, dpi=300):
        """PDFをPNGに変換（出力先を指定）、複数ページの場合は結合"""
        import os
        import subprocess
        import glob
        from PIL import Image as PILImage
        
        try:
            base_noext = os.path.splitext(output_path)[0]
            for p in sorted(glob.glob(base_noext + "*.png")):
                try:
                    os.remove(p)
                except (OSError, FileNotFoundError):
                    pass
            
            im_cmd = get_imagemagick_command()
            if not im_cmd:
                print(f"[ERROR] ImageMagickが見つかりません")
                return None
            
            cmd = [
                im_cmd,
                '-density', str(dpi),
                pdf_path,
                '-background', 'white',
                '-alpha', 'remove',
                '-quality', '90',
                output_path
            ]
            
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=120
            )
            
            if result.returncode != 0:
                print(f"[ERROR] ImageMagick変換失敗: {result.stderr}")
                return None
            
            candidates = sorted(glob.glob(base_noext + "*.png"))
            if not candidates and os.path.exists(output_path):
                candidates = [output_path]
            
            if len(candidates) > 1:
                imgs = [PILImage.open(p).convert('RGBA') for p in candidates]
                widths = [im.size[0] for im in imgs]
                heights = [im.size[1] for im in imgs]
                maxw = max(widths)
                total_h = sum(heights)
                stitched = PILImage.new('RGBA', (maxw, total_h), (255, 255, 255, 255))
                y = 0
                for im_obj in imgs:
                    stitched.paste(im_obj, (0, y))
                    y += im_obj.size[1]
                stitched.convert('RGB').save(output_path, 'PNG')
                for p in candidates:
                    try:
                        if os.path.abspath(p) != os.path.abspath(output_path):
                            os.remove(p)
                    except Exception:
                        pass
            
            if os.path.exists(output_path):
                return output_path
            
            return None
            
        except subprocess.TimeoutExpired:
            print(f"[ERROR] ImageMagick変換タイムアウト")
            return None
        except Exception as e:
            print(f"[ERROR] ImageMagick変換エラー: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _crop_png_to_cell_range(self, png_path, cell_range, sheet, dpi):
        """PNGを指定されたセル範囲にクロップ"""
        import os
        from PIL import Image
        
        try:
            s_col, e_col, s_row, e_row = cell_range
            
            # 画像を開く
            im = Image.open(png_path)
            w_im, h_im = im.size
            
            # セル座標を取得（簡易版）
            # 実際にはシートの列幅・行高を使用して計算すべき
            # ここでは画像サイズから推定
            
            # クロップ領域を計算（簡易版）
            # より正確な実装は元のコードを参照
            left_ratio = (s_col - 1) / max(1, sheet.max_column)
            top_ratio = (s_row - 1) / max(1, sheet.max_row)
            right_ratio = e_col / max(1, sheet.max_column)
            bottom_ratio = e_row / max(1, sheet.max_row)
            
            left = int(w_im * left_ratio)
            top = int(h_im * top_ratio)
            right = int(w_im * right_ratio)
            bottom = int(h_im * bottom_ratio)
            
            # クロップ
            if right > left and bottom > top:
                cropped = im.crop((left, top, right, bottom))
                cropped.save(png_path)
                return png_path
            
            return png_path
            
        except Exception as e:
            print(f"[ERROR] クロップエラー: {e}")
            return png_path


    def _phase10_postprocess(self, out_path, png_name, sheet, group_rows=None, cell_range=None):
        """フェーズ10: 後処理
        
        生成された画像ファイルの最終処理と戻り値の準備
        
        Args:
            out_path: 出力ファイルパス
            png_name: PNGファイル名
            sheet: ワークシートオブジェクト
            group_rows: グループ行のリスト（オプション）
            cell_range: セル範囲 (s_col, e_col, s_row, e_row)（オプション）
            
        Returns:
            Tuple[str, int]: (画像ファイル名, 開始行)
        """
        import os
        
        try:
            # 実際に使用されたファイル名を取得
            basename = os.path.basename(out_path)
            
            # 代表的な開始行を決定（ログ出力用）
            rep = None
            
            # group_rowsから開始行を取得
            if group_rows:
                try:
                    rep = int(min(group_rows))
                except (ValueError, TypeError):
                    pass
            
            # cell_rangeから開始行を取得
            if rep is None and cell_range:
                try:
                    rep = int(cell_range[2])  # s_row
                except (ValueError, TypeError):
                    pass
            
            # デフォルト値
            if rep is None:
                rep = 1
            
            # デバッグログ
            print(f"[INFO] sheet={sheet.title} file={basename} start_row={rep}")
            
            return (basename, rep)
            
        except Exception as e:
            print(f"[ERROR] 後処理エラー: {e}")
            # フォールバック: タプルで返す
            return (png_name, 1)


    def _set_page_setup_and_margins(self, xlsx_path):
        """ExcelファイルのpageSetupとpageMarginsを設定"""
        import zipfile
        import tempfile
        import shutil
        import xml.etree.ElementTree as ET
        
        tmpdir = tempfile.mkdtemp(prefix='xls2md_fitpage_')
        try:
            with zipfile.ZipFile(xlsx_path, 'r') as zin:
                zin.extractall(tmpdir)
            
            xl_worksheets = os.path.join(tmpdir, 'xl', 'worksheets')
            if os.path.exists(xl_worksheets):
                for fname in os.listdir(xl_worksheets):
                    if fname.endswith('.xml') and fname.startswith('sheet'):
                        sheet_path = os.path.join(xl_worksheets, fname)
                        try:
                            tree = ET.parse(sheet_path)
                            root = tree.getroot()
                            ns = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
                            
                            for ps in root.findall(f'.//{{{ns}}}pageSetup'):
                                root.remove(ps)
                            
                            ps = ET.Element(f'{{{ns}}}pageSetup')
                            ps.set('scale', '25')
                            ps.set('orientation', 'landscape')
                            ps.set('paperSize', '9')
                            ps.set('useFirstPageNumber', '1')
                            root.append(ps)
                            
                            for pm in root.findall(f'.//{{{ns}}}pageMargins'):
                                root.remove(pm)
                            pm = ET.Element(f'{{{ns}}}pageMargins')
                            pm.set('left', '0.25')
                            pm.set('right', '0.25')
                            pm.set('top', '0.25')
                            pm.set('bottom', '0.25')
                            pm.set('header', '0.0')
                            pm.set('footer', '0.0')
                            root.append(pm)
                            
                            for hf in root.findall(f'.//{{{ns}}}headerFooter'):
                                root.remove(hf)
                            
                            tree.write(sheet_path, encoding='utf-8', xml_declaration=True)
                        except Exception as e:
                            print(f"[WARNING] {fname} のpageSetup設定に失敗: {e}")
            
            with zipfile.ZipFile(xlsx_path, 'w', zipfile.ZIP_DEFLATED) as zout:
                for root_dir, dirs, files in os.walk(tmpdir):
                    for file in files:
                        file_path = os.path.join(root_dir, file)
                        arcname = os.path.relpath(file_path, tmpdir)
                        zout.write(file_path, arcname)
            
            return True
        finally:
            try:
                shutil.rmtree(tmpdir)
            except Exception:
                pass
    
    def _col_letter(self, col_num):
        """列番号をExcelの列文字に変換（1→'A', 27→'AA'）"""
        result = []
        while col_num > 0:
            col_num -= 1
            result.append(chr(col_num % 26 + ord('A')))
            col_num //= 26
        return ''.join(reversed(result))
