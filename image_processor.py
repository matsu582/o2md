#!/usr/bin/env python3
"""
画像・PDF処理モジュール
Office文書からの画像抽出、変換、PDF処理機能を提供
"""

import os
import subprocess
import tempfile
import shutil
from pathlib import Path
from typing import Optional, Tuple, List
from PIL import Image

from utils import get_libreoffice_path, get_imagemagick_command

LIBREOFFICE_PATH = get_libreoffice_path()
IMAGEMAGICK_CMD = get_imagemagick_command()


def convert_document_to_pdf(docx_path: str) -> Optional[str]:
    """
    Word文書をPDFに変換
    
    Args:
        docx_path: 変換元のWord文書パス
    
    Returns:
        変換されたPDFファイルのパス。失敗時はNone
    """
    try:
        temp_dir = tempfile.mkdtemp()
        
        cmd = [
            LIBREOFFICE_PATH,
            '--headless',
            '--convert-to', 'pdf',
            '--outdir', temp_dir,
            docx_path
        ]
        
        env = os.environ.copy()
        env['SAL_DISABLE_OPENCL'] = '1'
        
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=60, env=env)
        
        if result.returncode == 0:
            for file in os.listdir(temp_dir):
                if file.endswith('.pdf'):
                    pdf_path = os.path.join(temp_dir, file)
                    final_pdf_path = tempfile.mktemp(suffix='.pdf')
                    shutil.copy2(pdf_path, final_pdf_path)
                    shutil.rmtree(temp_dir)
                    print(f"[INFO] PDFに変換完了: {final_pdf_path}")
                    return final_pdf_path
        
        shutil.rmtree(temp_dir)
        print(f"[ERROR] PDF変換失敗: {result.stderr}")
        return None
        
    except Exception as e:
        print(f"[ERROR] PDF変換エラー: {e}")
        return None


def convert_pdf_to_png(pdf_path: str, output_path: str) -> bool:
    """
    PDFをPNGに変換
    
    Args:
        pdf_path: 変換元のPDFファイルパス
        output_path: 出力先のPNGファイルパス
    
    Returns:
        変換成功時True、失敗時False
    """
    try:
        print("[DEBUG] 高速PDF→PNG変換実行...")
        cmd_fast = [
            IMAGEMAGICK_CMD,
            '-density', '300',
            f'{pdf_path}[0]',
            '-colorspace', 'RGB',
            '-background', 'white',
            '-alpha', 'remove',
            '-resize', '200%',
            '-trim',
            '+repage',
            '-quality', '90',
            '-depth', '8',
            output_path
        ]
        
        result = subprocess.run(cmd_fast, capture_output=True, text=True, timeout=30)
        
        if result.returncode == 0 and os.path.exists(output_path):
            print(f"[INFO] 高速PNG変換完了: {output_path}")
            return True
        
        print(f"[WARNING] 高速変換失敗、代替手法を試行: {result.stderr}")
        
        if shutil.which('pdftoppm'):
            print("[DEBUG] pdftoppm高速変換試行...")
            cmd_ppm = ['pdftoppm', '-png', '-r', '200', '-singlefile', 
                      pdf_path, output_path.replace('.png', '')]
            
            result2 = subprocess.run(cmd_ppm, capture_output=True, text=True, timeout=20)
            if result2.returncode == 0 and os.path.exists(output_path):
                print(f"[INFO] pdftoppm変換完了: {output_path}")
                cmd_trim = [IMAGEMAGICK_CMD, output_path, '-trim', '+repage', output_path]
                subprocess.run(cmd_trim, capture_output=True, text=True, timeout=10)
                return True
        
        print("[DEBUG] 最小設定変換試行...")
        cmd_minimal = [
            IMAGEMAGICK_CMD,
            '-density', '150',
            f'{pdf_path}[0]',
            '-resize', '150%',
            '-trim',
            '+repage',
            output_path
        ]
        
        result3 = subprocess.run(cmd_minimal, capture_output=True, text=True, timeout=15)
        if result3.returncode == 0 and os.path.exists(output_path):
            print(f"[INFO] 最小設定変換完了: {output_path}")
            return True
        
        print(f"[ERROR] すべてのPNG変換手法が失敗しました")
        return False
        
    except Exception as e:
        print(f"[ERROR] PNG変換エラー: {e}")
        return False


def convert_excel_to_pdf(xlsx_path: str, tmpdir: str, apply_fit_to_page: bool = True) -> Optional[str]:
    """
    ExcelファイルをPDFに変換
    
    Args:
        xlsx_path: 変換元のExcelファイルパス
        tmpdir: 一時ディレクトリ
        apply_fit_to_page: 1ページに収める設定を適用するか
    
    Returns:
        変換されたPDFファイルのパス。失敗時はNone
    """
    try:
        cmd = [
            LIBREOFFICE_PATH,
            '--headless',
            '--convert-to', 'pdf',
            '--outdir', tmpdir,
            xlsx_path
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
        
        if result.returncode == 0:
            for file in os.listdir(tmpdir):
                if file.endswith('.pdf'):
                    return os.path.join(tmpdir, file)
        
        print(f"[ERROR] Excel→PDF変換失敗: {result.stderr}")
        return None
        
    except Exception as e:
        print(f"[ERROR] Excel→PDF変換エラー: {e}")
        return None


def convert_pdf_page_to_png(pdf_path: str, page_index: int, dpi: int,
                            output_path: str, trim: bool = True) -> bool:
    """
    PDFの特定ページをPNGに変換
    
    Args:
        pdf_path: PDFファイルパス
        page_index: ページ番号（0始まり）
        dpi: 解像度
        output_path: 出力先パス
        trim: 余白を除去するか
    
    Returns:
        変換成功時True、失敗時False
    """
    try:
        cmd = [
            IMAGEMAGICK_CMD,
            '-density', str(dpi),
            f'{pdf_path}[{page_index}]',
            '-quality', '100',
            '-background', 'white',
            '-alpha', 'remove'
        ]
        
        if trim:
            cmd.extend(['-trim', '+repage'])
        
        cmd.append(output_path)
        
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
        
        if result.returncode == 0 and os.path.exists(output_path):
            print(f"[INFO] PDF→PNG変換完了: {output_path}")
            return True
        
        print(f"[ERROR] PDF→PNG変換失敗: {result.stderr}")
        return False
        
    except Exception as e:
        print(f"[ERROR] PDF→PNG変換エラー: {e}")
        return False


def convert_vector_image(image_data: bytes, original_path: str) -> Optional[str]:
    """
    ベクター画像（EMF/WMF）をPNGに変換
    
    Args:
        image_data: 画像データ
        original_path: 元のファイルパス
    
    Returns:
        変換されたPNGファイルのパス。失敗時は元のパスまたはNone
    """
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=Path(original_path).suffix) as temp_file:
            temp_file.write(image_data)
            temp_path = temp_file.name
        
        output_path = original_path.replace('.emf', '.png').replace('.wmf', '.png')
        
        if convert_with_libreoffice(temp_path, output_path):
            os.unlink(temp_path)
            return output_path
        
        cmd = [
            IMAGEMAGICK_CMD,
            temp_path,
            '-density', '300',
            '-quality', '100',
            '-background', 'white',
            '-alpha', 'remove',
            output_path
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
        os.unlink(temp_path)
        
        if result.returncode == 0 and os.path.exists(output_path):
            print(f"[SUCCESS] ベクター画像変換完了（ImageMagick）: {output_path}")
            return output_path
        else:
            print(f"[ERROR] ベクター画像変換失敗: {result.stderr}")
            with open(original_path, 'wb') as f:
                f.write(image_data)
            return original_path
            
    except Exception as e:
        print(f"[ERROR] ベクター画像変換エラー: {e}")
        with open(original_path, 'wb') as f:
            f.write(image_data)
        return original_path


def convert_with_libreoffice(input_path: str, output_path: str) -> bool:
    """
    LibreOfficeを使用してベクター画像をPNGに変換
    
    Args:
        input_path: 入力ファイルパス
        output_path: 出力ファイルパス
    
    Returns:
        変換成功時True、失敗時False
    """
    try:
        temp_dir = tempfile.mkdtemp()
        
        cmd = [
            LIBREOFFICE_PATH,
            '--headless',
            '--convert-to', 'pdf',
            '--outdir', temp_dir,
            input_path
        ]
        
        subprocess.run(cmd, capture_output=True, text=True, timeout=30)
        
        pdf_path = None
        for file in os.listdir(temp_dir):
            if file.endswith('.pdf'):
                pdf_path = os.path.join(temp_dir, file)
                break
        
        if pdf_path and os.path.exists(pdf_path):
            cmd2 = [
                IMAGEMAGICK_CMD,
                pdf_path,
                '-density', '300',
                '-quality', '100',
                '-background', 'white',
                '-alpha', 'remove',
                '-trim',
                '+repage',
                output_path
            ]
            
            result2 = subprocess.run(cmd2, capture_output=True, text=True, timeout=30)
            shutil.rmtree(temp_dir)
            
            if result2.returncode == 0 and os.path.exists(output_path):
                print(f"[SUCCESS] ベクター画像変換完了（LibreOffice→PDF→PNG）: {output_path}")
                return True
        
        shutil.rmtree(temp_dir)
        return False
        
    except Exception as e:
        print(f"[ERROR] LibreOffice変換エラー: {e}")
        return False


def compute_sheet_cell_pixel_map(sheet, DPI: int = 300):
    """
    Excelシートのセルとピクセルのマッピングを計算
    
    元の実装と互換性を保つため、list形式で返します。
    col_x[0] == 0, col_x[i]は列i(1-based)の右端のピクセル位置
    
    Args:
        sheet: openpyxlのワークシート
        DPI: 解像度
    
    Returns:
        (列のピクセル位置list, 行のピクセル位置list)
    """
    try:
        from openpyxl.utils import get_column_letter
        import math
        
        max_col = sheet.max_column
        max_row = sheet.max_row
        col_pixels = []
        
        for c in range(1, max_col+1):
            cd = sheet.column_dimensions.get(get_column_letter(c))
            width = getattr(cd, 'width', None) if cd is not None else None
            if width is None:
                try:
                    from openpyxl.utils import units as _units
                    width = getattr(sheet.sheet_format, 'defaultColWidth', None) or _units.DEFAULT_COLUMN_WIDTH
                except Exception:
                    width = 8.43
            try:
                MAX_DIGIT_WIDTH = 7
                base_px = int(math.floor(((256.0 * float(width) + math.floor(128.0 / MAX_DIGIT_WIDTH)) / 256.0) * MAX_DIGIT_WIDTH))
                if base_px < 1:
                    base_px = 1
                scale = float(DPI) / 96.0 if DPI and DPI > 0 else 1.0
                px = max(1, int(round(base_px * scale)))
            except (ValueError, TypeError):
                try:
                    base = max(1, int(float(width) * 7 + 5))
                    px = max(1, int(round(base * (float(DPI) / 96.0))))
                except (ValueError, TypeError):
                    px = max(1, int(float(width) * 7 + 5))
            col_pixels.append(px)

        row_pixels = []
        for r in range(1, max_row+1):
            rd = sheet.row_dimensions.get(r)
            hpts = getattr(rd, 'height', None) if rd is not None else None
            if hpts is None:
                try:
                    from openpyxl.utils import units as _units
                    hpts = _units.DEFAULT_ROW_HEIGHT
                except Exception:
                    hpts = 15
            try:
                px = max(1, int(float(hpts) * DPI / 72.0))
            except (ValueError, TypeError):
                px = max(1, int(hpts * DPI / 72))
            row_pixels.append(px)

        col_x = [0]
        for v in col_pixels:
            col_x.append(col_x[-1] + v)
        row_y = [0]
        for v in row_pixels:
            row_y.append(row_y[-1] + v)

        return col_x, row_y
    except Exception:
        return [0], [0]


def snap_box_to_cell_bounds(box: Tuple[int, int, int, int], 
                            col_x: List[int], row_y: List[int], DPI: int = 300) -> Tuple[int, int, int, int]:
    """
    バウンディングボックスをセルの境界にスナップ
    
    元の実装と互換性を保つため、list形式のcol_x/row_yを期待します。
    
    Args:
        box: (left, top, right, bottom)のピクセル座標
        col_x: 列のピクセル位置リスト（インデックスベース）
        row_y: 行のピクセル位置リスト（インデックスベース）
        DPI: 解像度
    
    Returns:
        スナップされた(left, top, right, bottom)
    """
    try:
        l, t, r, btm = box
        try:
            tol = max(1, int(DPI / 300.0 * 3))
        except (ValueError, TypeError):
            tol = 3
        
        start_col = None
        for c in range(1, len(col_x)):
            if col_x[c] >= l - tol:
                start_col = c
                break
        if start_col is None:
            start_col = max(1, len(col_x)-1)
        
        end_col = None
        for c in range(1, len(col_x)):
            if col_x[c] >= r + tol:
                end_col = c
                break
        if end_col is None:
            end_col = max(1, len(col_x)-1)
        
        start_row = None
        for rw in range(1, len(row_y)):
            if row_y[rw] >= t - tol:
                start_row = rw
                break
        if start_row is None:
            start_row = max(1, len(row_y)-1)
        
        end_row = None
        for rw in range(1, len(row_y)):
            if row_y[rw] >= btm + tol:
                end_row = rw
                break
        if end_row is None:
            end_row = max(1, len(row_y)-1)
        
        if start_col > 0:
            snap_l = col_x[start_col - 1]
        else:
            snap_l = l
        
        snap_r = col_x[end_col] if end_col < len(col_x) else r
        
        if start_row > 0:
            snap_t = row_y[start_row - 1]
        else:
            snap_t = t
        
        snap_b = row_y[end_row] if end_row < len(row_y) else btm
        
        return (snap_l, snap_t, snap_r, snap_b)
    except Exception:
        return box
